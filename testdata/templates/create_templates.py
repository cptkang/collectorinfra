"""엑셀 양식 템플릿 3종 생성 스크립트.

testdata/templates/ 폴더에 다음 3개의 엑셀 양식 파일을 생성한다:
  1. server_list_template.xlsx - 서버 인벤토리 목록
  2. resource_status_template.xlsx - 리소스별 상태 보고서
  3. config_report_template.xlsx - 서버 설정 정보 보고서
"""

import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _apply_header_style(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    headers: list[str],
    widths: list[float],
    fill_color: str,
) -> None:
    """헤더 행에 스타일을 적용한다.

    Args:
        ws: 워크시트 객체
        headers: 헤더 텍스트 리스트
        widths: 각 컬럼의 너비 리스트
        fill_color: 배경색 (hex, 예: 'DAEEF3')
    """
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin")
    header_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    for col_idx, (header_text, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width


def create_server_list_template(output_dir: Path) -> Path:
    """서버 인벤토리 목록 양식을 생성한다.

    컬럼: 호스트명, IP주소, OS종류, CPU코어수, 제조사, 상태

    Args:
        output_dir: 출력 디렉토리 경로

    Returns:
        생성된 파일 경로
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "서버목록"

    headers = ["호스트명", "IP주소", "OS종류", "CPU코어수", "제조사", "상태"]
    widths = [20, 18, 15, 12, 20, 10]
    fill_color = "DAEEF3"  # 연한 파란색

    _apply_header_style(ws, headers, widths, fill_color)

    filepath = output_dir / "server_list_template.xlsx"
    wb.save(str(filepath))
    return filepath


def create_resource_status_template(output_dir: Path) -> Path:
    """리소스별 상태 보고서 양식을 생성한다.

    컬럼: 서버명, 리소스유형, 리소스명, 상태, 설명

    Args:
        output_dir: 출력 디렉토리 경로

    Returns:
        생성된 파일 경로
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "리소스상태"

    headers = ["서버명", "리소스유형", "리소스명", "상태", "설명"]
    widths = [20, 25, 25, 10, 40]
    fill_color = "E2EFDA"  # 연한 녹색

    _apply_header_style(ws, headers, widths, fill_color)

    filepath = output_dir / "resource_status_template.xlsx"
    wb.save(str(filepath))
    return filepath


def create_config_report_template(output_dir: Path) -> Path:
    """서버 설정 정보 보고서 양식을 생성한다.

    컬럼: 서버명, 설정항목, 설정값

    Args:
        output_dir: 출력 디렉토리 경로

    Returns:
        생성된 파일 경로
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "설정정보"

    headers = ["서버명", "설정항목", "설정값"]
    widths = [20, 25, 50]
    fill_color = "FFF2CC"  # 연한 노란색

    _apply_header_style(ws, headers, widths, fill_color)

    filepath = output_dir / "config_report_template.xlsx"
    wb.save(str(filepath))
    return filepath


def main() -> None:
    """3종의 엑셀 양식 템플릿을 생성한다."""
    output_dir = Path(__file__).parent
    output_dir.mkdir(parents=True, exist_ok=True)

    templates = [
        ("server_list_template.xlsx", create_server_list_template),
        ("resource_status_template.xlsx", create_resource_status_template),
        ("config_report_template.xlsx", create_config_report_template),
    ]

    for name, creator in templates:
        filepath = creator(output_dir)
        print(f"[OK] {filepath}")

    print(f"\n총 {len(templates)}개 템플릿 생성 완료: {output_dir}")


if __name__ == "__main__":
    main()
