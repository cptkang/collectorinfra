#!/usr/bin/env python3
"""
Excel 데이터를 PostgreSQL INSERT SQL로 변환하여 실행하는 스크립트.

sample/CMM_RESOURCE.xlsx, sample/CORE_CONFIG_PROP.xlsx의 데이터를 읽어
polestar PostgreSQL DB에 INSERT한다.

핵심: CMM_RESOURCE.RESOURCE_CONF_ID = CORE_CONFIG_PROP.CONFIGURATION_ID
     JOIN 관계가 올바르게 설정되어야 한다.

Usage:
    python scripts/excel_to_pg_insert.py                # SQL 생성 + DB 실행
    python scripts/excel_to_pg_insert.py --sql-only      # SQL 파일만 생성
    python scripts/excel_to_pg_insert.py --with-extra     # Excel + 추가 샘플 데이터
"""

import argparse
import sys
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
SAMPLE_DIR = BASE_DIR / "sample"
OUTPUT_DIR = BASE_DIR / "testdata" / "pg"

# DB 테이블 컬럼 정의 (DDL 순서)
CMM_COLUMNS = [
    "dtype", "id", "acl_id", "acl_manager_group_id", "acl_manager_id",
    "avail_status", "ctime", "dtime", "description", "group_path",
    "haschildren", "hostname", "id_ancestry", "identifier", "importance_id",
    "is_inherit_avail_depend", "is_inherit_custom_conf", "is_inherit_manager_zone",
    "inheritstatus", "inventorypollinginterval", "invisible", "ipaddress",
    "lc", "location", "longpollinginterval", "mtime", "is_maintenance",
    "manager_zone", "mesurementpollinginterval", "modifiedby", "name",
    "optlock", "order_num", "parent_resource_id", "platform_resource_id",
    "pollingpolicy", "priority", "resourceicon", "resource_key",
    "resourcestatus", "resource_type", "resourcetypeversion",
    "service_resource_id", "is_sync_desc", "is_sync_name", "uuid",
    "version", "system", "avail_depend_resource_id", "avail_depend_resource_id_2",
    "connection_conf_id", "custom_conf_id", "realtime_info_id",
    "resource_conf_id", "resource_path_id", "schedule_id",
    "resource_system_id", "group_resource_id", "business_group_resource_id",
]

CONFIG_COLUMNS = [
    "dtype", "id", "errormessage", "name", "time_stamp", "stringvalue",
    "is_lob", "stringvalue_short", "configuration_id", "parent_list_id",
    "parent_map_id", "propertydefinition_id",
]

# 문자열 컬럼 (나머지는 숫자/NULL)
CMM_STRING_COLS = {
    "dtype", "description", "group_path", "hostname", "id_ancestry",
    "identifier", "ipaddress", "location", "manager_zone", "modifiedby",
    "name", "resourceicon", "resource_key", "resource_type",
    "resourcetypeversion", "uuid", "version",
}

CONFIG_STRING_COLS = {
    "dtype", "errormessage", "name", "stringvalue", "stringvalue_short",
}


def sql_value(val, is_string: bool) -> str:
    """Python 값을 SQL 값으로 변환."""
    if val is None or val == "" or val == "NULL":
        return "NULL"
    if is_string:
        s = str(val).replace("'", "''")
        return f"'{s}'"
    try:
        v = int(val)
        return str(v)
    except (ValueError, TypeError):
        try:
            v = float(val)
            return str(v)
        except (ValueError, TypeError):
            s = str(val).replace("'", "''")
            return f"'{s}'"


def read_excel_rows(filepath: Path, columns: list[str], string_cols: set[str]) -> list[str]:
    """Excel 파일을 읽어 INSERT SQL 문 목록을 반환."""
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb.active

    excel_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {}
    for i, h in enumerate(excel_headers):
        if h:
            header_map[h.lower()] = i

    table_name = "polestar.cmm_resource" if "cmm" in filepath.stem.lower() else "polestar.core_config_prop"

    insert_lines = []
    seen_ids = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_list = list(row)
        id_idx = header_map.get("id")
        if id_idx is None or row_list[id_idx] is None or row_list[id_idx] == "":
            continue

        row_id = row_list[id_idx]
        if row_id in seen_ids:
            continue
        seen_ids.add(row_id)

        values = []
        for col in columns:
            idx = header_map.get(col.lower())
            if idx is not None and idx < len(row_list):
                val = row_list[idx]
            else:
                val = None
            values.append(sql_value(val, col in string_cols))

        cols_str = ", ".join(columns)
        vals_str = ", ".join(values)
        insert_lines.append(f"INSERT INTO {table_name} ({cols_str}) VALUES ({vals_str});")

    wb.close()
    return insert_lines


def generate_fix_existing_data() -> list[str]:
    """기존 30대 서버의 RESOURCE_CONF_ID를 CONFIGURATION_ID와 매핑하는 UPDATE SQL."""
    updates = []
    updates.append("-- 기존 테스트 데이터: platform.server의 RESOURCE_CONF_ID 설정")

    # hostname → configuration_id 매핑 (CORE_CONFIG_PROP에서 Hostname 속성 기준)
    # svr-web-01~10 → 301~310, svr-was-01~10 → 311~320, svr-db-01~10 → 321~330
    servers = []
    for i in range(1, 11):
        servers.append((f"svr-web-{i:02d}", 300 + i))
    for i in range(1, 11):
        servers.append((f"svr-was-{i:02d}", 310 + i))
    for i in range(1, 11):
        servers.append((f"svr-db-{i:02d}", 320 + i))

    for hostname, conf_id in servers:
        updates.append(
            f"UPDATE polestar.cmm_resource "
            f"SET resource_conf_id = {conf_id} "
            f"WHERE hostname = '{hostname}' AND resource_type = 'platform.server' "
            f"AND resource_conf_id IS NULL;"
        )

    return updates


def generate_extra_sample_data() -> tuple[list[str], list[str]]:
    """추가 샘플 데이터 INSERT SQL 생성.

    핵심: CMM_RESOURCE.RESOURCE_CONF_ID = CORE_CONFIG_PROP.CONFIGURATION_ID
    각 서버의 platform.server 행에 resource_conf_id를 설정하고,
    해당 configuration_id로 CORE_CONFIG_PROP에 속성을 INSERT한다.
    """
    cmm_lines = []
    config_lines = []

    cmm_id = 400001
    config_id = 600001
    # configuration_id: 기존 301~330 사용 중이므로 401부터
    conf_id_base = 401

    CTIME_BASE = 1640995200000  # 2022-01-01
    MTIME_BASE = 1719705600000  # 2024-06-30

    extra_servers = [
        ("svr-app-01", "APP", "VMware", "10.1.1.1", 4, "애플리케이션 서버 #1"),
        ("svr-app-02", "APP", "VMware", "10.1.1.2", 4, "애플리케이션 서버 #2"),
        ("svr-app-03", "APP", "HPE", "10.1.1.3", 8, "애플리케이션 서버 #3"),
        ("svr-mid-01", "MID", "Dell", "10.1.2.1", 8, "미들웨어 서버 #1"),
        ("svr-mid-02", "MID", "Dell", "10.1.2.2", 8, "미들웨어 서버 #2"),
        ("svr-bat-01", "BATCH", "VMware", "10.1.3.1", 4, "배치 서버 #1"),
        ("svr-bat-02", "BATCH", "VMware", "10.1.3.2", 4, "배치 서버 #2"),
        ("svr-mon-01", "MON", "HPE", "10.1.4.1", 2, "모니터링 서버 #1"),
        ("svr-bkp-01", "BACKUP", "Dell", "10.1.5.1", 4, "백업 서버 #1"),
        ("svr-log-01", "LOG", "HPE", "10.1.6.1", 4, "로그수집 서버 #1"),
    ]

    fs_lists = {
        "APP": ["/", "/boot", "/fsutil", "/fsapp", "/fslog", "/fshome"],
        "MID": ["/", "/boot", "/fsutil", "/fsapp", "/fslog", "/fshome", "/fsmid"],
        "BATCH": ["/", "/boot", "/fsutil", "/fsapp", "/fslog", "/fsbatch"],
        "MON": ["/", "/boot", "/fsutil", "/fslog"],
        "BACKUP": ["/", "/boot", "/fsutil", "/fslog", "/fsbackup", "/fsnfs"],
        "LOG": ["/", "/boot", "/fsutil", "/fslog", "/fselk"],
    }

    nic_lists = {
        "APP": ["ens192", "ens224"],
        "MID": ["bond0", "ens2f0", "ens5f0"],
        "BATCH": ["ens192"],
        "MON": ["ens192"],
        "BACKUP": ["ens192", "ens224", "bond0"],
        "LOG": ["ens192", "ens224"],
    }

    for idx, (hostname, group, platform, ip, cpu_cores, desc) in enumerate(extra_servers):
        ctime = CTIME_BASE + idx * 86400000 * 15
        mtime = MTIME_BASE + idx * 86400000 * 3
        conf_id_val = conf_id_base + idx  # 이 서버의 configuration_id

        is_abnormal = idx in (2, 6)
        svc_avail = 1 if is_abnormal else 0

        # ===== ServiceResource (최상위 platform.server) =====
        # resource_conf_id = conf_id_val → CORE_CONFIG_PROP.CONFIGURATION_ID와 매핑
        svc_id = cmm_id
        cmm_lines.append(
            f"INSERT INTO polestar.cmm_resource "
            f"(dtype, id, acl_id, avail_status, ctime, mtime, description, "
            f"haschildren, hostname, id_ancestry, importance_id, inheritstatus, "
            f"invisible, ipaddress, name, resource_key, resource_type, "
            f"resource_conf_id) VALUES "
            f"('ServiceResource', {svc_id}, 1, {svc_avail}, {ctime}, {mtime}, "
            f"'{desc}', 1, '{hostname}', '1>', 1, 0, 0, '{ip}', "
            f"'{hostname}', 'extra-{svc_id:06d}', 'platform.server', "
            f"{conf_id_val});"
        )
        cmm_id += 1

        # Monitor group
        mg_id = cmm_id
        cmm_lines.append(
            f"INSERT INTO polestar.cmm_resource "
            f"(dtype, id, acl_id, avail_status, ctime, mtime, "
            f"haschildren, hostname, id_ancestry, importance_id, inheritstatus, "
            f"invisible, ipaddress, name, parent_resource_id, platform_resource_id, "
            f"resource_key, resource_type, service_resource_id) VALUES "
            f"('ServiceResource', {mg_id}, 1, 0, {ctime}, {mtime}, "
            f"1, '{hostname}', '1>{svc_id}>', 1, 0, 0, '{ip}', "
            f"'monitor group', {svc_id}, {svc_id}, "
            f"'extra-{mg_id:06d}', 'management.MonitorGroup', {svc_id});"
        )
        cmm_id += 1

        def add_res(name, rtype, parent_id, desc_val=None, avail=0,
                    host_val=None, ip_val=None, haschildren_val=0, order=None,
                    res_conf_id=None):
            nonlocal cmm_id
            rid = cmm_id
            cmm_id += 1
            desc_sql = f"'{desc_val}'" if desc_val else "NULL"
            host_sql = f"'{host_val}'" if host_val else "NULL"
            ip_sql = f"'{ip_val}'" if ip_val else "NULL"
            order_sql = str(order) if order else "NULL"
            conf_sql = str(res_conf_id) if res_conf_id else "NULL"
            cmm_lines.append(
                f"INSERT INTO polestar.cmm_resource "
                f"(dtype, id, acl_id, avail_status, ctime, mtime, description, "
                f"haschildren, hostname, id_ancestry, importance_id, inheritstatus, "
                f"invisible, ipaddress, name, parent_resource_id, platform_resource_id, "
                f"resource_key, resource_type, service_resource_id, order_num, "
                f"resource_conf_id) VALUES "
                f"('Resource', {rid}, 1, {avail}, {ctime}, {mtime}, {desc_sql}, "
                f"{haschildren_val}, {host_sql}, '1>{svc_id}>{parent_id}>', 1, 0, "
                f"0, {ip_sql}, '{name}', {parent_id}, {svc_id}, "
                f"'extra-{rid:06d}', '{rtype}', {svc_id}, {order_sql}, "
                f"{conf_sql});"
            )
            return rid

        # ===== 서브리소스별 configuration_id (conf_id_val + offset) =====
        # Excel 패턴: server.Server, server.Cpus, server.Disks, server.Memory,
        #             server.VirtualMemory, server.NetworkInterface 에 resource_conf_id가 있음
        cpus_conf_id = conf_id_val + 100   # 501~510
        disks_conf_id = conf_id_val + 200  # 601~610
        mem_conf_id = conf_id_val + 300    # 701~710
        vmem_conf_id = conf_id_val + 400   # 801~810

        # CPU 컨테이너 + 코어
        cpus_id = add_res("CPU", "server.Cpus", svc_id, "CPU 관리",
                          host_val=hostname, ip_val=ip, haschildren_val=1,
                          res_conf_id=cpus_conf_id)
        for c in range(1, cpu_cores + 1):
            add_res(f"Core{c}", "server.Cpu", cpus_id, order=c)

        # 디스크
        add_res("디스크", "server.Disks", svc_id, "전체 디스크",
                host_val=hostname, ip_val=ip, res_conf_id=disks_conf_id)

        # 파일시스템
        fs_id = add_res("파일시스템", "server.FileSystems", svc_id, "파일시스템 관리",
                        host_val=hostname, ip_val=ip, haschildren_val=1)
        fs_list = fs_lists.get(group, ["/", "/boot"])
        for fs_name in fs_list:
            add_res(fs_name, "server.FileSystem", fs_id)

        # 메모리
        mem_id = add_res("메모리", "server.Memory", svc_id, "물리적 메모리",
                         host_val=hostname, ip_val=ip, haschildren_val=1,
                         res_conf_id=mem_conf_id)
        add_res("기타 메모리", "server.OtherMemory", mem_id)
        add_res("가상메모리", "server.VirtualMemory", mem_id,
                res_conf_id=vmem_conf_id)

        # 네트워크 세션
        add_res("네트워크 세션", "server.Netstat", svc_id, "네트워크 연결 정보",
                host_val=hostname, ip_val=ip)

        # 네트워크 인터페이스
        ni_id = add_res("Network Interfaces", "server.NetworkInterfaces", svc_id,
                        "Network Interface 관리", host_val=hostname, ip_val=ip,
                        haschildren_val=1)
        nics = nic_lists.get(group, ["ens192"])
        nic_conf_base = conf_id_val + 500  # 901~
        for nic_idx, nic_name in enumerate(nics):
            nic_conf_id = nic_conf_base + nic_idx
            add_res(nic_name, "server.NetworkInterface", ni_id,
                    res_conf_id=nic_conf_id)

        # 기타정보, 프로세스
        add_res("기타정보", "server.Other", svc_id, host_val=hostname, ip_val=ip)
        add_res("프로세스", "server.Process", svc_id, host_val=hostname, ip_val=ip)

        # LogMonitor, ProcessMonitor
        log_avail = 1 if idx % 3 == 0 else 0
        add_res("Log Monitor", "server.LogMonitor", mg_id, avail=log_avail)
        add_res("Syslog Monitor", "server.LogMonitor", mg_id,
                avail=1 if idx % 4 == 0 else 0)

        proc_monitors = {
            "APP": [("java", 0), ("httpd", 1 if idx % 3 == 0 else 0)],
            "MID": [("java", 0), ("tomcat", 1 if idx % 2 == 0 else 0)],
            "BATCH": [("cron", 0), ("java", 1 if idx % 3 == 0 else 0)],
            "MON": [("prometheus", 0), ("grafana", 0)],
            "BACKUP": [("rsync", 0), ("nfs", 1 if idx % 2 == 0 else 0)],
            "LOG": [("elasticsearch", 0), ("logstash", 0), ("kibana", 0)],
        }
        for pm_name, pm_avail in proc_monitors.get(group, [("app", 0)]):
            add_res(pm_name, "server.ProcessMonitor", mg_id, pm_name, avail=pm_avail)

        # ===== CORE_CONFIG_PROP =====
        if platform == "VMware":
            model = "VMware Virtual Platform"
            vendor = "VMware, Inc."
            serial = f"VMware-{hostname.replace('-', '').upper()}"
        elif platform == "HPE":
            model = "ProLiant DL380 Gen10 Plus"
            vendor = "HPE"
            serial = f"HOST-{hostname.replace('-', '').upper()}"
        else:
            model = "PowerEdge R750"
            vendor = "Dell Inc."
            serial = f"SVCTAG-{hostname.replace('-', '').upper()}"

        versions = ["7.6.26_6", "7.6.28_1", "8.4.12"]
        os_versions = [
            "3.10.0-957.el7.x86_64",
            "4.18.0-305.el8.x86_64",
            "5.14.0-70.el9.x86_64",
        ]

        ts = ctime + 3600000

        # server.Server 속성 (configuration_id = conf_id_val) — 12종
        server_props = [
            ("Vendor", vendor, 454),
            ("Model", model, 455),
            ("OSType", "LINUX", 456),
            ("OSVerson", os_versions[idx % 3], 457),
            ("GMT", "GMT+09:00", 458),
            ("SerialNumber", serial, 459),
            ("Hostname", hostname, 460),
            ("IPaddress", ip, 461),
            ("AgentVersion", versions[idx % 3], 462),
            ("InstallPath", "/fsutil/polestar/agent/NNPAgent/MAgent/", 463),
            ("AgentID", f"MA_{hostname}_{ctime}", 464),
            ("OSParameter", "kernel.shmmax = 68719476736", 465),
        ]
        for prop_name, prop_value, prop_def_id in server_props:
            pv = prop_value.replace("'", "''")
            config_lines.append(
                f"INSERT INTO polestar.core_config_prop "
                f"(dtype, id, name, time_stamp, is_lob, stringvalue_short, "
                f"configuration_id, propertydefinition_id) VALUES "
                f"('SIMPLE', {config_id}, '{prop_name}', {ts}, 0, '{pv}', "
                f"{conf_id_val}, {prop_def_id});"
            )
            config_id += 1

        # server.Cpus 속성 (configuration_id = cpus_conf_id)
        cpu_props = [
            ("NumberOfCPUs", str(cpu_cores), 470),
            ("ProcessorType", "Intel(R) Xeon(R) Gold 6248R", 471),
            ("CurrentClockSpeed", "3000", 472),
        ]
        for prop_name, prop_value, prop_def_id in cpu_props:
            config_lines.append(
                f"INSERT INTO polestar.core_config_prop "
                f"(dtype, id, name, time_stamp, is_lob, stringvalue_short, "
                f"configuration_id, propertydefinition_id) VALUES "
                f"('SIMPLE', {config_id}, '{prop_name}', {ts}, 0, '{prop_value}', "
                f"{cpus_conf_id}, {prop_def_id});"
            )
            config_id += 1

        # server.Disks 속성 (configuration_id = disks_conf_id)
        disk_sizes = {"APP": "107374182400", "MID": "214748364800",
                      "BATCH": "107374182400", "MON": "53687091200",
                      "BACKUP": "429496729600", "LOG": "214748364800"}
        config_lines.append(
            f"INSERT INTO polestar.core_config_prop "
            f"(dtype, id, name, time_stamp, is_lob, stringvalue_short, "
            f"configuration_id, propertydefinition_id) VALUES "
            f"('SIMPLE', {config_id}, 'DiskSize', {ts}, 0, "
            f"'{disk_sizes.get(group, '107374182400')}', "
            f"{disks_conf_id}, 480);"
        )
        config_id += 1

        # server.Memory 속성 (configuration_id = mem_conf_id)
        mem_sizes = {"APP": "8589934592", "MID": "17179869184",
                     "BATCH": "8589934592", "MON": "4294967296",
                     "BACKUP": "8589934592", "LOG": "17179869184"}
        config_lines.append(
            f"INSERT INTO polestar.core_config_prop "
            f"(dtype, id, name, time_stamp, is_lob, stringvalue_short, "
            f"configuration_id, propertydefinition_id) VALUES "
            f"('SIMPLE', {config_id}, 'TotalPhysicalMemory', {ts}, 0, "
            f"'{mem_sizes.get(group, '8589934592')}', "
            f"{mem_conf_id}, 490);"
        )
        config_id += 1

        # server.VirtualMemory 속성 (configuration_id = vmem_conf_id)
        config_lines.append(
            f"INSERT INTO polestar.core_config_prop "
            f"(dtype, id, name, time_stamp, is_lob, stringvalue_short, "
            f"configuration_id, propertydefinition_id) VALUES "
            f"('SIMPLE', {config_id}, 'SwapSize', {ts}, 0, '4294967296', "
            f"{vmem_conf_id}, 491);"
        )
        config_id += 1
        config_lines.append(
            f"INSERT INTO polestar.core_config_prop "
            f"(dtype, id, name, time_stamp, is_lob, stringvalue_short, "
            f"configuration_id, propertydefinition_id) VALUES "
            f"('SIMPLE', {config_id}, 'TotalVirtualMemory', {ts}, 0, "
            f"'{int(mem_sizes.get(group, '8589934592')) + 4294967296}', "
            f"{vmem_conf_id}, 492);"
        )
        config_id += 1

        # server.NetworkInterface 속성 (각 NIC별 configuration_id)
        nics = nic_lists.get(group, ["ens192"])
        for nic_idx, nic_name in enumerate(nics):
            nic_conf_id = nic_conf_base + nic_idx
            nic_ip = f"{ip.rsplit('.', 1)[0]}.{int(ip.rsplit('.', 1)[1]) + nic_idx}"
            nic_props = [
                ("InterfaceName", nic_name, 500),
                ("IPAddress", nic_ip, 501),
                ("MACAddress", f"00:50:56:{idx:02x}:{nic_idx:02x}:01", 502),
                ("Speed", "10000", 503),
            ]
            for prop_name, prop_value, prop_def_id in nic_props:
                pv = prop_value.replace("'", "''")
                config_lines.append(
                    f"INSERT INTO polestar.core_config_prop "
                    f"(dtype, id, name, time_stamp, is_lob, stringvalue_short, "
                    f"configuration_id, propertydefinition_id) VALUES "
                    f"('SIMPLE', {config_id}, '{prop_name}', {ts}, 0, '{pv}', "
                    f"{nic_conf_id}, {prop_def_id});"
                )
                config_id += 1

    return cmm_lines, config_lines


def main():
    parser = argparse.ArgumentParser(description="Excel → PostgreSQL INSERT")
    parser.add_argument("--sql-only", action="store_true", help="SQL 파일만 생성")
    parser.add_argument("--with-extra", action="store_true", help="추가 샘플 데이터 포함")
    args = parser.parse_args()

    cmm_file = SAMPLE_DIR / "CMM_RESOURCE.xlsx"
    config_file = SAMPLE_DIR / "CORE_CONFIG_PROP.xlsx"

    for f in (cmm_file, config_file):
        if not f.exists():
            # 파일명에 행 수 포함 패턴 시도: CMM_RESOURCE(873).xlsx
            alt = list(f.parent.glob(f"{f.stem}*{f.suffix}"))
            if alt:
                print(f"대체 파일 사용: {alt[0].name}")
            else:
                print(f"파일 없음: {f}")
                sys.exit(1)

    # 1. 기존 데이터 수정 (RESOURCE_CONF_ID 매핑)
    fix_lines = generate_fix_existing_data()
    print(f"기존 데이터 수정: {len(fix_lines) - 1} UPDATE")

    # 2. Excel INSERT
    print("Excel 파일 읽는 중...")
    cmm_inserts = read_excel_rows(cmm_file, CMM_COLUMNS, CMM_STRING_COLS)
    config_inserts = read_excel_rows(config_file, CONFIG_COLUMNS, CONFIG_STRING_COLS)
    print(f"  CMM_RESOURCE: {len(cmm_inserts)}행")
    print(f"  CORE_CONFIG_PROP: {len(config_inserts)}행")

    # 3. 추가 샘플 데이터
    extra_cmm = []
    extra_config = []
    if args.with_extra:
        print("추가 샘플 데이터 생성 중...")
        extra_cmm, extra_config = generate_extra_sample_data()
        print(f"  추가 CMM_RESOURCE: {len(extra_cmm)}행")
        print(f"  추가 CORE_CONFIG_PROP: {len(extra_config)}행")

    # 4. SQL 파일 생성
    all_cmm = cmm_inserts + extra_cmm
    all_config = config_inserts + extra_config

    sql_output = OUTPUT_DIR / "05_insert_excel_data.sql"

    lines = [
        "-- =============================================================================",
        "-- Excel 데이터 + 추가 샘플 INSERT",
        "-- 핵심: CMM_RESOURCE.RESOURCE_CONF_ID = CORE_CONFIG_PROP.CONFIGURATION_ID",
        f"-- 생성: scripts/excel_to_pg_insert.py",
        f"-- UPDATE: {len(fix_lines) - 1}건 (기존 RESOURCE_CONF_ID 매핑)",
        f"-- CMM_RESOURCE: {len(all_cmm)}행, CORE_CONFIG_PROP: {len(all_config)}행",
        "-- =============================================================================",
        "",
    ]

    # 기존 데이터 UPDATE
    lines.extend(fix_lines)
    lines.append("")

    # Excel + 추가 CMM_RESOURCE
    lines.append("-- CMM_RESOURCE INSERT")
    lines.append("")
    lines.extend(all_cmm)
    lines.append("")

    # Excel + 추가 CORE_CONFIG_PROP
    lines.append("-- CORE_CONFIG_PROP INSERT")
    lines.append("")
    lines.extend(all_config)
    lines.append("")
    lines.append("COMMIT;")
    lines.append("")

    sql_output.parent.mkdir(parents=True, exist_ok=True)
    with open(sql_output, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"\nSQL 파일 생성: {sql_output}")

    if args.sql_only:
        print("--sql-only: DB 실행 생략")
        return

    # 5. DB 실행
    print("\nPostgreSQL에 실행 중...")
    import subprocess
    result = subprocess.run(
        [
            "docker", "exec", "-i", "polestar_pg",
            "psql", "-U", "polestar_user", "-d", "infradb",
        ],
        input="\n".join(lines) + "\n",
        capture_output=True,
        text=True,
    )

    if result.returncode != 0:
        print(f"ERROR: {result.stderr}")
        sys.exit(1)

    insert_count = result.stdout.count("INSERT")
    update_count = result.stdout.count("UPDATE")
    error_lines = [l for l in result.stdout.split("\n") if "ERROR" in l]
    if error_lines:
        print("에러 발생:")
        for el in error_lines[:10]:
            print(f"  {el}")
    print(f"UPDATE: {update_count}건, INSERT: {insert_count}건")

    # 6. JOIN 검증
    print("\n=== JOIN 검증 ===")
    verify = subprocess.run(
        [
            "docker", "exec", "polestar_pg",
            "psql", "-U", "polestar_user", "-d", "infradb", "-c",
            """
            SELECT
                'JOIN 성공' as status,
                COUNT(DISTINCT cr.hostname) as servers,
                COUNT(DISTINCT cr.resource_conf_id) as conf_ids,
                COUNT(cc.id) as config_rows
            FROM polestar.cmm_resource cr
            JOIN polestar.core_config_prop cc
                ON cr.resource_conf_id = cc.configuration_id
            WHERE cr.resource_type = 'platform.server'
            UNION ALL
            SELECT
                'JOIN 실패 (CONF_ID 없음)',
                COUNT(DISTINCT cr.hostname),
                0,
                0
            FROM polestar.cmm_resource cr
            WHERE cr.resource_type = 'platform.server'
              AND (cr.resource_conf_id IS NULL
                   OR cr.resource_conf_id NOT IN (
                       SELECT DISTINCT configuration_id
                       FROM polestar.core_config_prop
                       WHERE configuration_id IS NOT NULL));
            """,
        ],
        capture_output=True, text=True,
    )
    print(verify.stdout)

    # 총 현황
    verify2 = subprocess.run(
        [
            "docker", "exec", "polestar_pg",
            "psql", "-U", "polestar_user", "-d", "infradb", "-c",
            "SELECT 'cmm_resource' as tbl, COUNT(*) as cnt FROM polestar.cmm_resource "
            "UNION ALL "
            "SELECT 'core_config_prop', COUNT(*) FROM polestar.core_config_prop;",
        ],
        capture_output=True, text=True,
    )
    print("최종 데이터 현황:")
    print(verify2.stdout)


if __name__ == "__main__":
    main()
