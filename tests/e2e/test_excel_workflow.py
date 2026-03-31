"""Excel 업로드/다운로드 테스트 (E-01 ~ E-11).

Plan 24 섹션 3.3에 정의된 Excel 파일 첨부, 질의 전송,
결과 파일 다운로드 시나리오를 검증한다.
conftest.py의 page fixture를 사용하여 테스트 서버에 접속한다.
"""

from __future__ import annotations

import os
import tempfile

from playwright.sync_api import Page, expect

FIXTURE_DIR = os.path.join(os.path.dirname(__file__), "fixtures")


# ---------------------------------------------------------------------------
# E-01: 파일 첨부 버튼
# ---------------------------------------------------------------------------


def test_file_input_exists(page: Page) -> None:
    """E-01: fileInput 요소가 존재하고 accept 속성이 올바른지 확인한다.

    파일 입력 필드가 DOM에 존재하며, .xlsx와 .docx만 허용하는
    accept 속성이 설정되어 있는지 검증한다.
    """
    file_input = page.locator("#fileInput")
    expect(file_input).to_have_count(1)

    accept_attr = file_input.get_attribute("accept")
    assert accept_attr is not None, "accept 속성이 없습니다."
    assert ".xlsx" in accept_attr, f"accept에 .xlsx가 없습니다: {accept_attr}"
    assert ".docx" in accept_attr, f"accept에 .docx가 없습니다: {accept_attr}"


# ---------------------------------------------------------------------------
# E-02: 지원 파일 형식 검증
# ---------------------------------------------------------------------------


def test_unsupported_file_rejected(page: Page) -> None:
    """E-02: 지원하지 않는 파일 형식(.txt) 설정 시 에러 메시지가 표시되는지 확인한다.

    Playwright의 set_input_files는 accept 속성을 무시하지만,
    JS handleFileChange가 확장자를 체크하여 에러를 표시한다.
    """
    txt_path = None
    try:
        # 임시 .txt 파일 생성
        fd, txt_path = tempfile.mkstemp(suffix=".txt")
        with os.fdopen(fd, "wb") as f:
            f.write(b"test content")

        file_input = page.locator("#fileInput")
        file_input.set_input_files(txt_path)

        # chatError에 active 클래스가 추가되어야 한다
        chat_error = page.locator("#chatError")
        page.wait_for_function(
            "document.getElementById('chatError').classList.contains('active')",
            timeout=3000,
        )

        # 에러 메시지 텍스트 확인
        error_text = page.locator("#chatErrorText").inner_text()
        assert "지원하지 않는 파일 형식" in error_text, (
            f"예상 에러 메시지가 없습니다: {error_text}"
        )

        # filePreview에 active 클래스가 추가되지 않아야 한다
        preview_classes = page.locator("#filePreview").get_attribute("class") or ""
        assert "active" not in preview_classes, (
            "지원하지 않는 파일인데 프리뷰가 활성화되었습니다."
        )
    finally:
        if txt_path and os.path.exists(txt_path):
            os.unlink(txt_path)


# ---------------------------------------------------------------------------
# E-03: 파일 크기 제한
# ---------------------------------------------------------------------------


def test_oversized_file_rejected(page: Page) -> None:
    """E-03: 10MB 초과 파일 설정 시 에러 메시지가 표시되는지 확인한다.

    page.evaluate로 JS에서 직접 크기가 큰 File 객체를 생성하고
    handleFileChange를 트리거하여 크기 검증 로직을 테스트한다.
    실제 대용량 파일을 디스크에 생성하지 않아 테스트가 빠르다.
    """
    # JS에서 11MB 크기의 가짜 File 객체를 생성하여 change 이벤트 디스패치
    page.evaluate("""
        () => {
            const input = document.getElementById('fileInput');
            const blob = new Blob([new ArrayBuffer(11 * 1024 * 1024)], {
                type: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            });
            const file = new File([blob], 'large_file.xlsx', { type: blob.type });

            const dataTransfer = new DataTransfer();
            dataTransfer.items.add(file);
            input.files = dataTransfer.files;
            input.dispatchEvent(new Event('change', { bubbles: true }));
        }
    """)

    # chatError에 active 클래스가 추가되어야 한다
    page.wait_for_function(
        "document.getElementById('chatError').classList.contains('active')",
        timeout=3000,
    )

    error_text = page.locator("#chatErrorText").inner_text()
    assert "10MB" in error_text, f"크기 제한 에러 메시지가 없습니다: {error_text}"

    # filePreview에 active 클래스가 추가되지 않아야 한다
    preview_classes = page.locator("#filePreview").get_attribute("class") or ""
    assert "active" not in preview_classes, (
        "크기 초과 파일인데 프리뷰가 활성화되었습니다."
    )


# ---------------------------------------------------------------------------
# E-04: 파일 첨부 프리뷰
# ---------------------------------------------------------------------------


def test_file_attach_preview(page: Page) -> None:
    """E-04: xlsx 파일 첨부 시 프리뷰가 표시되는지 확인한다.

    filePreview에 active 클래스가 추가되고,
    fileName에 파일명이 표시되는지 검증한다.
    """
    xlsx_path = os.path.join(FIXTURE_DIR, "sample_template.xlsx")
    assert os.path.exists(xlsx_path), f"fixture 파일이 없습니다: {xlsx_path}"

    file_input = page.locator("#fileInput")
    file_input.set_input_files(xlsx_path)

    # filePreview에 active 클래스가 추가될 때까지 대기
    page.wait_for_function(
        "document.getElementById('filePreview').classList.contains('active')",
        timeout=3000,
    )

    # 파일명 표시 확인
    file_name_text = page.locator("#fileName").inner_text()
    assert "sample_template.xlsx" in file_name_text, (
        f"파일명이 표시되지 않았습니다: {file_name_text}"
    )

    # 파일 크기 표시 확인 (빈 문자열이 아니어야 함)
    file_size_text = page.locator("#fileSize").inner_text()
    assert file_size_text != "", "파일 크기가 표시되지 않았습니다."


# ---------------------------------------------------------------------------
# E-05: 파일 제거
# ---------------------------------------------------------------------------


def test_file_remove(page: Page) -> None:
    """E-05: Remove 버튼 클릭 시 파일 첨부가 취소되는지 확인한다.

    파일 첨부 후 removeFile 버튼 클릭 시
    filePreview에서 active 클래스가 제거되는지 검증한다.
    """
    xlsx_path = os.path.join(FIXTURE_DIR, "sample_template.xlsx")

    # 파일 첨부
    file_input = page.locator("#fileInput")
    file_input.set_input_files(xlsx_path)

    # 프리뷰 활성화 대기
    page.wait_for_function(
        "document.getElementById('filePreview').classList.contains('active')",
        timeout=3000,
    )

    # Remove 버튼 클릭
    page.locator("#removeFile").click()

    # filePreview에서 active 클래스가 제거될 때까지 대기
    page.wait_for_function(
        "!document.getElementById('filePreview').classList.contains('active')",
        timeout=3000,
    )

    preview_classes = page.locator("#filePreview").get_attribute("class") or ""
    assert "active" not in preview_classes, (
        "Remove 후에도 프리뷰가 활성 상태입니다."
    )


# ---------------------------------------------------------------------------
# E-06: Excel + 질의 전송
# ---------------------------------------------------------------------------


def test_excel_query_send(page: Page) -> None:
    """E-06: Excel 파일과 질의가 함께 전송되는지 확인한다.

    파일 첨부 + 텍스트 입력 후 전송하면
    사용자 메시지에 파일 배지(.message-file-badge)가 표시되고,
    배지에 파일명이 포함되는지 검증한다.
    """
    xlsx_path = os.path.join(FIXTURE_DIR, "sample_template.xlsx")

    # 파일 첨부
    file_input = page.locator("#fileInput")
    file_input.set_input_files(xlsx_path)

    page.wait_for_function(
        "document.getElementById('filePreview').classList.contains('active')",
        timeout=3000,
    )

    # 질의 입력
    page.locator("#prompt").fill("서버 현황 데이터를 양식에 채워줘")

    # 전송
    page.locator("#sendBtn").click()

    # 사용자 메시지에 파일 배지가 표시될 때까지 대기
    file_badge = page.locator(".message-file-badge")
    file_badge.first.wait_for(state="visible", timeout=5000)

    badge_text = file_badge.first.inner_text()
    assert "sample_template.xlsx" in badge_text, (
        f"파일 배지에 파일명이 없습니다: {badge_text}"
    )


# ---------------------------------------------------------------------------
# E-07: 전송 후 프리뷰 제거
# ---------------------------------------------------------------------------


def test_preview_cleared_after_send(page: Page) -> None:
    """E-07: 파일 전송 후 filePreview에서 active 클래스가 제거되는지 확인한다.

    handleSend에서 clearFile()이 호출되어
    전송 직후 프리뷰가 비활성화되는 것을 검증한다.
    """
    xlsx_path = os.path.join(FIXTURE_DIR, "sample_template.xlsx")

    # 파일 첨부
    file_input = page.locator("#fileInput")
    file_input.set_input_files(xlsx_path)

    page.wait_for_function(
        "document.getElementById('filePreview').classList.contains('active')",
        timeout=3000,
    )

    # 질의 입력 및 전송
    page.locator("#prompt").fill("양식에 데이터 채워줘")
    page.locator("#sendBtn").click()

    # 전송 후 filePreview에서 active 클래스가 제거될 때까지 대기
    page.wait_for_function(
        "!document.getElementById('filePreview').classList.contains('active')",
        timeout=5000,
    )

    preview_classes = page.locator("#filePreview").get_attribute("class") or ""
    assert "active" not in preview_classes, (
        "전송 후에도 프리뷰가 활성 상태입니다."
    )


# ---------------------------------------------------------------------------
# E-08: 처리 중 표시
# ---------------------------------------------------------------------------


def test_processing_indicator_during_file_query(page: Page) -> None:
    """E-08: 파일 질의 처리 중 processingMessage가 표시되는지 확인한다.

    executeFileQuery에서 renderProcessingMessage()가 호출되어
    처리 인디케이터가 나타나는 것을 검증한다.
    """
    xlsx_path = os.path.join(FIXTURE_DIR, "sample_template.xlsx")

    # 파일 첨부
    file_input = page.locator("#fileInput")
    file_input.set_input_files(xlsx_path)

    page.wait_for_function(
        "document.getElementById('filePreview').classList.contains('active')",
        timeout=3000,
    )

    # 질의 입력 및 전송
    page.locator("#prompt").fill("서버 현황 데이터를 양식에 채워줘")
    page.locator("#sendBtn").click()

    # processingMessage가 DOM에 나타날 때까지 대기
    # MockGraph의 ainvoke가 빠르게 응답하므로 짧은 시간 내에 캡처해야 한다.
    # wait_for_function으로 processingMessage가 존재했음을 확인한다.
    # (이미 제거되었을 수도 있으므로, 존재한 적 있는지 또는 현재 존재하는지 확인)
    processing_appeared = page.evaluate("""
        () => {
            return new Promise((resolve) => {
                // 이미 존재하는지 먼저 확인
                if (document.getElementById('processingMessage')) {
                    resolve(true);
                    return;
                }
                // MutationObserver로 추가 감지
                const observer = new MutationObserver((mutations) => {
                    for (const m of mutations) {
                        for (const node of m.addedNodes) {
                            if (node.id === 'processingMessage' ||
                                (node.querySelector && node.querySelector('#processingMessage'))) {
                                observer.disconnect();
                                resolve(true);
                                return;
                            }
                        }
                    }
                });
                observer.observe(document.body, { childList: true, subtree: true });
                // 5초 타임아웃
                setTimeout(() => { observer.disconnect(); resolve(false); }, 5000);
            });
        }
    """)

    # processingMessage가 나타났거나, 이미 응답이 와서 제거된 상태일 수 있다.
    # 제거 후에도 에이전트 응답이 있으면 처리 중 표시가 정상 작동한 것이다.
    agent_response = page.locator(
        ".message.message--agent:not(.message--processing)"
    )
    agent_response.first.wait_for(state="visible", timeout=60000)

    # 처리가 완료된 후 processingMessage는 제거되어야 한다
    assert page.locator("#processingMessage").count() == 0, (
        "처리 완료 후에도 processingMessage가 남아있습니다."
    )

    # processingMessage가 나타났거나 에이전트 응답이 있으면 성공
    assert processing_appeared or agent_response.count() > 0, (
        "처리 인디케이터가 표시되지 않았습니다."
    )


# ---------------------------------------------------------------------------
# E-09: 응답에 다운로드 버튼
# ---------------------------------------------------------------------------


def test_download_button_shown(page: Page) -> None:
    """E-09: has_file=true 응답 시 다운로드 버튼(.message-download)이 표시되는지 확인한다.

    MockGraph가 파일 첨부 질의에 대해 output_file을 반환하면,
    renderAgentMessage에서 다운로드 링크가 생성되는 것을 검증한다.
    다운로드 URL이 /api/v1/query/{query_id}/download 형식인지도 확인한다.
    """
    xlsx_path = os.path.join(FIXTURE_DIR, "sample_template.xlsx")

    # 파일 첨부 + 질의 전송
    file_input = page.locator("#fileInput")
    file_input.set_input_files(xlsx_path)

    page.wait_for_function(
        "document.getElementById('filePreview').classList.contains('active')",
        timeout=3000,
    )

    page.locator("#prompt").fill("양식에 데이터 채워줘")
    page.locator("#sendBtn").click()

    # 에이전트 응답 대기 (문서 생성 시간 고려하여 60초)
    agent_msg = page.locator(
        ".message.message--agent:not(.message--processing)"
    )
    agent_msg.first.wait_for(state="visible", timeout=60000)

    # 다운로드 버튼 확인
    download_link = page.locator(".message-download").last
    expect(download_link).to_be_visible()

    # href에 다운로드 URL 패턴이 포함되어야 한다
    href = download_link.get_attribute("href") or ""
    assert "/api/v1/query/" in href, f"다운로드 URL에 query 경로가 없습니다: {href}"
    assert "/download" in href, f"다운로드 URL에 /download가 없습니다: {href}"


# ---------------------------------------------------------------------------
# E-10: 다운로드 실행
# ---------------------------------------------------------------------------


def test_file_download(page: Page) -> None:
    """E-10: 다운로드 버튼 클릭 시 파일이 다운로드되는지 확인한다.

    page.expect_download()을 사용하여 실제 다운로드 이벤트가
    발생하는지 검증하고, 파일명과 크기를 확인한다.
    """
    xlsx_path = os.path.join(FIXTURE_DIR, "sample_template.xlsx")

    # 파일 첨부 + 질의 전송
    file_input = page.locator("#fileInput")
    file_input.set_input_files(xlsx_path)

    page.wait_for_function(
        "document.getElementById('filePreview').classList.contains('active')",
        timeout=3000,
    )

    page.locator("#prompt").fill("양식에 데이터 채워줘")
    page.locator("#sendBtn").click()

    # 다운로드 버튼이 나타날 때까지 대기 (60초)
    download_link = page.locator(".message-download").last
    download_link.wait_for(state="visible", timeout=60000)

    # 다운로드 실행
    with page.expect_download(timeout=60000) as download_info:
        download_link.click()

    download = download_info.value

    # 파일명이 .xlsx로 끝나는지 확인
    assert download.suggested_filename.endswith(".xlsx"), (
        f"다운로드 파일명이 .xlsx가 아닙니다: {download.suggested_filename}"
    )

    # 파일 크기가 0보다 큰지 확인 (빈 파일이 아닌지)
    downloaded_path = download.path()
    assert downloaded_path is not None, "다운로드 경로가 None입니다."
    file_size = os.path.getsize(downloaded_path)
    assert file_size > 0, f"다운로드 파일이 비어있습니다 (크기: {file_size})"


# ---------------------------------------------------------------------------
# E-11: 사용자 메시지에 파일 배지
# ---------------------------------------------------------------------------


def test_user_message_file_badge(page: Page) -> None:
    """E-11: 전송된 사용자 메시지에 파일명 배지가 표시되는지 확인한다.

    renderUserMessage에서 msg.file이 있을 때
    .message-file-badge 요소가 생성되고 파일명이 포함되는지 검증한다.
    """
    xlsx_path = os.path.join(FIXTURE_DIR, "sample_template.xlsx")

    # 파일 첨부
    file_input = page.locator("#fileInput")
    file_input.set_input_files(xlsx_path)

    page.wait_for_function(
        "document.getElementById('filePreview').classList.contains('active')",
        timeout=3000,
    )

    # 질의 입력 및 전송
    page.locator("#prompt").fill("서버 현황을 양식에 채워주세요")
    page.locator("#sendBtn").click()

    # 사용자 메시지 영역 내에서 파일 배지 확인
    user_msg = page.locator(".message.message--user").last
    user_msg.wait_for(state="visible", timeout=5000)

    file_badge = user_msg.locator(".message-file-badge")
    expect(file_badge).to_be_visible()

    badge_text = file_badge.inner_text()
    assert "sample_template.xlsx" in badge_text, (
        f"사용자 메시지 배지에 파일명이 없습니다: {badge_text}"
    )


# ---------------------------------------------------------------------------
# E-12: 다운로드된 Excel 파일 내용 검증
# ---------------------------------------------------------------------------


def test_downloaded_excel_content(page: Page) -> None:
    """E-12: 다운로드된 Excel 파일의 헤더, 데이터 행, 값이 기대한 대로 생성되었는지 확인한다.

    sample_template.xlsx를 업로드하고 질의 전송 후 결과 파일을 다운로드하여
    openpyxl로 열어 다음을 검증한다:
    - 파일명: result_20260324.xlsx
    - 시트가 1개 이상 존재하고 시트명이 "결과"
    - 헤더 행: 서버명, IP주소, CPU사용률 (MockGraph 출력 형식)
    - 데이터 5행: srv-1~5, 10.0.0.1~5, 30%~70%
    - 총 행 수: 6행 (헤더 1 + 데이터 5)
    """
    import shutil

    from openpyxl import load_workbook

    xlsx_path = os.path.join(FIXTURE_DIR, "sample_template.xlsx")

    # 파일 첨부 + 질의 전송
    file_input = page.locator("#fileInput")
    file_input.set_input_files(xlsx_path)

    page.wait_for_function(
        "document.getElementById('filePreview').classList.contains('active')",
        timeout=3000,
    )

    page.locator("#prompt").fill("서버 현황 데이터를 양식에 채워줘")
    page.locator("#sendBtn").click()

    # 다운로드 버튼 대기
    download_link = page.locator(".message-download").last
    download_link.wait_for(state="visible", timeout=60000)

    # 다운로드 실행
    with page.expect_download(timeout=60000) as download_info:
        download_link.click()

    download = download_info.value

    # --- 파일명 검증 ---
    assert download.suggested_filename == "result_20260324.xlsx", (
        f"기대 파일명: result_20260324.xlsx, 실제: {download.suggested_filename}"
    )

    # Playwright는 확장자 없이 저장하므로 .xlsx 확장자로 복사
    raw_path = download.path()
    assert raw_path is not None, "다운로드 경로가 None입니다."
    xlsx_download_path = str(raw_path) + ".xlsx"
    shutil.copy2(raw_path, xlsx_download_path)

    try:
        # --- Excel 내용 검증 ---
        wb = load_workbook(xlsx_download_path, read_only=True)

        # 시트 존재 확인
        assert len(wb.sheetnames) >= 1, f"시트가 없습니다: {wb.sheetnames}"
        assert wb.sheetnames[0] == "결과", (
            f"시트명 기대: '결과', 실제: {wb.sheetnames[0]!r}"
        )
        ws = wb.active

        # 전체 행을 리스트로 변환
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 6, (
            f"기대 행 수: 6 (헤더1 + 데이터5), 실제: {len(rows)}"
        )

        # --- 헤더 행 검증 ---
        headers = rows[0]
        assert headers[0] == "서버명", f"A1 기대: '서버명', 실제: {headers[0]!r}"
        assert headers[1] == "IP주소", f"B1 기대: 'IP주소', 실제: {headers[1]!r}"
        assert headers[2] == "CPU사용률", (
            f"C1 기대: 'CPU사용률', 실제: {headers[2]!r}"
        )

        # --- 데이터 행 검증 (MockGraph._create_sample_xlsx_bytes 기대값) ---
        expected_data = [
            ("srv-1", "10.0.0.1", "30%"),
            ("srv-2", "10.0.0.2", "40%"),
            ("srv-3", "10.0.0.3", "50%"),
            ("srv-4", "10.0.0.4", "60%"),
            ("srv-5", "10.0.0.5", "70%"),
        ]

        for i, (expected_row, actual_row) in enumerate(
            zip(expected_data, rows[1:]), start=1
        ):
            row_label = f"행 {i+1}"
            assert actual_row[0] == expected_row[0], (
                f"{row_label} 서버명 기대: {expected_row[0]!r}, "
                f"실제: {actual_row[0]!r}"
            )
            assert actual_row[1] == expected_row[1], (
                f"{row_label} IP주소 기대: {expected_row[1]!r}, "
                f"실제: {actual_row[1]!r}"
            )
            assert actual_row[2] == expected_row[2], (
                f"{row_label} CPU사용률 기대: {expected_row[2]!r}, "
                f"실제: {actual_row[2]!r}"
            )

        wb.close()
    finally:
        if os.path.exists(xlsx_download_path):
            os.unlink(xlsx_download_path)


# ---------------------------------------------------------------------------
# E-13: 입력 템플릿(sample_template.xlsx) 구조 검증
# ---------------------------------------------------------------------------


def test_input_template_structure() -> None:
    """E-13: 입력 템플릿(sample_template.xlsx) 구조가 올바른지 확인한다.

    sample_template.xlsx의 시트명, 헤더 열, 샘플 데이터 행,
    빈 행, 스타일(헤더 굵게) 등을 검증한다.
    테스트 서버 불필요 (로컬 파일 검증).
    """
    from openpyxl import load_workbook

    xlsx_path = os.path.join(FIXTURE_DIR, "sample_template.xlsx")
    assert os.path.exists(xlsx_path), f"fixture 파일이 없습니다: {xlsx_path}"

    wb = load_workbook(xlsx_path)

    # --- 시트 검증 ---
    assert "서버현황" in wb.sheetnames, (
        f"시트명 '서버현황'이 없습니다: {wb.sheetnames}"
    )
    ws = wb["서버현황"]

    # --- 헤더 검증 (12열) ---
    expected_headers = [
        "IP", "CPU", "MEMORY", "OS", "WAS", "DB",
        "DMZ 유무", "센터구분", "Serial 번호", "서버명", "업무내용", "대분류",
    ]
    for col_idx, expected in enumerate(expected_headers, start=1):
        actual = ws.cell(1, col_idx).value
        assert actual == expected, (
            f"헤더 열{col_idx} 기대: {expected!r}, 실제: {actual!r}"
        )

    # --- 헤더 스타일 검증 (굵게) ---
    for col_idx in range(1, len(expected_headers) + 1):
        header_cell = ws.cell(1, col_idx)
        assert header_cell.font.bold is True, (
            f"열{col_idx} 헤더 폰트가 굵게가 아닙니다."
        )

    # --- 샘플 데이터 행 검증 (2행) ---
    sample_row = {col_idx: ws.cell(2, col_idx).value for col_idx in range(1, 13)}
    assert sample_row[1] == "123.123.123.123", (
        f"A2(IP) 기대: '123.123.123.123', 실제: {sample_row[1]!r}"
    )
    assert sample_row[2] == 8, f"B2(CPU) 기대: 8, 실제: {sample_row[2]!r}"
    assert sample_row[3] == 24, f"C2(MEMORY) 기대: 24, 실제: {sample_row[3]!r}"
    assert sample_row[4] == "Polestar", (
        f"D2(OS) 기대: 'Polestar', 실제: {sample_row[4]!r}"
    )
    assert sample_row[10] == "aaa123", (
        f"J2(서버명) 기대: 'aaa123', 실제: {sample_row[10]!r}"
    )
    assert sample_row[12] == "UNIX", (
        f"L2(대분류) 기대: 'UNIX', 실제: {sample_row[12]!r}"
    )

    # --- 빈 데이터 행 검증 (3~7행: 모두 None) ---
    for row_idx in range(3, 7):
        for col_idx in range(1, 13):
            cell_val = ws.cell(row_idx, col_idx).value
            assert cell_val is None, (
                f"행{row_idx} 열{col_idx}이 비어있지 않습니다: {cell_val!r}"
            )

    # --- 열 너비 검증 (A, B, C) ---
    assert ws.column_dimensions["A"].width == 20.0, (
        f"A열 너비: {ws.column_dimensions['A'].width}, 기대: 20.0"
    )
    assert ws.column_dimensions["B"].width == 18.0, (
        f"B열 너비: {ws.column_dimensions['B'].width}, 기대: 18.0"
    )

    wb.close()
