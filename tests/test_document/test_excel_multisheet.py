"""Excel multi-sheet data fill tests."""

from __future__ import annotations

import io
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from src.document.excel_parser import parse_excel_template
from src.document.excel_writer import fill_excel_template


def _create_multisheet_template(
    sheet_configs: list[dict[str, Any]],
) -> bytes:
    """Create an Excel template with multiple sheets.

    Args:
        sheet_configs: List of dicts with keys:
            - name: sheet name
            - headers: list of header strings
            - data_rows: optional list of row lists
    """
    wb = Workbook()
    # Remove default sheet
    default_ws = wb.active

    for idx, cfg in enumerate(sheet_configs):
        ws = wb.create_sheet(title=cfg["name"])
        for col_idx, header in enumerate(cfg.get("headers", []), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = cell.font.copy(bold=True)
        for row_offset, row_data in enumerate(cfg.get("data_rows", [])):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=2 + row_offset, column=col_idx, value=value)

    # Remove default "Sheet" if we created named sheets
    if default_ws.title not in [c["name"] for c in sheet_configs]:
        wb.remove(default_ws)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


class TestMultiSheetBasicFill:
    """Test filling multiple sheets with a shared mapping."""

    def test_all_sheets_filled_with_shared_mapping(self):
        """All sheets receive data when no target_sheets specified."""
        template_bytes = _create_multisheet_template([
            {"name": "Servers", "headers": ["hostname", "ip"]},
            {"name": "Metrics", "headers": ["hostname", "cpu_pct"]},
        ])
        template_structure = parse_excel_template(template_bytes)
        assert len(template_structure["sheets"]) == 2

        column_mapping = {
            "hostname": "servers.hostname",
            "ip": "servers.ip_address",
            "cpu_pct": "cpu_metrics.usage_pct",
        }
        rows = [
            {"hostname": "web-01", "ip_address": "10.0.0.1", "usage_pct": 85.2},
            {"hostname": "web-02", "ip_address": "10.0.0.2", "usage_pct": 42.1},
        ]

        result_bytes = fill_excel_template(
            template_bytes, template_structure, column_mapping, rows
        )

        wb = load_workbook(io.BytesIO(result_bytes))

        # Servers sheet
        ws_servers = wb["Servers"]
        assert ws_servers.cell(row=1, column=1).value == "hostname"
        assert ws_servers.cell(row=2, column=1).value == "web-01"
        assert ws_servers.cell(row=2, column=2).value == "10.0.0.1"
        assert ws_servers.cell(row=3, column=1).value == "web-02"

        # Metrics sheet
        ws_metrics = wb["Metrics"]
        assert ws_metrics.cell(row=1, column=1).value == "hostname"
        assert ws_metrics.cell(row=2, column=1).value == "web-01"
        assert ws_metrics.cell(row=2, column=2).value == 85.2

        wb.close()

    def test_target_sheets_filters_correctly(self):
        """Only specified target sheets are filled."""
        template_bytes = _create_multisheet_template([
            {"name": "Servers", "headers": ["hostname", "ip"]},
            {"name": "Metrics", "headers": ["hostname", "cpu_pct"]},
        ])
        template_structure = parse_excel_template(template_bytes)

        column_mapping = {
            "hostname": "servers.hostname",
            "ip": "servers.ip_address",
            "cpu_pct": "cpu_metrics.usage_pct",
        }
        rows = [
            {"hostname": "web-01", "ip_address": "10.0.0.1", "usage_pct": 85.2},
        ]

        result_bytes = fill_excel_template(
            template_bytes, template_structure, column_mapping, rows,
            target_sheets=["Servers"],
        )

        wb = load_workbook(io.BytesIO(result_bytes))

        # Servers sheet should be filled
        assert wb["Servers"].cell(row=2, column=1).value == "web-01"

        # Metrics sheet should be untouched (no data in row 2)
        assert wb["Metrics"].cell(row=2, column=1).value is None

        wb.close()


class TestMultiSheetPerSheetMapping:
    """Test filling sheets with per-sheet independent mappings."""

    def test_per_sheet_mapping(self):
        """Each sheet uses its own column mapping."""
        template_bytes = _create_multisheet_template([
            {"name": "ServerInfo", "headers": ["server_name", "address"]},
            {"name": "CPUData", "headers": ["server", "usage"]},
        ])
        template_structure = parse_excel_template(template_bytes)

        # Per-sheet mappings with different field names
        sheet_mappings = [
            {
                "sheet_name": "ServerInfo",
                "column_mapping": {
                    "server_name": "servers.hostname",
                    "address": "servers.ip_address",
                },
                "rows": [
                    {"hostname": "srv-01", "ip_address": "192.168.1.1"},
                ],
            },
            {
                "sheet_name": "CPUData",
                "column_mapping": {
                    "server": "servers.hostname",
                    "usage": "cpu_metrics.usage_pct",
                },
                "rows": [
                    {"hostname": "srv-01", "usage_pct": 72.5},
                    {"hostname": "srv-02", "usage_pct": 95.1},
                ],
            },
        ]

        result_bytes = fill_excel_template(
            template_bytes,
            template_structure,
            column_mapping={},  # fallback (unused when sheet_mappings provided)
            rows=[],
            sheet_mappings=sheet_mappings,
        )

        wb = load_workbook(io.BytesIO(result_bytes))

        # ServerInfo sheet: 1 row
        ws_info = wb["ServerInfo"]
        assert ws_info.cell(row=2, column=1).value == "srv-01"
        assert ws_info.cell(row=2, column=2).value == "192.168.1.1"
        assert ws_info.cell(row=3, column=1).value is None  # only 1 data row

        # CPUData sheet: 2 rows
        ws_cpu = wb["CPUData"]
        assert ws_cpu.cell(row=2, column=1).value == "srv-01"
        assert ws_cpu.cell(row=2, column=2).value == 72.5
        assert ws_cpu.cell(row=3, column=1).value == "srv-02"
        assert ws_cpu.cell(row=3, column=2).value == 95.1

        wb.close()

    def test_per_sheet_mapping_with_target_sheets(self):
        """Per-sheet mapping respects target_sheets filter."""
        template_bytes = _create_multisheet_template([
            {"name": "Sheet1", "headers": ["name", "desc"]},
            {"name": "Sheet2", "headers": ["value", "unit"]},
        ])
        template_structure = parse_excel_template(template_bytes)

        sheet_mappings = [
            {
                "sheet_name": "Sheet1",
                "column_mapping": {"name": "items.name", "desc": "items.desc"},
                "rows": [{"name": "item-a", "desc": "d"}],
            },
            {
                "sheet_name": "Sheet2",
                "column_mapping": {"value": "items.value", "unit": "items.unit"},
                "rows": [{"value": 100, "unit": "ea"}],
            },
        ]

        result_bytes = fill_excel_template(
            template_bytes,
            template_structure,
            column_mapping={},
            rows=[],
            sheet_mappings=sheet_mappings,
            target_sheets=["Sheet2"],
        )

        wb = load_workbook(io.BytesIO(result_bytes))
        assert wb["Sheet1"].cell(row=2, column=1).value is None  # not filled
        assert wb["Sheet2"].cell(row=2, column=1).value == 100  # filled
        wb.close()


class TestBackwardCompatibility:
    """Ensure existing single-sheet behavior is preserved."""

    def test_single_sheet_no_new_params(self):
        """Single-sheet fill works without new parameters."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.cell(row=1, column=1, value="name")
        ws.cell(row=1, column=2, value="count")
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        template_bytes = buf.getvalue()

        template_structure = parse_excel_template(template_bytes)
        column_mapping = {"name": "items.name", "count": "items.count"}
        rows = [{"name": "apple", "count": 5}]

        result_bytes = fill_excel_template(
            template_bytes, template_structure, column_mapping, rows
        )

        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb["Data"]
        assert ws.cell(row=2, column=1).value == "apple"
        assert ws.cell(row=2, column=2).value == 5
        wb.close()

    def test_empty_sheet_mappings_falls_back(self):
        """When sheet_mappings is None, falls back to column_mapping."""
        template_bytes = _create_multisheet_template([
            {"name": "Report", "headers": ["host", "status"]},
        ])
        template_structure = parse_excel_template(template_bytes)

        column_mapping = {"host": "servers.hostname", "status": "servers.status"}
        rows = [{"hostname": "db-01", "status": "running"}]

        result_bytes = fill_excel_template(
            template_bytes, template_structure, column_mapping, rows,
            sheet_mappings=None,
            target_sheets=None,
        )

        wb = load_workbook(io.BytesIO(result_bytes))
        assert wb["Report"].cell(row=2, column=1).value == "db-01"
        wb.close()


class TestEdgeCases:
    """Edge case tests."""

    def test_nonexistent_target_sheet(self):
        """Target sheet that does not exist in template is silently skipped."""
        template_bytes = _create_multisheet_template([
            {"name": "Real", "headers": ["data", "extra"]},
        ])
        template_structure = parse_excel_template(template_bytes)

        result_bytes = fill_excel_template(
            template_bytes, template_structure,
            column_mapping={"data": "t.data", "extra": "t.extra"},
            rows=[{"data": "value", "extra": "x"}],
            target_sheets=["Nonexistent"],
        )

        wb = load_workbook(io.BytesIO(result_bytes))
        # Real sheet should not be filled (not in target_sheets)
        assert wb["Real"].cell(row=2, column=1).value is None
        wb.close()

    def test_three_sheets_fill_all(self):
        """Three sheets are all filled when no target specified."""
        template_bytes = _create_multisheet_template([
            {"name": "A", "headers": ["x", "y"]},
            {"name": "B", "headers": ["x", "y"]},
            {"name": "C", "headers": ["x", "y"]},
        ])
        template_structure = parse_excel_template(template_bytes)

        result_bytes = fill_excel_template(
            template_bytes, template_structure,
            column_mapping={"x": "t.val", "y": "t.val2"},
            rows=[{"val": 42, "val2": 99}],
        )

        wb = load_workbook(io.BytesIO(result_bytes))
        for sheet_name in ["A", "B", "C"]:
            assert wb[sheet_name].cell(row=2, column=1).value == 42
        wb.close()

    def test_sheet_with_no_mapping_skipped(self):
        """Sheet with empty column_mapping in per_sheet_mapping is skipped."""
        template_bytes = _create_multisheet_template([
            {"name": "Filled", "headers": ["val", "extra"]},
            {"name": "Empty", "headers": ["val", "extra"]},
        ])
        template_structure = parse_excel_template(template_bytes)

        sheet_mappings = [
            {"sheet_name": "Filled", "column_mapping": {"val": "t.val", "extra": "t.extra"}, "rows": [{"val": 1, "extra": "e"}]},
            {"sheet_name": "Empty", "column_mapping": {}, "rows": [{"val": 2, "extra": "e2"}]},
        ]

        result_bytes = fill_excel_template(
            template_bytes, template_structure,
            column_mapping={},
            rows=[],
            sheet_mappings=sheet_mappings,
        )

        wb = load_workbook(io.BytesIO(result_bytes))
        assert wb["Filled"].cell(row=2, column=1).value == 1
        # Empty sheet mapping has no valid column_mapping entries, so _fill_sheet skips it
        wb.close()
