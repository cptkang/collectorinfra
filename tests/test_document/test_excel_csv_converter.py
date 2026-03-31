"""Excel → CSV 변환 모듈 단위 테스트.

Plan 19: excel_csv_converter.py의 CsvSheetData, excel_to_csv() 함수,
CSV 변환 실패 시 template_structure 폴백 로직을 검증한다.
"""

from __future__ import annotations

import io
from dataclasses import fields as dataclass_fields
from datetime import date, datetime, time

import pytest
from openpyxl import Workbook

from src.document.excel_csv_converter import (
    CsvConversionError,
    CsvSheetData,
    _build_csv_text,
    _format_cell_value,
    excel_to_csv,
)


# ===== 테스트용 Excel 생성 헬퍼 =====


def _make_excel_bytes(
    headers: list[str] | None = None,
    data_rows: list[list] | None = None,
    sheet_name: str = "Sheet1",
    header_row: int = 1,
    merged_cells: list[str] | None = None,
    extra_sheets: list[dict] | None = None,
) -> bytes:
    """테스트용 Excel 파일 바이너리를 프로그래밍적으로 생성한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if headers:
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=header_row, column=col_idx, value=header)

    if data_rows:
        start_row = header_row + 1
        for row_offset, row_data in enumerate(data_rows):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=start_row + row_offset, column=col_idx, value=value)

    if merged_cells:
        for cell_range in merged_cells:
            ws.merge_cells(cell_range)

    if extra_sheets:
        for extra in extra_sheets:
            extra_ws = wb.create_sheet(title=extra.get("name", "Extra"))
            extra_headers = extra.get("headers", [])
            extra_header_row = extra.get("header_row", 1)
            for col_idx, header in enumerate(extra_headers, 1):
                extra_ws.cell(row=extra_header_row, column=col_idx, value=header)
            extra_data = extra.get("data_rows", [])
            for row_offset, row_data in enumerate(extra_data):
                for col_idx, value in enumerate(row_data, 1):
                    extra_ws.cell(
                        row=extra_header_row + 1 + row_offset,
                        column=col_idx,
                        value=value,
                    )

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


# ===== 테스트 클래스 =====


class TestExcelToCsvSingleSheet:
    """단일 시트 → CsvSheetData 변환 테스트."""

    def test_excel_to_csv_single_sheet(self):
        """단일 시트 Excel을 CsvSheetData로 변환한다."""
        data = _make_excel_bytes(
            headers=["서버명", "IP주소", "CPU사용률"],
            data_rows=[
                ["web-01", "10.0.0.1", 45.2],
                ["db-01", "10.0.1.1", 82.1],
            ],
        )

        result = excel_to_csv(data)

        assert len(result) == 1
        assert "Sheet1" in result

        sheet = result["Sheet1"]
        assert isinstance(sheet, CsvSheetData)
        assert sheet.sheet_name == "Sheet1"
        assert sheet.headers == ["서버명", "IP주소", "CPU사용률"]
        assert sheet.header_row_index == 1
        assert sheet.data_start_row == 2
        assert len(sheet.example_rows) == 2
        assert sheet.example_rows[0] == ["web-01", "10.0.0.1", "45.2"]
        assert sheet.example_rows[1] == ["db-01", "10.0.1.1", "82.1"]

        # csv_text에 헤더와 데이터가 모두 포함
        assert "서버명" in sheet.csv_text
        assert "web-01" in sheet.csv_text


class TestExcelToCsvMultiSheet:
    """멀티시트 → 시트별 CsvSheetData 변환 테스트."""

    def test_excel_to_csv_multi_sheet(self):
        """2개 이상 시트를 시트별 CsvSheetData dict로 변환한다."""
        data = _make_excel_bytes(
            headers=["서버명", "IP주소"],
            data_rows=[["web-01", "10.0.0.1"]],
            extra_sheets=[
                {
                    "name": "CPU메트릭",
                    "headers": ["호스트", "사용률"],
                    "data_rows": [["web-01", 85.0]],
                },
            ],
        )

        result = excel_to_csv(data)

        assert len(result) == 2
        assert "Sheet1" in result
        assert "CPU메트릭" in result

        sheet1 = result["Sheet1"]
        assert sheet1.headers == ["서버명", "IP주소"]

        sheet2 = result["CPU메트릭"]
        assert sheet2.headers == ["호스트", "사용률"]
        assert len(sheet2.example_rows) == 1

    def test_excel_to_csv_specific_sheet(self):
        """sheet_name 지정 시 해당 시트만 변환한다."""
        data = _make_excel_bytes(
            headers=["서버명", "IP"],
            extra_sheets=[
                {"name": "CPU", "headers": ["호스트", "사용률"]},
            ],
        )

        result = excel_to_csv(data, sheet_name="CPU")

        assert len(result) == 1
        assert "CPU" in result


class TestExcelToCsvWithExampleRows:
    """예시 데이터 행 추출 테스트."""

    def test_excel_to_csv_with_example_rows(self):
        """데이터 행이 있는 Excel에서 example_rows를 올바르게 추출한다."""
        rows = [[f"server-{i}", f"10.0.0.{i}", i * 10.5] for i in range(1, 6)]
        data = _make_excel_bytes(
            headers=["서버명", "IP", "CPU"],
            data_rows=rows,
        )

        result = excel_to_csv(data)
        sheet = result["Sheet1"]

        assert len(sheet.example_rows) == 5
        assert sheet.example_rows[0][0] == "server-1"

    def test_example_rows_max_50(self):
        """example_rows는 최대 50행으로 제한된다."""
        rows = [[f"server-{i}", f"10.0.0.{i}"] for i in range(100)]
        data = _make_excel_bytes(
            headers=["서버명", "IP"],
            data_rows=rows,
        )

        result = excel_to_csv(data)
        sheet = result["Sheet1"]

        assert len(sheet.example_rows) <= 50


class TestExcelToCsvEmptyTemplate:
    """빈 양식 (헤더만, 데이터 없음) 테스트."""

    def test_excel_to_csv_empty_template(self):
        """헤더만 있고 데이터가 없는 Excel → example_rows=[], csv_text에 헤더만."""
        data = _make_excel_bytes(
            headers=["서버명", "IP주소", "CPU사용률", "메모리사용률"],
        )

        result = excel_to_csv(data)
        sheet = result["Sheet1"]

        assert sheet.headers == ["서버명", "IP주소", "CPU사용률", "메모리사용률"]
        assert sheet.example_rows == []
        assert "서버명" in sheet.csv_text
        # csv_text에는 헤더만 존재 (데이터 행 없음)
        lines = sheet.csv_text.strip().splitlines()
        assert len(lines) == 1  # 헤더 행만


class TestExcelToCsvMergedCells:
    """병합 셀 포함 Excel 테스트."""

    def test_excel_to_csv_merged_cells(self):
        """병합 셀이 있는 Excel에서 첫 셀 값만 추출된다."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged"

        # 헤더: A1="카테고리", B1="서버명", C1="IP"
        ws.cell(row=1, column=1, value="카테고리")
        ws.cell(row=1, column=2, value="서버명")
        ws.cell(row=1, column=3, value="IP")

        # 데이터: A2:A3 병합, B2="web-01", C2="10.0.0.1"
        ws.cell(row=2, column=1, value="웹서버")
        ws.cell(row=2, column=2, value="web-01")
        ws.cell(row=2, column=3, value="10.0.0.1")
        ws.cell(row=3, column=2, value="web-02")
        ws.cell(row=3, column=3, value="10.0.0.2")
        ws.merge_cells("A2:A3")

        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        file_data = buf.getvalue()

        result = excel_to_csv(file_data)
        sheet = result["Merged"]

        assert sheet.headers == ["카테고리", "서버명", "IP"]
        # 병합된 첫 셀 값("웹서버")이 첫 행에 포함
        assert len(sheet.example_rows) >= 1
        assert sheet.example_rows[0][0] == "웹서버"


class TestExcelToCsvDateValues:
    """datetime 셀 → ISO 형식 문자열 변환 테스트."""

    def test_excel_to_csv_date_values(self):
        """datetime 셀이 ISO 형식 문자열로 변환된다."""
        dt = datetime(2026, 3, 23, 14, 30, 0)
        d = date(2026, 3, 23)

        data = _make_excel_bytes(
            headers=["서버명", "등록일시", "만료일"],
            data_rows=[
                ["web-01", dt, d],
            ],
        )

        result = excel_to_csv(data)
        sheet = result["Sheet1"]

        assert len(sheet.example_rows) == 1
        # datetime → ISO format
        assert sheet.example_rows[0][1] == "2026-03-23T14:30:00"
        # date → openpyxl은 date를 datetime으로 변환하므로 ISO datetime 형식
        assert sheet.example_rows[0][2] in ("2026-03-23", "2026-03-23T00:00:00")


class TestExcelToCsvFallbackNoHeader:
    """헤더 탐지 실패 → template_structure 폴백 테스트."""

    def test_excel_to_csv_fallback_no_header(self):
        """헤더를 탐지할 수 없는 시트에서 template_structure 폴백을 수행한다."""
        # 모든 행에 셀이 1개만 있는 Excel → 헤더 탐지 실패
        wb = Workbook()
        ws = wb.active
        ws.title = "NoHeader"
        ws.cell(row=1, column=1, value="제목만 있는 행")
        ws.cell(row=2, column=1, value="한 열만 있음")

        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        file_data = buf.getvalue()

        result = excel_to_csv(file_data)

        # 헤더 탐지 실패 시 폴백 경로로 진행하거나 빈 결과
        if "NoHeader" in result:
            sheet = result["NoHeader"]
            assert isinstance(sheet, CsvSheetData)
            # 폴백 시 example_rows는 빈 리스트
            assert sheet.example_rows == []
        else:
            # template_structure 폴백도 헤더를 찾지 못하면 빈 dict
            assert len(result) == 0


class TestExcelToCsvFallbackComplex:
    """복잡한 구조의 Excel에서 폴백 동작 테스트."""

    def test_excel_to_csv_fallback_complex(self):
        """복잡한 Excel 구조(정상 시트 + 비정상 시트 혼합)에서 폴백을 수행한다.

        정상 시트는 CSV 변환, 비정상 시트는 template_structure 폴백.
        """
        wb = Workbook()

        # 시트1: 정상 (헤더 2개 이상)
        ws1 = wb.active
        ws1.title = "정상시트"
        ws1.cell(row=1, column=1, value="서버명")
        ws1.cell(row=1, column=2, value="IP")
        ws1.cell(row=2, column=1, value="web-01")
        ws1.cell(row=2, column=2, value="10.0.0.1")

        # 시트2: 비정상 (단일 열만 존재 → 헤더 탐지 실패 → 폴백)
        ws2 = wb.create_sheet(title="비정상시트")
        ws2.cell(row=1, column=1, value="메모")

        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        file_data = buf.getvalue()

        result = excel_to_csv(file_data)

        # 정상 시트는 정상 변환
        assert "정상시트" in result
        normal_sheet = result["정상시트"]
        assert normal_sheet.headers == ["서버명", "IP"]
        assert len(normal_sheet.example_rows) == 1

        # 비정상 시트는 폴백 (template_structure 기반 또는 빈 결과)
        if "비정상시트" in result:
            fallback_sheet = result["비정상시트"]
            assert isinstance(fallback_sheet, CsvSheetData)
            assert fallback_sheet.example_rows == []


class TestCsvSheetDataStructure:
    """CsvSheetData 데이터클래스 필드 정합성 테스트."""

    def test_csv_sheet_data_structure(self):
        """CsvSheetData의 필드가 Plan 19 스펙과 일치하는지 확인한다."""
        expected_fields = {
            "sheet_name": str,
            "headers": list,
            "example_rows": list,
            "csv_text": str,
            "header_row_index": int,
            "data_start_row": int,
        }

        actual_fields = {f.name: f.type for f in dataclass_fields(CsvSheetData)}

        for name, expected_type in expected_fields.items():
            assert name in actual_fields, f"Missing field: {name}"

    def test_csv_sheet_data_instantiation(self):
        """CsvSheetData를 직접 생성하여 필드에 접근할 수 있다."""
        sd = CsvSheetData(
            sheet_name="Test",
            headers=["A", "B"],
            example_rows=[["1", "2"]],
            csv_text="A,B\n1,2\n",
            header_row_index=1,
            data_start_row=2,
        )
        assert sd.sheet_name == "Test"
        assert sd.headers == ["A", "B"]
        assert sd.example_rows == [["1", "2"]]
        assert sd.header_row_index == 1
        assert sd.data_start_row == 2


class TestFormatCellValue:
    """_format_cell_value 유틸 함수 테스트."""

    def test_none_returns_empty_string(self):
        assert _format_cell_value(None) == ""

    def test_datetime_returns_iso(self):
        assert _format_cell_value(datetime(2026, 1, 15, 10, 30)) == "2026-01-15T10:30:00"

    def test_date_returns_iso(self):
        assert _format_cell_value(date(2026, 1, 15)) == "2026-01-15"

    def test_time_returns_iso(self):
        assert _format_cell_value(time(14, 30, 0)) == "14:30:00"

    def test_number_returns_string(self):
        assert _format_cell_value(42.5) == "42.5"

    def test_string_passthrough(self):
        assert _format_cell_value("hello") == "hello"


class TestBuildCsvText:
    """_build_csv_text 유틸 함수 테스트."""

    def test_headers_only(self):
        text = _build_csv_text(["A", "B", "C"], [])
        assert text.strip() == "A,B,C"

    def test_headers_with_rows(self):
        text = _build_csv_text(["A", "B"], [["1", "2"], ["3", "4"]])
        lines = text.strip().splitlines()
        assert len(lines) == 3
        assert lines[0] == "A,B"
        assert lines[1] == "1,2"

    def test_empty_headers(self):
        text = _build_csv_text([], [])
        assert text.strip() == ""


class TestInvalidFile:
    """잘못된 파일 입력 테스트."""

    def test_invalid_bytes_raises_value_error(self):
        """유효하지 않은 바이트는 ValueError를 발생시킨다."""
        with pytest.raises(ValueError, match="Excel 파일을 읽을 수 없습니다"):
            excel_to_csv(b"this is not an excel file")
