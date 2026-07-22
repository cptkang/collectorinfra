"""Excel 양식 데이터 채우기 단위 테스트."""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook, load_workbook

from src.document.excel_parser import parse_excel_template
from src.document.excel_writer import fill_excel_template


def _create_template_bytes(
    headers: list[str],
    data_rows: list[list] | None = None,
    formulas: dict[str, str] | None = None,
) -> bytes:
    """테스트용 Excel 양식 파일을 생성한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)

    if data_rows:
        for row_offset, row_data in enumerate(data_rows):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=2 + row_offset, column=col_idx, value=value)

    if formulas:
        for cell_ref, formula in formulas.items():
            ws[cell_ref] = formula

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


class TestFillExcelTemplate:
    """fill_excel_template 함수 테스트."""

    def test_basic_data_fill(self):
        """기본 데이터 채우기."""
        template_bytes = _create_template_bytes(["서버명", "IP주소", "상태"])
        template_structure = parse_excel_template(template_bytes)
        column_mapping = {
            "서버명": "servers.hostname",
            "IP주소": "servers.ip_address",
            "상태": "servers.status",
        }
        rows = [
            {"hostname": "web-01", "ip_address": "10.0.0.1", "status": "active"},
            {"hostname": "web-02", "ip_address": "10.0.0.2", "status": "inactive"},
        ]

        result_bytes, total_filled = fill_excel_template(
            template_bytes, template_structure, column_mapping, rows
        )

        # 결과 검증
        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "web-01"
        assert ws.cell(row=2, column=2).value == "10.0.0.1"
        assert ws.cell(row=2, column=3).value == "active"
        assert ws.cell(row=3, column=1).value == "web-02"
        wb.close()

    def test_preserves_headers(self):
        """헤더 행이 보존된다."""
        template_bytes = _create_template_bytes(["서버명", "IP주소"])
        template_structure = parse_excel_template(template_bytes)
        column_mapping = {"서버명": "servers.hostname", "IP주소": "servers.ip_address"}
        rows = [{"hostname": "web-01", "ip_address": "10.0.0.1"}]

        result_bytes, _ = fill_excel_template(
            template_bytes, template_structure, column_mapping, rows
        )

        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "서버명"
        assert ws.cell(row=1, column=2).value == "IP주소"
        wb.close()

    def test_formula_cells_preserved(self):
        """수식 셀은 건너뛴다."""
        template_bytes = _create_template_bytes(
            ["이름", "값", "계산"],
            formulas={"C2": "=B2*2"},
        )
        template_structure = parse_excel_template(template_bytes)
        column_mapping = {
            "이름": "items.name",
            "값": "items.value",
            "계산": "items.calc",
        }
        rows = [{"name": "test", "value": 100, "calc": 999}]

        result_bytes, _ = fill_excel_template(
            template_bytes, template_structure, column_mapping, rows
        )

        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "test"
        assert ws.cell(row=2, column=2).value == 100
        # 수식 셀은 원래 수식이 보존됨
        assert ws["C2"].value == "=B2*2"
        wb.close()

    def test_unmapped_columns_ignored(self):
        """매핑되지 않은 컬럼은 비어있다."""
        template_bytes = _create_template_bytes(["서버명", "비고"])
        template_structure = parse_excel_template(template_bytes)
        column_mapping = {
            "서버명": "servers.hostname",
            "비고": None,  # 매핑 불가
        }
        rows = [{"hostname": "web-01"}]

        result_bytes, _ = fill_excel_template(
            template_bytes, template_structure, column_mapping, rows
        )

        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "web-01"
        # 매핑 없는 컬럼은 값이 없어야 함
        assert ws.cell(row=2, column=2).value is None
        wb.close()

    def test_empty_rows(self):
        """빈 결과 행 처리."""
        template_bytes = _create_template_bytes(["서버명", "IP"])
        template_structure = parse_excel_template(template_bytes)
        column_mapping = {"서버명": "servers.hostname", "IP": "servers.ip"}

        result_bytes, _ = fill_excel_template(
            template_bytes, template_structure, column_mapping, []
        )

        # 에러 없이 파일이 생성되어야 함
        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "서버명"
        wb.close()

    def test_column_name_case_insensitive(self):
        """컬럼명 대소문자 무시 매칭."""
        template_bytes = _create_template_bytes(["서버명", "IP"])
        template_structure = parse_excel_template(template_bytes)
        column_mapping = {"서버명": "servers.Hostname", "IP": "servers.ip"}
        rows = [{"hostname": "web-01", "ip": "10.0.0.1"}]

        result_bytes, _ = fill_excel_template(
            template_bytes, template_structure, column_mapping, rows
        )

        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "web-01"
        wb.close()

    def test_multiple_rows(self):
        """여러 행 데이터 채우기."""
        template_bytes = _create_template_bytes(["이름", "값"])
        template_structure = parse_excel_template(template_bytes)
        column_mapping = {"이름": "items.name", "값": "items.value"}
        rows = [
            {"name": f"item-{i}", "value": i * 10}
            for i in range(10)
        ]

        result_bytes, _ = fill_excel_template(
            template_bytes, template_structure, column_mapping, rows
        )

        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "item-0"
        assert ws.cell(row=11, column=1).value == "item-9"
        assert ws.cell(row=11, column=2).value == 90
        wb.close()

    def test_invalid_file_raises_error(self):
        """유효하지 않은 파일은 ValueError를 발생시킨다."""
        with pytest.raises(ValueError):
            fill_excel_template(
                b"invalid",
                {"sheets": []},
                {},
                [],
            )


class TestNumericStringNormalization:
    """DB2(b0) 경로에서 문자열로 도착한 통계값의 숫자 변환 (2026-07-16).

    버그: 은행존(DB2) 사용률 통계(CPU/메모리 평균·최고)가 DECIMAL→문자열("6.51")로 도착해
    엑셀에 텍스트로 저장됨(gp/yd PostgreSQL은 숫자로 도착해 정상 — 엔진 비대칭).
    수정: 사용률 통계 필드(classify_metric_field 판정)에 한해 숫자형 문자열을 숫자로 변환.
    OS 버전("7.9")·IP 같은 진짜 텍스트 필드는 변환하지 않는다.
    """

    def test_metric_string_value_becomes_number(self):
        """통계 헤더의 문자열 "6.51"은 float 6.51로 저장된다."""
        template_bytes = _create_template_bytes(["서버 이름", "CPU 평균", "메모리 최고"])
        template_structure = parse_excel_template(template_bytes)
        # 사용률 필드는 field_mapper가 미매핑(None)으로 넘긴다(D-066 후속3)
        column_mapping = {"서버 이름": "cmm_resource.name", "CPU 평균": None, "메모리 최고": None}
        rows = [{"name": "srv-01", "CPU 평균": "6.51", "메모리 최고": "85"}]

        result_bytes, _ = fill_excel_template(
            template_bytes, template_structure, column_mapping, rows
        )

        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active
        assert ws.cell(row=2, column=2).value == 6.51
        assert isinstance(ws.cell(row=2, column=2).value, float)
        # 정수형 문자열은 int로
        assert ws.cell(row=2, column=3).value == 85
        assert isinstance(ws.cell(row=2, column=3).value, int)
        wb.close()

    def test_non_metric_string_stays_text(self):
        """OS 버전 같은 비통계 필드의 숫자형 문자열("7.9")은 텍스트 유지."""
        template_bytes = _create_template_bytes(["서버 이름", "OS 버전", "IP주소"])
        template_structure = parse_excel_template(template_bytes)
        column_mapping = {
            "서버 이름": "cmm_resource.name",
            "OS 버전": "OSVerson",
            "IP주소": "cmm_resource.ipaddress",
        }
        rows = [{"name": "srv-01", "OSVerson": "7.9", "ipaddress": "10.1.2.3"}]

        result_bytes, _ = fill_excel_template(
            template_bytes, template_structure, column_mapping, rows
        )

        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "7.9"
        assert isinstance(ws.cell(row=2, column=2).value, str)
        assert ws.cell(row=2, column=3).value == "10.1.2.3"
        wb.close()

    def test_non_numeric_string_in_metric_field_stays_text(self):
        """통계 헤더라도 숫자로 파싱 불가한 값("N/A")은 원본 유지."""
        template_bytes = _create_template_bytes(["서버 이름", "CPU 평균"])
        template_structure = parse_excel_template(template_bytes)
        column_mapping = {"서버 이름": "cmm_resource.name", "CPU 평균": None}
        rows = [{"name": "srv-01", "CPU 평균": "N/A"}]

        result_bytes, _ = fill_excel_template(
            template_bytes, template_structure, column_mapping, rows
        )

        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "N/A"
        wb.close()
