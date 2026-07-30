"""Excel 양식 파서 단위 테스트."""

from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import Workbook

from src.document.excel_parser import parse_excel_template

_FORMS_DIR = Path(__file__).resolve().parent.parent / "fixtures" / "forms"


def _create_excel_bytes(
    headers: list[str] | None = None,
    data_rows: list[list] | None = None,
    sheet_name: str = "Sheet1",
    header_row: int = 1,
    merged_cells: list[str] | None = None,
    formulas: dict[str, str] | None = None,
    extra_sheets: list[dict] | None = None,
) -> bytes:
    """테스트용 Excel 파일 바이너리를 생성한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if headers:
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=header_row, column=col_idx, value=header)

    if data_rows:
        start_row = header_row + 1
        for row_offset, row_data in enumerate(data_rows):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=start_row + row_offset, column=col_idx, value=value)

    if merged_cells:
        for cell_range in merged_cells:
            ws.merge_cells(cell_range)

    if formulas:
        for cell_ref, formula in formulas.items():
            ws[cell_ref] = formula

    if extra_sheets:
        for extra in extra_sheets:
            extra_ws = wb.create_sheet(title=extra.get("name", "Extra"))
            extra_headers = extra.get("headers", [])
            for col_idx, header in enumerate(extra_headers, 1):
                extra_ws.cell(row=1, column=col_idx, value=header)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


class TestParseExcelTemplate:
    """parse_excel_template 함수 테스트."""

    def test_basic_single_sheet(self):
        """기본 단일 시트 파싱."""
        data = _create_excel_bytes(
            headers=["서버명", "IP주소", "CPU 사용률"],
            data_rows=[
                ["web-01", "10.0.0.1", 85.2],
                ["web-02", "10.0.0.2", 72.0],
            ],
        )

        result = parse_excel_template(data)

        assert result["file_type"] == "xlsx"
        assert len(result["sheets"]) == 1

        sheet = result["sheets"][0]
        assert sheet["name"] == "Sheet1"
        assert sheet["headers"] == ["서버명", "IP주소", "CPU 사용률"]
        assert sheet["header_row"] == 1
        assert sheet["data_start_row"] == 2
        assert sheet["max_column"] == 3

    def test_header_detection_non_first_row(self):
        """헤더가 첫 번째 행이 아닌 경우 탐지."""
        wb = Workbook()
        ws = wb.active
        # 1행: 빈 행, 2행: 제목 (1셀만), 3행: 헤더
        ws.cell(row=2, column=1, value="보고서 제목")
        ws.cell(row=3, column=1, value="서버명")
        ws.cell(row=3, column=2, value="IP주소")
        ws.cell(row=3, column=3, value="상태")

        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        result = parse_excel_template(data)
        sheet = result["sheets"][0]
        assert sheet["header_row"] == 3
        assert sheet["data_start_row"] == 4
        assert "서버명" in sheet["headers"]

    def test_empty_sheet_skipped(self):
        """빈 시트는 결과에 포함되지 않는다."""
        wb = Workbook()
        wb.active.title = "Empty"
        buf = io.BytesIO()
        wb.save(buf)

        result = parse_excel_template(buf.getvalue())
        assert result["sheets"] == []

    def test_merged_cells_detected(self):
        """병합 셀 정보가 수집된다."""
        data = _create_excel_bytes(
            headers=["서버명", "IP", "상태"],
            merged_cells=["A1:B1"],
        )

        result = parse_excel_template(data)
        sheet = result["sheets"][0]
        assert len(sheet["merged_cells"]) > 0

    def test_formula_cells_detected(self):
        """수식 셀이 탐지된다."""
        data = _create_excel_bytes(
            headers=["항목", "값", "합계"],
            data_rows=[["a", 100, None]],
            formulas={"C2": "=SUM(B2:B100)"},
        )

        result = parse_excel_template(data)
        sheet = result["sheets"][0]
        assert "C2" in sheet["formula_cells"]

    def test_multi_sheet(self):
        """다중 시트 파싱."""
        data = _create_excel_bytes(
            headers=["서버명", "IP"],
            extra_sheets=[
                {"name": "CPU", "headers": ["서버", "사용률"]},
            ],
        )

        result = parse_excel_template(data)
        assert len(result["sheets"]) == 2
        sheet_names = [s["name"] for s in result["sheets"]]
        assert "Sheet1" in sheet_names
        assert "CPU" in sheet_names

    def test_invalid_file_raises_error(self):
        """유효하지 않은 파일은 ValueError를 발생시킨다."""
        with pytest.raises(ValueError, match="Excel 파일을 읽을 수 없습니다"):
            parse_excel_template(b"not a valid excel file")

    def test_data_end_row_detection(self):
        """데이터 영역 끝 행이 올바르게 탐지된다."""
        data = _create_excel_bytes(
            headers=["이름", "값"],
            data_rows=[
                ["a", 1],
                ["b", 2],
                ["c", 3],
            ],
        )

        result = parse_excel_template(data)
        sheet = result["sheets"][0]
        assert sheet["data_end_row"] == 4  # row 2, 3, 4

    def test_header_cells_structure(self):
        """header_cells에 col과 value 정보가 있다."""
        data = _create_excel_bytes(headers=["이름", "나이"])

        result = parse_excel_template(data)
        sheet = result["sheets"][0]
        assert len(sheet["header_cells"]) == 2
        assert sheet["header_cells"][0]["col"] == 1
        assert sheet["header_cells"][0]["value"] == "이름"

    def test_single_header_block_rows(self):
        """단일 헤더는 header_block_rows가 그 행 하나다."""
        data = _create_excel_bytes(headers=["서버명", "IP"])

        sheet = parse_excel_template(data)["sheets"][0]
        assert sheet["header_block_rows"] == [1]


class TestMultiRowHeaderBlock:
    """2행 병합 헤더 블록 결합 테스트 (D-112)."""

    def _parse_form(self, filename: str) -> dict:
        data = (_FORMS_DIR / filename).read_bytes()
        result = parse_excel_template(data)
        assert len(result["sheets"]) == 1
        return result["sheets"][0]

    def test_cpu_form_composite_headers(self):
        """금감원 CPU 양식: 그룹+서브 헤더가 복합 필드명으로 결합된다."""
        sheet = self._parse_form("CPU_양식.xlsx")

        assert sheet["header_block_rows"] == [5, 6]
        assert sheet["header_row"] == 6  # 블록 하단 — data_start_row 계약 유지
        assert sheet["data_start_row"] == 7
        assert sheet["max_column"] == 17

        headers = sheet["headers"]
        # 세로 병합(5:6)은 단독 값
        assert "제조사(모델명)" in headers
        assert "호스트명" in headers
        assert "비고" in headers
        # 가로 병합 그룹 + 서브 헤더 결합
        assert "월중평균사용률(최근 6개월간)|M" in headers
        assert "월중평균사용률(최근 6개월간)|M+5" in headers
        assert "월중 Peak시 사용률(최근 6개월간)|M" in headers
        assert "월중 Peak시 사용률(최근 6개월간)|M+5" in headers
        # 비병합 상하 값 결합
        assert "구분|분류" in headers
        assert "처리능력|(TPMC)" in headers

    def test_cpu_form_column_coordinates(self):
        """복합 필드명의 열 좌표가 정확하다 (E=5 평균 M, K=11 Peak M)."""
        sheet = self._parse_form("CPU_양식.xlsx")
        col_of = {hc["value"]: hc["col"] for hc in sheet["header_cells"]}

        assert col_of["호스트명"] == 3
        assert col_of["월중평균사용률(최근 6개월간)|M"] == 5
        assert col_of["월중평균사용률(최근 6개월간)|M+5"] == 10
        assert col_of["월중 Peak시 사용률(최근 6개월간)|M"] == 11
        assert col_of["월중 Peak시 사용률(최근 6개월간)|M+5"] == 16
        assert col_of["비고"] == 17

    def test_memory_form_composite_headers(self):
        """금감원 메모리 양식: 처리능력 단위가 (GB)로 결합된다."""
        sheet = self._parse_form("메모리_양식.xlsx")

        assert sheet["header_block_rows"] == [5, 6]
        assert sheet["data_start_row"] == 7
        assert "처리능력|(GB)" in sheet["headers"]
        assert "월중평균사용률(최근 6개월간)|M+2" in sheet["headers"]

    def test_server_list_form_composite_headers(self):
        """금감원 서버 목록 양식: 서브 헤더(설치장소 등)가 유실되지 않는다."""
        sheet = self._parse_form("서버_목록_리스트_양식.xlsx")

        assert sheet["header_block_rows"] == [5, 6]
        assert sheet["data_start_row"] == 7
        assert sheet["max_column"] == 12

        headers = sheet["headers"]
        assert "서버명" in headers
        assert "OS(버전정보포함)" in headers
        assert "IP" in headers
        # 가로 병합 그룹 아래 서브 헤더 (줄바꿈은 공백으로 정규화)
        assert "서버위치|설치장소 (주센터, 재해복구센터 등)" in headers
        assert "서버위치|설치구간 (DMZ, 내부망, 개발망 등)" in headers
        assert "접근통제 및 추가인증|적용 솔루션명" in headers
        assert "접근통제 및 추가인증|추가인증수단" in headers
        # 상단 행이 빈 열(L)은 하단 값 단독
        assert "비고" in headers

    def test_no_combine_without_merge_evidence(self):
        """병합 증거가 없으면 인접 행(제목 등)과 결합하지 않는다."""
        wb = Workbook()
        ws = wb.active
        # 4행: 여러 셀로 쪼개진 소제목 (서버 목록 양식 4행 패턴)
        ws.cell(row=4, column=1, value="나. 전체 서버 현황")
        ws.cell(row=4, column=4, value="(작성자성명, 직함, 내선번호)")
        # 5행: 실제 헤더 (병합 없음)
        for col, name in enumerate(["서버명", "IP", "OS", "상태"], 1):
            ws.cell(row=5, column=col, value=name)
        buf = io.BytesIO()
        wb.save(buf)

        sheet = parse_excel_template(buf.getvalue())["sheets"][0]
        assert sheet["header_block_rows"] == [5]
        assert sheet["headers"] == ["서버명", "IP", "OS", "상태"]
        assert sheet["data_start_row"] == 6

    def test_no_combine_with_data_row_below(self):
        """헤더 바로 아래가 데이터 행(수치 지배)이면 결합하지 않는다."""
        wb = Workbook()
        ws = wb.active
        # 세로 병합이 헤더-데이터 행에 걸쳐 있어도 수치 게이트로 차단
        for col, name in enumerate(["서버명", "CPU", "메모리"], 1):
            ws.cell(row=1, column=col, value=name)
        ws.cell(row=2, column=1, value="web-01")
        ws.cell(row=2, column=2, value=85.2)
        ws.cell(row=2, column=3, value=72.0)
        ws.merge_cells("A1:A2")  # 인위적 세로 병합
        buf = io.BytesIO()
        wb.save(buf)

        sheet = parse_excel_template(buf.getvalue())["sheets"][0]
        assert sheet["header_block_rows"] == [1]
        assert sheet["data_start_row"] == 2

    def test_single_header_with_horizontal_merge_unchanged(self):
        """헤더에 가로 병합이 있어도 그 아래 서브 값이 2개 미만이면 결합하지 않는다."""
        data = _create_excel_bytes(
            headers=["서버명", "IP", "상태"],  # A1:B1 병합으로 IP는 소거됨
            data_rows=[["web-01", None, "정상"]],
            merged_cells=["A1:B1"],
        )

        sheet = parse_excel_template(data)["sheets"][0]
        assert sheet["header_block_rows"] == [1]
        assert sheet["headers"] == ["서버명", "상태"]
        assert sheet["data_start_row"] == 2
