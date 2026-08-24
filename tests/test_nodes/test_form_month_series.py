"""금감원 취합자료 월 시리즈(M~M+5) 폼필 파이프라인 테스트 (plans/72 Phase 3, D-146/D-147/D-148).

인식기(순수 함수) → 단일/멀티 SQL 경로 대칭 배선 → writer 채움 → 응답 사유 노출까지
결정적 경로 전체를 고정한다. LLM은 어디에도 개입하지 않아야 한다.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from types import SimpleNamespace

from src.db_adapters.polestar.assembler import (
    apply_capacity_scope_rule,
    find_vendor_model_concat,
    recognize_month_series,
)

_FORMS_DIR = Path(__file__).resolve().parent.parent / "fixtures" / "forms"

_TODAY = date(2026, 7, 28)  # 결정적 기준일 — M+5(마지막 완결 월)=202606

_AVG = "월중평균사용률(최근 6개월간)"
_PEAK = "월중 Peak시 사용률(최근 6개월간)"


def _cpu_form_mapping() -> dict:
    mapping: dict = {
        "구분|분류": None,
        "제조사(모델명)": "EAV:Model",
        "호스트명": "cmm_resource.hostname",
        "처리능력|(TPMC)": None,
        "비고": None,
    }
    for k in range(6):
        sub = "M" if k == 0 else f"M+{k}"
        mapping[f"{_AVG}|{sub}"] = None
        mapping[f"{_PEAK}|{sub}"] = None
    return mapping


_EAV_META = {
    "_structure_meta": {
        "patterns": [
            {
                "type": "eav",
                "entity_table": "cmm_resource",
                "config_table": "core_config_prop",
                "attribute_column": "name",
                "value_column": "stringvalue_short",
                "direct_join": {
                    "entity_column": "resource_conf_id",
                    "config_column": "configuration_id",
                },
                "known_attributes": [
                    {"name": "Vendor", "description": "서버 제조사 [resource_type: server.Server]"},
                    {"name": "Model", "description": "모델명 [resource_type: server.Server]"},
                    # 실 프로필과 동일한 대문자 충돌(server.Cpus의 MODEL/VENDOR) 재현 —
                    # eav_attr_resource_types의 upper 키는 후자로 덮인다
                    {"name": "MODEL", "description": "CPU 모델 [resource_type: server.Cpus]"},
                    {"name": "VENDOR", "description": "CPU 제조사 [resource_type: server.Cpus]"},
                    {"name": "TotalSize", "description": "메모리 용량 [resource_type: server.Memory]"},
                ],
            }
        ]
    }
}


class TestRecognizeMonthSeries:
    """월 시리즈 인식기(D-146) — 구조 패턴 기반, 기관명·시트제목 하드코딩 없음."""

    def test_cpu_form_relative_months(self):
        ms = recognize_month_series(
            _cpu_form_mapping(),
            context_text="16. IT장비 및 통신회선 사용현황 나. 주요 업무 CPU 사용현황",
            user_query="양식 채워줘",
            today=_TODAY,
        )
        assert ms is not None
        assert ms.resource_type == "server.Cpus"
        assert len(ms.measures) == 12
        # 기본 앵커: M+5 = 실행일 기준 지난달(202606) → M=202601 (Q3 확정)
        assert ms.anchor == ("202601", "202606")
        assert ms.month_by_field[f"{_AVG}|M"] == "202601"
        assert ms.month_by_field[f"{_PEAK}|M+5"] == "202606"
        # alias는 복합 필드명 그대로(행 키=양식 헤더 아키텍처), 평균→avg_val, peak→max_val
        by_alias = {m[0]: m for m in ms.measures}
        assert by_alias[f"{_AVG}|M+2"][1:] == ("server.Cpus", "avg_val", "202603")
        assert by_alias[f"{_PEAK}|M+2"][1:] == ("server.Cpus", "max_val", "202603")

    def test_user_specified_anchor_month_is_end(self):
        """사용자가 월을 명시하면 그 월이 M+max(끝 월)가 된다(Q3: 기준월=M+5)."""
        ms = recognize_month_series(
            _cpu_form_mapping(),
            context_text="CPU 사용현황",
            user_query="2026년 3월 기준으로 채워줘",
            today=_TODAY,
        )
        assert ms.anchor == ("202510", "202603")
        assert ms.month_by_field[f"{_AVG}|M+5"] == "202603"

    def test_memory_context_noun(self):
        """'주기억장치'(관용 표현)로 메모리 리소스를 판정한다."""
        ms = recognize_month_series(
            _cpu_form_mapping(),
            context_text="다. 주요 업무 주기억장치 사용현황",
            today=_TODAY,
        )
        assert ms.resource_type == "server.Memory"

    def test_no_resource_noun_falls_back(self):
        """리소스 명사가 어디에도 없으면 판정 불가 → None(기존 경로 폴백)."""
        assert recognize_month_series(
            _cpu_form_mapping(), context_text="사용현황", user_query="양식 채워줘",
            today=_TODAY,
        ) is None

    def test_absolute_months_without_year(self):
        """'1월'~'3월' 절대 표기(연도 미상)는 마지막 완결 월 이하 최근 발생으로 보정."""
        mapping = {f"{_AVG}|{m}월": None for m in (1, 2, 3)}
        ms = recognize_month_series(mapping, context_text="CPU", today=_TODAY)
        assert ms.anchor == ("202601", "202603")

    def test_absolute_future_month_wraps_to_previous_year(self):
        """실행일(7월) 기준 '12월'은 아직 안 온 달 → 작년 12월로 보정."""
        mapping = {f"{_AVG}|12월": None, f"{_AVG}|1월": None}
        ms = recognize_month_series(mapping, context_text="CPU", today=_TODAY)
        assert ms.month_by_field[f"{_AVG}|12월"] == "202512"
        assert ms.month_by_field[f"{_AVG}|1월"] == "202601"

    def test_mixed_relative_absolute_rejected(self):
        mapping = {f"{_AVG}|M": None, f"{_AVG}|3월": None}
        assert recognize_month_series(mapping, context_text="CPU", today=_TODAY) is None

    def test_mapped_fields_not_consumed(self):
        """이미 정상 매핑된 필드는 인식 대상이 아니다(cmm_metric_stat 오매핑은 대상)."""
        mapping = {
            f"{_AVG}|M": "cmm_metric_stat_m.avg_val",  # LLM 오매핑 → 인식 대상
            f"{_AVG}|M+1": None,
        }
        ms = recognize_month_series(mapping, context_text="CPU", today=_TODAY)
        assert ms is not None and len(ms.measures) == 2

    def test_overlong_alias_rejected(self):
        """PG 식별자 63바이트 초과 필드명은 잘려 월 서픽스가 소실되므로 양식 전체 폴백."""
        long_group = "월중평균사용률" * 4  # 84바이트
        mapping = {f"{long_group}|M": None, f"{long_group}|M+1": None}
        assert recognize_month_series(mapping, context_text="CPU", today=_TODAY) is None

    def test_non_month_form_ignored(self):
        """서버 목록 양식(월 서브헤더 없음)은 발동하지 않는다."""
        mapping = {
            "서버위치|설치장소 (주센터, 재해복구센터 등)": None,
            "접근통제 및 추가인증|적용 솔루션명": None,
            "서버명": "cmm_resource.name",
        }
        assert recognize_month_series(mapping, context_text="18. DB 및 서버 운영현황", today=_TODAY) is None


class TestCapacityScopeRule:
    """'처리능력|(GB)' 요청 스코프 규칙(Q1/D-148) — 유사어 등록 없이 3중 문맥으로만."""

    _ATTR_RT = {"TOTALSIZE": "server.Memory", "MODEL": "server.Server"}

    def test_gb_capacity_mapped_for_memory_form(self):
        updates = apply_capacity_scope_rule(
            {"처리능력|(GB)": None, "비고": None}, self._ATTR_RT, "server.Memory"
        )
        assert updates == {"처리능력|(GB)": "EAV:TotalSize"}

    def test_tpmc_forced_blank(self):
        """CPU 양식의 (TPMC)는 오염 매핑이 있어도 강제 None(공란 보장) — 라이브 4차 실측:
        field_mapper 캐시가 '처리능력'을 TotalSize로 매핑해 TPMC에 메모리 용량이 채워짐."""
        updates = apply_capacity_scope_rule(
            {"처리능력|(TPMC)": "EAV:TotalSize"}, self._ATTR_RT, "server.Cpus"
        )
        assert updates == {"처리능력|(TPMC)": None}

    def test_requires_profile_attribute(self):
        """프로필에 TotalSize가 없으면 용량 매핑 대신 강제 공란(프로필 게이트)."""
        assert apply_capacity_scope_rule(
            {"처리능력|(GB)": None}, {}, "server.Memory"
        ) == {"처리능력|(GB)": None}


class TestVendorModelConcat:
    """'제조사(모델명)' Vendor+Model 결합 규칙(D-148) — 라이브 실측 반쪽 매핑 교정."""

    def test_rule_picks_server_scoped_exact_case_names(self):
        """server.Cpus의 MODEL/VENDOR(대문자 충돌)가 아니라 server.Server의
        Vendor/Model 정확한 이름을 고른다."""
        rules = find_vendor_model_concat(_EAV_META, {"제조사(모델명)": "EAV:Model"})
        assert rules == [("제조사(모델명)", "Vendor", "Model")]

    def test_rule_requires_both_attrs(self):
        meta = {"_structure_meta": {"patterns": [{
            "type": "eav",
            "known_attributes": [
                {"name": "Vendor", "description": "제조사 [resource_type: server.Server]"},
            ],
        }]}}
        assert find_vendor_model_concat(meta, {"제조사(모델명)": None}) == []

    def test_single_path_sql_has_concat_alias_once(self):
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        state = {
            "column_mapping": _cpu_form_mapping(),
            "schema_info": _EAV_META,
            "template_structure": {"sheets": [{"title_text": "나. 주요 업무 CPU 사용현황", "name": "Sheet1"}]},
            "active_db_engine": "postgresql",
            "active_db_id": "",
        }
        result = _try_build_form_fill_pivot_sql(state, 100_000, "2026년 6월 기준 CPU 양식 채워줘")
        sql = result["sql"]
        # Vendor(Model) 결합 라인 — alias는 정확히 1회(단독 매핑과 중복 없음)
        assert sql.count('AS "제조사(모델명)"') == 1
        assert "cc.name='Vendor'" in sql
        assert "NULLIF(COALESCE(" in sql
        assert "'(' ||" in sql


class TestSinglePathWiring:
    """단일 경로(query_generator._try_build_form_fill_pivot_sql) 배선."""

    def _state(self, mapping: dict, title: str) -> dict:
        return {
            "column_mapping": mapping,
            "schema_info": _EAV_META,
            "template_structure": {"sheets": [{"title_text": title, "name": "Sheet1"}]},
            "active_db_engine": "postgresql",
            "active_db_id": "",
        }

    def test_cpu_form_builds_month_pivot_sql(self):
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        state = self._state(_cpu_form_mapping(), "나. 주요 업무 CPU 사용현황")
        result = _try_build_form_fill_pivot_sql(
            state, 100_000, "2026년 6월 기준 CPU 양식 채워줘"
        )
        assert result is not None
        sql = result["sql"]
        # 월별 CASE 피벗 + alias=복합 필드명
        assert "s.stat_date='202603'" in sql
        assert f'AS "{_AVG}|M+2"' in sql
        assert f'AS "{_PEAK}|M+5"' in sql
        # 기존 필드도 통합 피벗에 포함
        assert 'AS "호스트명"' in sql
        assert "cc.name='Model'" in sql
        # 서버당 1행 계약 유지
        assert "GROUP BY COALESCE(c.platform_resource_id, c.id)" in sql
        anchor = result["month_anchor"]
        assert anchor["start"] == "202601" and anchor["end"] == "202606"
        assert len(anchor["fields"]) == 12
        # 월 필드는 state 매핑 강제 None(writer 필드명 조회 — N:1 역매핑 동일값 방지)
        updates = result["mapping_updates"]
        assert updates[f"{_AVG}|M+2"] is None
        assert updates[f"{_PEAK}|M+5"] is None
        # 비고 규칙(사용자 확정): SQL은 등록명, writer 매핑은 None
        assert 'THEN c.name END) AS "비고"' in sql
        assert updates["비고"] is None

    def test_memory_form_applies_capacity_rule(self):
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        mapping = _cpu_form_mapping()
        del mapping["처리능력|(TPMC)"]
        mapping["처리능력|(GB)"] = None
        state = self._state(mapping, "다. 주요 업무 주기억장치 사용현황")
        result = _try_build_form_fill_pivot_sql(
            state, 100_000, "2026년 6월 기준 메모리 양식 채워줘"
        )
        assert result is not None
        assert result["mapping_updates"]["처리능력|(GB)"] == "EAV:TotalSize"
        assert result["mapping_updates"][f"{_AVG}|M"] is None
        sql = result["sql"]
        assert "cc.name='TotalSize'" in sql
        assert "c.resource_type='server.Memory' AND s.definition_name" in sql

    def test_non_form_query_unchanged(self):
        """양식 업로드가 아니면(월 시리즈·자식 EAV도 없음) 기존과 동일하게 None(LLM 경로 유지)."""
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        state = self._state({"서버명": "cmm_resource.name"}, "서버 목록")
        state["template_structure"] = None  # 파일 없는 일반 질의
        assert _try_build_form_fill_pivot_sql(state, 1000, "서버 목록 조회") is None


class TestD116FormIntentGate:
    """D-149 게이트 확장 — 양식 업로드는 월 시리즈·자식 EAV 없어도 항상 결정적 조립.

    라이브 실측(2026-07-30): 단순 양식(서버 이름·IP·OS·코어·메모리)이 CM DB들에서 LLM
    폴백으로 떨어져 `column "r.name" must appear in the GROUP BY clause`로 전멸.
    결정적 조립기는 별칭 c + 전 SELECT 집계 + GROUP BY COALESCE라 구조적으로 불가능한
    에러 — 경로 선택이 per-DB LLM 매핑에 종속된 것이 근본 원인이며 게이트로 고정한다.
    """

    def _state(self, mapping: dict, schema_info: dict | None = None) -> dict:
        return {
            "column_mapping": mapping,
            "schema_info": _EAV_META if schema_info is None else schema_info,
            "template_structure": {"sheets": [{"title_text": "서버 목록", "name": "Sheet1"}]},
            "active_db_engine": "postgresql",
            "active_db_id": "",
        }

    def test_regular_only_form_builds_pivot(self):
        """직접 칼럼 매핑만 있는 양식(월·자식 EAV 없음) → 결정적 피벗(기존엔 None→LLM)."""
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        state = self._state({
            "서버 이름": "cmm_resource.name",
            "IP 주소": "cmm_resource.ipaddress",
        })
        result = _try_build_form_fill_pivot_sql(state, 1000, "서버 목록 양식 채워줘")
        assert result is not None
        sql = result["sql"]
        assert 'AS "서버 이름"' in sql and 'AS "IP 주소"' in sql
        # 서버당 1행 계약 — GROUP BY 에러류가 구조적으로 불가능한 형태
        assert "GROUP BY COALESCE(c.platform_resource_id, c.id)" in sql
        assert "FROM cmm_resource c" in sql
        assert " r." not in sql  # LLM 폴백 예시 별칭 부재 증명

    def test_form_without_eav_pattern_returns_none(self):
        """eav_pattern 부재 DB(비폴스타)는 form_intent여도 발동하지 않는다(현행 LLM 경로)."""
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        state = self._state({"서버 이름": "cmm_resource.name"}, schema_info={})
        assert _try_build_form_fill_pivot_sql(state, 1000, "양식 채워줘") is None

    def test_form_all_unmapped_injects_identity(self):
        """매핑이 전무해도 폼필이면 식별 컬럼 주입으로 결정적 조립(빈 SELECT 방지)."""
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        state = self._state({"구분": None, "용도": None})
        result = _try_build_form_fill_pivot_sql(state, 1000, "양식 채워줘")
        assert result is not None
        assert 'AS "server_name"' in result["sql"]
        assert 'AS "hostname"' in result["sql"]

    def test_simple_form_fixture_end_to_end(self):
        """단순 양식(라이브 실측 케이스) 파서→매핑→조립 e2e — 5필드 전부 결정적 SELECT."""
        import io

        from openpyxl import Workbook

        from src.document.excel_parser import parse_excel_template
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        wb = Workbook()
        ws = wb.active
        ws.title = "서버현황"
        ws.append(["서버 이름", "IP 주소", "OS 버전", "CPU 코어 수", "메모리 용량"])
        ws.append(["", "", "", "", ""])
        buf = io.BytesIO()
        wb.save(buf)
        template = parse_excel_template(buf.getvalue())
        assert template["sheets"], "단일 헤더 양식 파싱 실패"

        meta = {
            "_structure_meta": {
                "patterns": [{
                    "type": "eav",
                    "entity_table": "cmm_resource",
                    "config_table": "core_config_prop",
                    "attribute_column": "name",
                    "value_column": "stringvalue_short",
                    "direct_join": {
                        "entity_column": "resource_conf_id",
                        "config_column": "configuration_id",
                    },
                    "known_attributes": [
                        {"name": "OSVerson", "description": "OS 버전 [resource_type: server.Server]"},
                        {"name": "LOGICALCORE", "description": "논리 코어 [resource_type: server.Cpus]"},
                        {"name": "TotalSize", "description": "메모리 용량 [resource_type: server.Memory]"},
                    ],
                }]
            }
        }
        state = {
            "column_mapping": {
                "서버 이름": "cmm_resource.name",
                "IP 주소": "cmm_resource.ipaddress",
                "OS 버전": "EAV:OSVerson",
                "CPU 코어 수": "EAV:LOGICALCORE",
                "메모리 용량": "EAV:TotalSize",
            },
            "schema_info": meta,
            "template_structure": template,
            "active_db_engine": "postgresql",
            "active_db_id": "",
        }
        result = _try_build_form_fill_pivot_sql(state, 1000, "서버 양식 채워줘")
        assert result is not None
        sql = result["sql"]
        for field in ["서버 이름", "IP 주소", "OS 버전", "CPU 코어 수", "메모리 용량"]:
            assert f'AS "{field}"' in sql
        assert "'server.Cpus'" in sql and "'server.Memory'" in sql
        assert "GROUP BY COALESCE(c.platform_resource_id, c.id)" in sql

    async def test_multi_path_form_intent_symmetric(self):
        """멀티 경로 대칭 — form_intent + eav_pattern이면 LLM에 도달하지 않는다."""
        from src.nodes.multi_db_executor import _generate_sql

        sql = await _generate_sql(
            llm=None,  # 결정적 경로 — LLM에 도달하면 안 된다
            parsed_requirements={"original_query": "서버 목록 양식 채워줘"},
            schema_info={
                "tables": {"cmm_resource": {"columns": [
                    {"name": "name", "type": "text"},
                    {"name": "ipaddress", "type": "text"},
                    {"name": "hostname", "type": "text"},
                ]}},
                **_EAV_META,
            },
            sub_query_context="서버 목록 양식 채우기",
            default_limit=1000,
            column_mapping={"서버 이름": "cmm_resource.name", "IP 주소": "cmm_resource.ipaddress"},
            db_engine="postgresql",
            db_id="",
            unmapped_fields=[],
            app_config=SimpleNamespace(
                text2sql=SimpleNamespace(semantic_compose=False, generic_llm_mapping=False),
                get_polestar_db_ids=lambda: set(),
            ),
            form_context_text="서버 목록",
            form_fill_out={},
            form_intent=True,
            mapping_sources={},
        )
        assert 'AS "서버 이름"' in sql and 'AS "IP 주소"' in sql
        assert "GROUP BY COALESCE(c.platform_resource_id, c.id)" in sql
        assert " r." not in sql


class TestLlmInferredDemotion:
    """D-149 — 폼필에서 llm_inferred 매핑 채움 금지(침묵 오염 → 공란+역질문 후보)."""

    def _state(self, mapping: dict, sources: dict) -> dict:
        return {
            "column_mapping": mapping,
            "mapping_sources": sources,
            "schema_info": _EAV_META,
            "template_structure": {"sheets": [{"title_text": "서버 목록", "name": "Sheet1"}]},
            "active_db_engine": "postgresql",
            "active_db_id": "",
        }

    def test_inferred_dropped_synonym_kept(self):
        """llm_inferred는 제외(화이트리스트 통과 칼럼이어도), synonym 출처는 채움 유지."""
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        state = self._state(
            {"서버 이름": "cmm_resource.name", "도입일자": "cmm_resource.description"},
            {"서버 이름": "synonym", "도입일자": "llm_inferred"},
        )
        result = _try_build_form_fill_pivot_sql(state, 1000, "양식 채워줘")
        assert result is not None
        sql = result["sql"]
        assert 'AS "서버 이름"' in sql
        assert "도입일자" not in sql  # 라이브 실측: epoch 원값 오염의 원천 차단
        # writer 역조회 차단 — state 매핑 강제 None(엄격 필드명 조회 → 공란)
        assert result["mapping_updates"]["도입일자"] is None

    def test_inferred_metric_field_recovered_by_name(self):
        """llm_inferred 사용률류(집계어 명시)는 매핑 값을 버리되 필드명 기반 피벗으로 회수.

        집계어 없는 bare '사용률'은 평균/피크가 모호해 회수하지 않는다(공란+역질문 후보 —
        classify_metric_field의 보수 계약).
        """
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        state = self._state(
            {
                "서버 이름": "cmm_resource.name",
                "CPU 평균 사용률": "cmm_metric_stat_m.avg_val",
                "CPU 사용률": "cmm_metric_stat_m.avg_val",
            },
            {
                "서버 이름": "synonym",
                "CPU 평균 사용률": "llm_inferred",
                "CPU 사용률": "llm_inferred",
            },
        )
        result = _try_build_form_fill_pivot_sql(state, 1000, "양식 채워줘")
        assert result is not None
        sql = result["sql"]
        assert 'AS "CPU 평균 사용률"' in sql  # 집계어 명시 → 결정적 회수
        assert "'Utilization'" in sql
        assert 'AS "CPU 사용률"' not in sql  # 집계 모호 → 공란(역질문 후보)
        assert result["mapping_updates"]["CPU 사용률"] is None

    def test_no_form_intent_no_demotion(self):
        """양식 턴이 아니면 강등하지 않는다(일반 질의 경로 무영향)."""
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        state = self._state(
            {"코어수": "EAV:LOGICALCORE"},
            {"코어수": "llm_inferred"},
        )
        state["template_structure"] = None
        state["schema_info"] = {
            "_structure_meta": {"patterns": [{
                "type": "eav",
                "entity_table": "cmm_resource",
                "config_table": "core_config_prop",
                "attribute_column": "name",
                "value_column": "stringvalue_short",
                "direct_join": {
                    "entity_column": "resource_conf_id",
                    "config_column": "configuration_id",
                },
                "known_attributes": [
                    {"name": "LOGICALCORE", "description": "논리 코어 [resource_type: server.Cpus]"},
                ],
            }]}
        }
        result = _try_build_form_fill_pivot_sql(state, 1000, "코어수 조회")
        # 자식 EAV 존재 → 기존 게이트로 조립되며, 비폼필이라 llm_inferred도 유지
        assert result is not None
        assert "cc.name='LOGICALCORE'" in result["sql"]

    async def test_multi_path_demotion_symmetric(self):
        """멀티 경로 대칭 — 강등 + form_fill_out.mapping_updates 강제 None."""
        from src.nodes.multi_db_executor import _generate_sql

        form_fill_out: dict = {}
        sql = await _generate_sql(
            llm=None,
            parsed_requirements={"original_query": "서버 목록 양식 채워줘"},
            schema_info={
                "tables": {"cmm_resource": {"columns": [
                    {"name": "name", "type": "text"},
                    {"name": "description", "type": "text"},
                    {"name": "hostname", "type": "text"},
                ]}},
                **_EAV_META,
            },
            sub_query_context="서버 목록 양식 채우기",
            default_limit=1000,
            column_mapping={"서버 이름": "cmm_resource.name", "도입일자": "cmm_resource.description"},
            db_engine="postgresql",
            db_id="",
            unmapped_fields=[],
            app_config=SimpleNamespace(
                text2sql=SimpleNamespace(semantic_compose=False, generic_llm_mapping=False),
                get_polestar_db_ids=lambda: set(),
            ),
            form_context_text="서버 목록",
            form_fill_out=form_fill_out,
            form_intent=True,
            mapping_sources={"서버 이름": "synonym", "도입일자": "llm_inferred"},
        )
        assert 'AS "서버 이름"' in sql
        assert "도입일자" not in sql
        assert form_fill_out["mapping_updates"]["도입일자"] is None


class TestMultiPathWiring:
    """멀티 경로(multi_db_executor._generate_sql) 대칭 배선 — 단일 경로와 동일 SQL 형태."""

    async def test_month_series_deterministic_sql(self):
        from src.nodes.multi_db_executor import _generate_sql

        schema_info = {
            "tables": {"cmm_resource": {"columns": [{"name": "hostname", "type": "text"}]}},
            **_EAV_META,
        }
        month_fields = [f"{_AVG}|{'M' if k == 0 else f'M+{k}'}" for k in range(6)]
        month_fields += [f"{_PEAK}|{'M' if k == 0 else f'M+{k}'}" for k in range(6)]
        form_fill_out: dict = {}
        sql = await _generate_sql(
            llm=None,  # 결정적 경로 — LLM에 도달하면 안 된다
            parsed_requirements={"original_query": "2026년 6월 기준 CPU 양식 채워줘"},
            schema_info=schema_info,
            sub_query_context="CPU 양식 채우기",
            default_limit=100_000,
            column_mapping={"호스트명": "cmm_resource.hostname"},
            db_engine="postgresql",
            db_id="",
            unmapped_fields=month_fields,
            app_config=SimpleNamespace(
                text2sql=SimpleNamespace(semantic_compose=False, generic_llm_mapping=False),
                get_polestar_db_ids=lambda: set(),
            ),
            form_context_text="나. 주요 업무 CPU 사용현황",
            form_fill_out=form_fill_out,
        )
        assert "s.stat_date='202601'" in sql and "s.stat_date='202606'" in sql
        assert f'AS "{_AVG}|M"' in sql
        assert f'AS "{_PEAK}|M+5"' in sql
        assert 'AS "호스트명"' in sql
        assert form_fill_out["month_anchor"]["end"] == "202606"
        assert len(form_fill_out["month_anchor"]["fields"]) == 12

    async def test_fires_with_empty_db_mapping(self):
        """per-DB 매핑이 비어도(라이브 실측 FIX-1) 월 인식이 발동하고 식별 컬럼을 주입한다."""
        from src.nodes.multi_db_executor import _generate_sql

        month_fields = [f"{_AVG}|{'M' if k == 0 else f'M+{k}'}" for k in range(6)]
        form_fill_out: dict = {}
        sql = await _generate_sql(
            llm=None,
            parsed_requirements={"original_query": "2026년 6월 기준 CPU 양식 채워줘"},
            schema_info={"tables": {"cmm_resource": {"columns": []}}, **_EAV_META},
            sub_query_context="CPU 양식 채우기",
            default_limit=100_000,
            column_mapping={},  # per-DB 매핑 공백
            db_engine="postgresql",
            db_id="",
            unmapped_fields=month_fields,
            app_config=SimpleNamespace(
                text2sql=SimpleNamespace(semantic_compose=False, generic_llm_mapping=False),
                get_polestar_db_ids=lambda: set(),
            ),
            form_context_text="나. 주요 업무 CPU 사용현황",
            form_fill_out=form_fill_out,
        )
        assert f'AS "{_AVG}|M+3"' in sql
        # 식별 컬럼 결정적 주입(양식 헤더와 무충돌 라틴 alias)
        assert 'AS "server_name"' in sql and 'AS "hostname"' in sql
        assert form_fill_out["month_anchor"]["start"] == "202601"


class TestWriterFillsMonthColumns:
    """실물 CPU 양식 픽스처에 월 칼럼이 채워지는지(행 키=복합 필드명 아키텍처) e2e."""

    def test_fill_real_cpu_form(self):
        import openpyxl
        import io

        from src.document.excel_parser import parse_excel_template
        from src.document.excel_writer import fill_excel_template

        file_data = (_FORMS_DIR / "CPU_양식.xlsx").read_bytes()
        template = parse_excel_template(file_data)
        mapping = _cpu_form_mapping()
        row = {
            "호스트명": "web-01",
            "제조사(모델명)": "Dell R740",
            f"{_AVG}|M": 11.5,
            f"{_AVG}|M+5": 16.5,
            f"{_PEAK}|M": 91.1,
            f"{_PEAK}|M+5": 96.6,
        }
        stats: dict[str, int] = {}
        out_bytes, filled = fill_excel_template(
            file_data, template, mapping, [row], fill_stats=stats
        )
        assert filled >= 6
        # 필드별 실제 채움 통계(D-147 판정 근거)
        assert stats["호스트명"] == 1
        assert stats[f"{_AVG}|M"] == 1
        assert stats["처리능력|(TPMC)"] == 0
        assert stats["비고"] == 0

        ws = openpyxl.load_workbook(io.BytesIO(out_bytes)).worksheets[0]
        assert ws.cell(row=7, column=3).value == "web-01"   # C7 호스트명
        assert ws.cell(row=7, column=5).value == 11.5       # E7 평균 M
        assert ws.cell(row=7, column=10).value == 16.5      # J7 평균 M+5
        assert ws.cell(row=7, column=11).value == 91.1      # K7 Peak M
        assert ws.cell(row=7, column=16).value == 96.6      # P7 Peak M+5
        assert ws.cell(row=7, column=4).value is None       # D7 TPMC — 공란 유지

    def test_db2_string_values_converted_numeric(self):
        """DB2 경로의 문자열 통계값("6.51")도 월 칼럼에서 숫자로 변환된다(numeric hint)."""
        import openpyxl
        import io

        from src.document.excel_parser import parse_excel_template
        from src.document.excel_writer import fill_excel_template

        file_data = (_FORMS_DIR / "CPU_양식.xlsx").read_bytes()
        template = parse_excel_template(file_data)
        row = {f"{_AVG}|M+1": "6.51"}
        out_bytes, _ = fill_excel_template(
            file_data, template, _cpu_form_mapping(), [row]
        )
        ws = openpyxl.load_workbook(io.BytesIO(out_bytes)).worksheets[0]
        assert ws.cell(row=7, column=6).value == 6.51


class TestFormFillNotes:
    """응답의 기준월 명시(§2.4) + 미작성 사유(D-147)."""

    def test_notes_appended(self):
        from src.nodes.output_generator import _append_form_fill_notes

        state = {
            "form_month_anchor": {
                "start": "202601", "end": "202606",
                "resource_type": "server.Cpus",
                "fields": [f"{_AVG}|M", f"{_AVG}|M+1"],
            },
            "column_mapping": {
                "호스트명": "cmm_resource.hostname",
                "구분|분류": None,
                "처리능력|(TPMC)": None,
                f"{_AVG}|M": None,     # 월 시리즈 — 미작성 사유에서 제외돼야 함
                f"{_AVG}|M+1": None,
            },
        }
        out = _append_form_fill_notes("본문", state)
        assert "2026년 1월" in out and "2026년 6월" in out
        assert "미작성 항목" in out
        assert "구분 > 분류" in out
        assert "처리능력 > (TPMC)" in out
        assert f"{_AVG} > M" not in out  # 월 필드는 채움 대상 — 사유 목록에 없어야 함

    def test_no_anchor_no_change(self):
        from src.nodes.output_generator import _append_form_fill_notes

        assert _append_form_fill_notes("본문", {"column_mapping": {"a": "t.c"}}) == "본문"

    def test_fill_stats_overrides_mapping_based_judgment(self):
        """실제로 채워진 칼럼(매핑 None이어도)은 미작성 목록에 오르지 않는다(라이브 실측 교정)."""
        from src.nodes.output_generator import _append_form_fill_notes

        state = {
            "form_month_anchor": {"start": "202601", "end": "202606", "fields": [f"{_AVG}|M"]},
            "column_mapping": {"제조사(모델명)": None, "처리능력|(TPMC)": None},
        }
        fill_stats = {"제조사(모델명)": 20, "처리능력|(TPMC)": 0, f"{_AVG}|M": 20, "비고": 0}
        out = _append_form_fill_notes("본문", state, fill_stats=fill_stats)
        unfilled_block = out.split("[미작성 항목]")[1]
        assert "처리능력 > (TPMC)" in unfilled_block
        assert "비고" in unfilled_block
        assert "제조사(모델명)" not in unfilled_block  # 채워졌으므로 제외
        assert "2026년 1월부터 2026년 6월까지" in out

    def test_all_month_columns_zero_warns_sql_check(self):
        """월 칼럼이 전부 0건이면 D-050 취지의 생성 SQL 확인 안내를 낸다."""
        from src.nodes.output_generator import _append_form_fill_notes

        fields = [f"{_AVG}|M", f"{_AVG}|M+1"]
        state = {
            "form_month_anchor": {"start": "202601", "end": "202606", "fields": fields},
            "column_mapping": {},
        }
        out = _append_form_fill_notes(
            "본문", state, fill_stats={fields[0]: 0, fields[1]: 0, "호스트명": 10}
        )
        assert "[확인 필요]" in out and "생성 SQL" in out

    def test_parser_title_text_contains_form_context(self):
        """파서가 추출한 title_text에 리소스 판정 문맥(CPU)이 담긴다."""
        from src.document.excel_parser import parse_excel_template

        sheet = parse_excel_template((_FORMS_DIR / "CPU_양식.xlsx").read_bytes())["sheets"][0]
        assert "CPU" in sheet["title_text"]
        assert sheet["header_block_rows"] == [5, 6]


class TestHallucinatedColumnGuard:
    """환각 매핑 칼럼의 결정적 피벗 유입 차단(FIX-5 — 라이브 c.category 실측)."""

    _SCHEMA_TABLES = {
        "tables": {
            "cmm_resource": {
                "columns": [
                    {"name": "name", "type": "text"},
                    {"name": "hostname", "type": "text"},
                    {"name": "id", "type": "bigint"},
                ],
            },
            "no_column_info": {"columns": []},
        }
    }

    def test_helper_drops_missing_column_only(self):
        from src.utils.query_gen_common import drop_entries_missing_columns

        entries = [
            ("호스트명", "cmm_resource.hostname"),        # 존재 → 유지
            ("구분|분류", "cmm_resource.category"),        # 부재 → 제외
            ("기타", "no_column_info.whatever"),           # 칼럼 정보 없음 → 유지(오탐 방지)
            ("미지테이블", "unknown_table.col"),           # 테이블 미지 → 유지
        ]
        kept, dropped = drop_entries_missing_columns(entries, self._SCHEMA_TABLES)
        assert ("호스트명", "cmm_resource.hostname") in kept
        assert ("기타", "no_column_info.whatever") in kept
        assert ("미지테이블", "unknown_table.col") in kept
        assert dropped == [("구분|분류", "cmm_resource.category")]

    def test_single_path_excludes_hallucinated_column(self):
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        mapping = _cpu_form_mapping()
        mapping["구분|분류"] = "cmm_resource.category"  # 환각 매핑
        state = {
            "column_mapping": mapping,
            "schema_info": {**self._SCHEMA_TABLES, **_EAV_META},
            "template_structure": {"sheets": [{"title_text": "CPU 사용현황", "name": "Sheet1"}]},
            "active_db_engine": "postgresql",
            "active_db_id": "",
        }
        result = _try_build_form_fill_pivot_sql(state, 100_000, "2026년 6월 기준")
        sql = result["sql"]
        assert "category" not in sql
        assert 'AS "호스트명"' in sql

    async def test_multi_path_excludes_hallucinated_column(self):
        from src.nodes.multi_db_executor import _generate_sql

        month_fields = [f"{_AVG}|{'M' if k == 0 else f'M+{k}'}" for k in range(6)]
        sql = await _generate_sql(
            llm=None,
            parsed_requirements={"original_query": "2026년 6월 기준 CPU 양식 채워줘"},
            schema_info={**self._SCHEMA_TABLES, **_EAV_META},
            sub_query_context="CPU 양식 채우기",
            default_limit=100_000,
            column_mapping={"구분|분류": "cmm_resource.category", "호스트명": "cmm_resource.hostname"},
            db_engine="postgresql",
            db_id="",
            unmapped_fields=month_fields,
            app_config=SimpleNamespace(
                text2sql=SimpleNamespace(semantic_compose=False, generic_llm_mapping=False),
                get_polestar_db_ids=lambda: set(),
            ),
            form_context_text="나. 주요 업무 CPU 사용현황",
            form_fill_out={},
        )
        assert "category" not in sql
        assert 'AS "호스트명"' in sql


class TestInsufficiencyLoopSuppression:
    """결정적 월 피벗 결과가 부족 판정 재시도(LLM 재생성)로 덮이는 회귀 방지(FIX-6)."""

    async def test_month_anchor_suppresses_retry(self, monkeypatch):
        import importlib

        ro = importlib.import_module("src.nodes.result_organizer")

        async def _always_insufficient(*args, **kwargs):
            return False

        monkeypatch.setattr(ro, "_check_data_sufficiency", _always_insufficient)
        state = {
            "query_results": [{"호스트명": "web-01", f"{_AVG}|M": 11.1}],
            "parsed_requirements": {"output_format": "xlsx"},
            "template_structure": None,
            "column_mapping": {"호스트명": "cmm_resource.hostname"},
            "retry_count": 0,
            "form_month_anchor": {"start": "202601", "end": "202606", "fields": [f"{_AVG}|M"]},
        }
        result = await ro.result_organizer(state)
        assert result.get("error_message") != "data_insufficient"
        assert result["organized_data"]["is_sufficient"] is not False

    async def test_form_turn_without_month_anchor_also_suppressed(self, monkeypatch):
        """월 시리즈 없는 폼필 턴(서버목록류)도 억제 — 라이브 실측(2026-07-30 3존):
        의도적 공란이 부족 판정 → 재시도 턴이 결정적 조립을 스킵 → LLM 폴백이 계약을 덮음."""
        import importlib

        ro = importlib.import_module("src.nodes.result_organizer")

        async def _always_insufficient(*args, **kwargs):
            return False

        monkeypatch.setattr(ro, "_check_data_sufficiency", _always_insufficient)
        state = {
            "query_results": [{"서버 이름": "web-01", "IP 주소": "10.0.0.1"}],
            "parsed_requirements": {"output_format": "xlsx"},
            "template_structure": {"sheets": [{"name": "Sheet1"}]},
            "column_mapping": {"서버 이름": "cmm_resource.name"},
            "retry_count": 0,
        }
        result = await ro.result_organizer(state)
        assert result.get("error_message") != "data_insufficient"
        assert result["organized_data"]["is_sufficient"] is not False

    async def test_without_anchor_retry_still_requested(self, monkeypatch):
        import importlib

        ro = importlib.import_module("src.nodes.result_organizer")

        async def _always_insufficient(*args, **kwargs):
            return False

        monkeypatch.setattr(ro, "_check_data_sufficiency", _always_insufficient)
        state = {
            "query_results": [{"호스트명": "web-01"}],
            "parsed_requirements": {"output_format": "xlsx"},
            "template_structure": None,
            "column_mapping": None,
            "retry_count": 0,
        }
        result = await ro.result_organizer(state)
        assert result.get("error_message") == "data_insufficient"


class TestWriterMappingForceNull:
    """월 필드 N:1 오염 매핑의 강제 None 갱신(라이브 3차 실측 — Excel 동일값 복제 방지)."""

    def test_polluted_metric_mapping_forced_to_none(self):
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        mapping = _cpu_form_mapping()
        # field_mapper 오염 재현: 월 필드 전부가 같은 metric 칼럼으로 매핑됨
        for k in range(6):
            sub = "M" if k == 0 else f"M+{k}"
            mapping[f"{_AVG}|{sub}"] = "cmm_metric_stat_m.avg_val"
            mapping[f"{_PEAK}|{sub}"] = "cmm_metric_stat_m.max_val"
        state = {
            "column_mapping": mapping,
            "schema_info": _EAV_META,
            "template_structure": {"sheets": [{"title_text": "CPU 사용현황", "name": "Sheet1"}]},
            "active_db_engine": "postgresql",
            "active_db_id": "",
        }
        result = _try_build_form_fill_pivot_sql(state, 100_000, "2026년 6월 기준")
        assert result is not None
        # 인식은 발동(월별 리터럴 SQL) + 오염 매핑은 전부 None으로 갱신
        assert "s.stat_date='202603'" in result["sql"]
        for k in range(6):
            sub = "M" if k == 0 else f"M+{k}"
            assert result["mapping_updates"][f"{_AVG}|{sub}"] is None
            assert result["mapping_updates"][f"{_PEAK}|{sub}"] is None


class TestInferredMappingDisplayNoneSafety:
    """오케스트레이션 경로에서 column_mapping이 None 값으로 실려도 크래시 없어야 한다
    (라이브 실측 2026-07-30 B0 CPU: 'NoneType' object has no attribute 'get' 후보)."""

    def test_none_column_mapping_does_not_crash(self):
        from src.nodes.output_generator import _append_inferred_mapping_info

        state = {
            "mapping_sources": {"구분": "llm_inferred"},
            "column_mapping": None,
            "db_column_mapping": None,
        }
        assert _append_inferred_mapping_info("응답", state) == "응답"


class TestResponseTableMarkdownSafety:
    """복합 필드명 '|'의 Markdown 표 충돌 방지 — 표시용 키 결정적 치환."""

    def test_pipe_keys_replaced_in_prompt(self):
        from src.nodes.output_generator import _build_response_prompt

        rows = [{f"{_AVG}|M+2": 11.5, "처리능력|(TPMC)": None, "호스트명": "web-01"}]
        prompt = _build_response_prompt("질의", "요약", rows)
        assert f"{_AVG} > M+2" in prompt
        assert "처리능력 > (TPMC)" in prompt
        assert f"{_AVG}|M+2" not in prompt  # 원본 '|' 키는 프롬프트에 없어야 함


class TestWriterStrictFieldLookup:
    """필드명 폴백 열의 엄격 조회(FIX-11) — NULL 월 키 누락 시 이웃 월 복제 방지."""

    def test_missing_month_key_stays_blank(self):
        """6월 생성 서버(행에 M+5 키만 존재): M~M+4는 공란, M+5만 채워져야 한다."""
        import io
        import openpyxl

        from src.document.excel_parser import parse_excel_template
        from src.document.excel_writer import fill_excel_template

        file_data = (_FORMS_DIR / "CPU_양식.xlsx").read_bytes()
        template = parse_excel_template(file_data)
        # NULL 칼럼이 직렬화에서 누락된 행 재현 — M+5만 존재
        row = {"호스트명": "new-06", f"{_AVG}|M+5": 7.7}
        out_bytes, _ = fill_excel_template(file_data, template, _cpu_form_mapping(), [row])
        ws = openpyxl.load_workbook(io.BytesIO(out_bytes)).worksheets[0]
        assert ws.cell(row=7, column=5).value is None    # E7 = 평균 M — 복제 금지
        assert ws.cell(row=7, column=9).value is None    # I7 = 평균 M+4 — 복제 금지
        assert ws.cell(row=7, column=10).value == 7.7    # J7 = 평균 M+5

    def test_db2_lowercase_latin_key_still_matches(self):
        """DB2 라틴 소문자화('…|m+1') 행 키는 정규화 동등으로 흡수한다(엄격 조회 범위 내)."""
        import io
        import openpyxl

        from src.document.excel_parser import parse_excel_template
        from src.document.excel_writer import fill_excel_template

        file_data = (_FORMS_DIR / "CPU_양식.xlsx").read_bytes()
        template = parse_excel_template(file_data)
        row = {f"{_AVG}|m+1".replace("M", "m"): 6.51}  # 라틴만 소문자화된 키
        out_bytes, _ = fill_excel_template(file_data, template, _cpu_form_mapping(), [row])
        ws = openpyxl.load_workbook(io.BytesIO(out_bytes)).worksheets[0]
        assert ws.cell(row=7, column=6).value == 6.51    # F7 = 평균 M+1


class TestUnverifiableSchemaWhitelist:
    """스키마에 칼럼 목록이 없을 때 entity 안전 화이트리스트(FIX-13 — gp+yd category 차단)."""

    def test_category_dropped_hostname_kept(self):
        from src.db_adapters.polestar.assembler import filter_pivot_regular_entries

        entries = [
            ("호스트명", "cmm_resource.hostname"),   # 화이트리스트 → 유지
            ("비고", "cmm_resource.name"),           # 화이트리스트 → 유지
            ("구분|분류", "cmm_resource.category"),   # 화이트리스트 밖 → 제외
        ]
        kept, dropped = filter_pivot_regular_entries(entries, {}, "cmm_resource")
        assert ("호스트명", "cmm_resource.hostname") in kept
        assert ("비고", "cmm_resource.name") in kept
        assert dropped == [("구분|분류", "cmm_resource.category")]

    def test_single_path_blocks_category_without_schema_columns(self):
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        mapping = _cpu_form_mapping()
        mapping["구분|분류"] = "cmm_resource.category"
        state = {
            "column_mapping": mapping,
            "schema_info": _EAV_META,  # tables 정보 없음 — 검증 불가 상황
            "template_structure": {"sheets": [{"title_text": "CPU 사용현황", "name": "Sheet1"}]},
            "active_db_engine": "postgresql",
            "active_db_id": "",
        }
        result = _try_build_form_fill_pivot_sql(state, 100_000, "2026년 6월 기준")
        assert "category" not in result["sql"]
        assert 'AS "호스트명"' in result["sql"]


class TestLlmFallbackMonthBlock:
    """LLM 폴백 경로의 월 리터럴 강제(라이브 5차 실측 — CURRENT_DATE 역방향 월 뒤집힘 차단)."""

    def test_block_contains_literal_months(self):
        from src.db_adapters.polestar.assembler import build_month_series_block

        ms = recognize_month_series(
            _cpu_form_mapping(), context_text="CPU 사용현황",
            user_query="2026년 6월 기준", today=_TODAY,
        )
        block = build_month_series_block(ms)
        assert f'"{_AVG}|M" ← s.stat_date = \'202601\'' in block
        assert f'"{_PEAK}|M+5" ← s.stat_date = \'202606\'' in block
        assert "CURRENT_DATE 동적 계산 금지" in block
        assert build_month_series_block(None) == ""

    async def test_multi_retry_prompt_gets_month_block(self):
        """멀티 재생성(error_context) 턴: 결정적 조립은 스킵하되 프롬프트에 월 리터럴 강제."""
        from src.nodes.multi_db_executor import _generate_sql

        class _FakeLLM:
            def __init__(self):
                self.captured = None

            async def ainvoke(self, messages):
                self.captured = messages
                return SimpleNamespace(content="SELECT 1;")

        llm = _FakeLLM()
        month_fields = [f"{_AVG}|{'M' if k == 0 else f'M+{k}'}" for k in range(6)]
        await _generate_sql(
            llm=llm,
            parsed_requirements={"original_query": "2026년 6월 기준 CPU 양식 채워줘"},
            schema_info={"tables": {"cmm_resource": {"columns": [{"name": "hostname", "type": "text"}]}}, **_EAV_META},
            sub_query_context="CPU 양식 채우기",
            default_limit=100_000,
            error_context="SQL 검증 실패: something",
            column_mapping={"호스트명": "cmm_resource.hostname"},
            db_engine="postgresql",
            db_id="",
            unmapped_fields=month_fields,
            app_config=SimpleNamespace(
                text2sql=SimpleNamespace(
                    semantic_compose=False, generic_llm_mapping=False,
                    multi_candidate=False, complexity_gate=False,
                    # 병합(multiintent) 신규 소비 플래그 — 명시 OFF(테스트 config는
                    # 검증 대상 필드를 명시, Known Mistakes)
                    query_history_fewshot=False, query_history_top_k=0,
                    query_history_min_score=0.0, path_parity=False,
                    multi_full_validation=False, prompt_canonical_render=False,
                    value_retrieval=False,
                ),
                query=SimpleNamespace(default_limit=1000),
                get_polestar_db_ids=lambda: set(),
            ),
            form_context_text="나. 주요 업무 CPU 사용현황",
            form_fill_out={},
        )
        human = llm.captured[-1].content
        assert "월별 칼럼 매핑 강제" in human
        assert "s.stat_date = '202601'" in human


class TestForeignTableRegularEntryDropped:
    """entity 외 테이블의 유효 매핑이 c.* 로 재작성되는 결함 차단 (라이브 6차 실측 확정 FIX-15).

    category는 cmm_resource_type에 **실존**하는 칼럼이라 부재-칼럼 검증(FIX-5)·화이트리스트
    (FIX-13)를 전부 통과했고, 조립기가 테이블명을 떼고 c.(entity)에 붙여 5회 연속
    `column c.category does not exist`를 만들었다.
    """

    def test_existing_column_on_other_table_dropped(self):
        from src.db_adapters.polestar.assembler import filter_pivot_regular_entries

        schema = {"tables": {
            "polestar.cmm_resource_type": {"columns": [{"name": "category", "type": "text"}]},
            "polestar.cmm_resource": {"columns": [{"name": "hostname", "type": "text"},
                                                   {"name": "name", "type": "text"}]},
        }}
        entries = [
            ("구분|분류", "cmm_resource_type.category"),  # 타 테이블 실존 칼럼 → 제외
            ("호스트명", "cmm_resource.hostname"),        # entity 칼럼 → 유지
        ]
        kept, dropped = filter_pivot_regular_entries(entries, schema, "cmm_resource")
        assert kept == [("호스트명", "cmm_resource.hostname")]
        assert ("구분|분류", "cmm_resource_type.category") in dropped

    def test_single_path_excludes_foreign_table_column(self):
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        mapping = _cpu_form_mapping()
        mapping["구분|분류"] = "cmm_resource_type.category"
        state = {
            "column_mapping": mapping,
            "schema_info": {
                "tables": {"polestar.cmm_resource_type": {"columns": [{"name": "category", "type": "text"}]}},
                **_EAV_META,
            },
            "template_structure": {"sheets": [{"title_text": "CPU 사용현황", "name": "Sheet1"}]},
            "active_db_engine": "postgresql",
            "active_db_id": "",
        }
        result = _try_build_form_fill_pivot_sql(state, 100_000, "2026년 6월 기준")
        assert "category" not in result["sql"]
        assert 'AS "호스트명"' in result["sql"]


class TestFormFillHitl:
    """Plan 73 Phase 2 (D-151) — 역질문 답변 검증·적용·후보·미해결 수집·writer 상수."""

    _META = _EAV_META
    _SCHEMA = {
        "tables": {"cmm_resource": {"columns": [
            {"name": "name", "type": "text"},
            {"name": "hostname", "type": "text"},
            {"name": "description", "type": "text"},
        ]}},
        **_EAV_META,
    }

    def _eav_pattern(self):
        return self._META["_structure_meta"]["patterns"][0]

    # ── resolve_form_fill_answers (검증기) ──

    def test_answers_validated_and_partitioned(self):
        from src.db_adapters.polestar.assembler import resolve_form_fill_answers

        answers = {
            "구분": {"action": "blank", "value": None},
            "도입일자": {"action": "column", "value": "description"},
            "제조사": {"action": "eav", "value": "Vendor"},
            "용도": {"action": "literal", "value": "웹서버"},
            "설치장소": {"action": "column", "value": "no_such_column"},   # 존재성 실패
            "빈값": {"action": "literal", "value": ""},                    # 값 없음
            "이상한것": {"action": "wormhole", "value": "x"},              # 미지 액션
        }
        ov, mu, lit = resolve_form_fill_answers(answers, self._SCHEMA, self._eav_pattern())
        assert ov["구분"]["applied"] and mu["구분"] is None
        assert ov["도입일자"]["applied"] and mu["도입일자"] == "cmm_resource.description"
        assert ov["제조사"]["applied"] and mu["제조사"] == "EAV:Vendor"
        assert ov["용도"]["applied"] and lit == {"용도": "웹서버"} and mu["용도"] is None
        assert not ov["설치장소"]["applied"] and "존재성" in ov["설치장소"]["reason"]
        assert not ov["빈값"]["applied"]
        assert not ov["이상한것"]["applied"]
        assert "설치장소" not in mu and "빈값" not in lit

    def test_protected_month_fields_rejected(self):
        from src.db_adapters.polestar.assembler import resolve_form_fill_answers

        ov, mu, _ = resolve_form_fill_answers(
            {f"{_AVG}|M": {"action": "blank", "value": None}},
            self._SCHEMA, self._eav_pattern(),
            protected_fields={f"{_AVG}|M"},
        )
        assert not ov[f"{_AVG}|M"]["applied"]
        assert mu == {}

    # ── build_form_fill_candidates (후보) ──

    def test_candidates_from_schema_and_profile(self):
        from src.db_adapters.polestar.assembler import build_form_fill_candidates

        cands = build_form_fill_candidates(self._SCHEMA, self._eav_pattern())
        values = {c["value"] for c in cands}
        assert "column:description" in values
        assert "eav:Vendor" in values and "eav:TotalSize" in values
        vendor = next(c for c in cands if c["value"] == "eav:Vendor")
        assert "제조사" in vendor["label"]  # known_attributes 설명이 한글 라벨로

    def test_candidates_whitelist_when_no_schema_columns(self):
        from src.db_adapters.polestar.assembler import build_form_fill_candidates

        cands = build_form_fill_candidates(self._META, self._eav_pattern())
        values = {c["value"] for c in cands}
        assert "column:hostname" in values and "column:name" in values  # 안전 화이트리스트

    # ── 단일 경로 오버라이드 적용 ──

    def test_single_path_applies_answers(self):
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        state = {
            "column_mapping": {"서버 이름": "cmm_resource.name", "도입일자": None, "구분": None},
            "schema_info": self._SCHEMA,
            "template_structure": {"sheets": [{"title_text": "서버 목록", "name": "Sheet1"}]},
            "active_db_engine": "postgresql",
            "active_db_id": "",
            "form_fill_answers": {
                "도입일자": {"action": "column", "value": "description"},
                "구분": {"action": "literal", "value": "운영"},
            },
        }
        result = _try_build_form_fill_pivot_sql(state, 1000, "양식 채워줘")
        assert result is not None
        sql = result["sql"]
        assert 'THEN c.description END) AS "도입일자"' in sql
        assert "구분" not in sql  # literal은 SQL 제외 — writer 상수 기입
        assert result["literals"] == {"구분": "운영"}
        assert result["overrides"]["도입일자"]["applied"]
        assert result["mapping_updates"]["도입일자"] == "cmm_resource.description"
        assert result["candidates"], "후보 목록이 함께 산출되어야 한다"

    async def test_multi_path_applies_answers_symmetric(self):
        from src.nodes.multi_db_executor import _generate_sql

        form_fill_out: dict = {}
        sql = await _generate_sql(
            llm=None,
            parsed_requirements={"original_query": "[양식 미해결 항목 답변]"},
            schema_info=self._SCHEMA,
            sub_query_context="양식 재채움",
            default_limit=1000,
            column_mapping={"서버 이름": "cmm_resource.name"},
            db_engine="postgresql",
            db_id="",
            unmapped_fields=["도입일자"],
            app_config=SimpleNamespace(
                text2sql=SimpleNamespace(semantic_compose=False, generic_llm_mapping=False),
                get_polestar_db_ids=lambda: set(),
            ),
            form_context_text="서버 목록",
            form_fill_out=form_fill_out,
            form_intent=True,
            mapping_sources={"서버 이름": "synonym"},
            form_fill_answers={"도입일자": {"action": "column", "value": "description"}},
        )
        assert 'THEN c.description END) AS "도입일자"' in sql
        assert form_fill_out["overrides"]["도입일자"]["applied"]
        assert form_fill_out["mapping_updates"]["도입일자"] == "cmm_resource.description"
        assert form_fill_out["candidates"]

    # ── writer 직접 입력 상수 ──

    def test_writer_fills_literal_values(self):
        import io

        from openpyxl import Workbook, load_workbook

        from src.document.excel_parser import parse_excel_template
        from src.document.excel_writer import fill_excel_template

        wb = Workbook()
        ws = wb.active
        ws.title = "서버현황"
        ws.append(["서버 이름", "용도"])
        buf = io.BytesIO()
        wb.save(buf)
        template = parse_excel_template(buf.getvalue())

        fill_stats: dict = {}
        out_bytes, filled = fill_excel_template(
            file_data=buf.getvalue(),
            template_structure=template,
            column_mapping={"서버 이름": None, "용도": None},
            rows=[{"서버 이름": "web-01"}, {"서버 이름": "web-02"}],
            fill_stats=fill_stats,
            literal_values={"용도": "웹서버"},
        )
        out_ws = load_workbook(io.BytesIO(out_bytes)).active
        assert out_ws.cell(row=2, column=2).value == "웹서버"
        assert out_ws.cell(row=3, column=2).value == "웹서버"
        assert fill_stats["용도"] == 2
        assert fill_stats["서버 이름"] == 2

    # ── 미해결 수집 → 역질문 페이로드/대기 상태 ──

    def test_hitl_payload_from_fill_stats(self):
        from src.nodes.output_generator import _build_form_fill_hitl

        state = {
            "template_structure": {"sheets": [{"name": "Sheet1"}]},
            "uploaded_file": b"xlsx-bytes",
            "file_type": "xlsx",
            "parsed_requirements": {"original_query": "양식 채워줘", "output_format": "xlsx"},
            "form_month_anchor": {"fields": [f"{_AVG}|M"]},
            "form_fill_literals": {"용도": "웹서버"},
            "form_fill_overrides": {"구분": {"action": "blank", "applied": True}},
            "form_fill_candidates": [{"value": "column:name", "label": "cmm_resource.name", "kind": "column"}],
            "selected_db_ids": ["polestar_cm_gp", "polestar_cm_yd"],
        }
        fill_stats = {
            "서버 이름": 5, f"{_AVG}|M": 0, "용도": 5,
            "구분": 0, "도입일자": 0,
        }
        clar, pending = _build_form_fill_hitl(state, fill_stats)
        assert clar is not None and pending is not None
        names = {f["name"] for f in clar["fields"]}
        # 월 시리즈·직접입력·사용자 지정 공란은 역질문 제외
        assert names == {"도입일자"}
        assert clar["candidates"]
        assert pending["uploaded_file"] == b"xlsx-bytes"
        assert pending["file_type"] == "xlsx"
        assert pending["unresolved"] == ["도입일자"]
        # FIX-26: 존 체크박스 런의 확정 존이 pending에 보존돼야 답변 턴이 복원한다
        assert pending["db_ids"] == ["polestar_cm_gp", "polestar_cm_yd"]

    def test_hitl_pending_db_ids_fallback_without_selection(self):
        """FIX-26 폴백: 존 체크박스 선택이 없는 런(텍스트 위치어 런)은 라우팅 확정
        (target_databases∪active_db_id∪mapped_db_ids)에서 존을 보존한다."""
        from src.nodes.output_generator import _build_form_fill_hitl

        state = {
            "template_structure": {"sheets": [{"name": "Sheet1"}]},
            "uploaded_file": b"xlsx-bytes",
            "file_type": "xlsx",
            "parsed_requirements": {"original_query": "김포 서버로 채워줘"},
            "selected_db_ids": None,
            "target_databases": [{"db_id": "polestar_cm_gp"}],
            "active_db_id": "polestar_cm_gp",
        }
        clar, pending = _build_form_fill_hitl(state, {"서버 이름": 5, "도입일자": 0})
        assert pending is not None
        assert pending["db_ids"] == ["polestar_cm_gp"]

    def test_hitl_none_when_all_zero_or_resolved(self):
        from src.nodes.output_generator import _build_form_fill_hitl

        state = {"template_structure": {"sheets": []}, "parsed_requirements": {}}
        # 전 필드 0건 = 데이터/SQL 문제 — 역질문하지 않는다(D-050)
        assert _build_form_fill_hitl(state, {"a": 0, "b": 0}) == (None, None)
        # 미해결 없음 — pending 자기정리 대상
        assert _build_form_fill_hitl(state, {"a": 3, "b": 1}) == (None, None)

    # ── 라우트: 답변 턴 delta 조립 ──

    def test_route_restores_pending_file_on_answer_turn(self):
        from src.api.routes.query import _build_turn_input_state
        from src.api.schemas import QueryRequest

        body = QueryRequest(
            query="[양식 미해결 항목 답변]",
            form_fill_answers={"도입일자": {"action": "blank", "value": None}},
        )
        checkpoint = {
            "pending_form_fill": {
                "uploaded_file": b"xlsx-bytes",
                "file_type": "xlsx",
                "original_query": "양식 채워줘",
                "unresolved": ["도입일자"],
                "db_ids": ["polestar_cm_gp", "polestar_cm_yd"],
            }
        }
        delta = _build_turn_input_state(body, "t1", checkpoint, {"sub": "u1"})
        assert delta["uploaded_file"] == b"xlsx-bytes"
        assert delta["file_type"] == "xlsx"
        assert delta["form_fill_answers"] == {"도입일자": {"action": "blank", "value": None}}
        # 요청 스코프 초기화 계약(D-064)은 유지된 delta 위에 복원되어야 한다
        assert delta["form_fill_overrides"] is None
        # FIX-17: 파이프라인 입력은 원 질의로 복원(위치 힌트 유실 → priority 공백 →
        # 전 DB 유사어 프롬프트 413 + 오라우팅 라이브 실측) + 폼필 LIMIT 복원
        assert delta["user_query"] == "양식 채워줘"
        from src.utils.query_gen_common import _ALL_QUERY_LIMIT

        assert delta["resolved_limit"] == _ALL_QUERY_LIMIT  # 폼필 기본 = 전량 상향값(D-134 정합)
        # FIX-26: 존 체크박스 런("채워줘")은 원 질의에 위치어가 없어 pending 보존 존을
        # selected_db_ids로 복원해야 기본 DB(b0) 침묵 오라우팅·타 DB 존재성 검증 탈락을 막는다
        assert delta["selected_db_ids"] == ["polestar_cm_gp", "polestar_cm_yd"]

    def test_route_body_selection_wins_over_pending_db_ids(self):
        """FIX-26 우선순위: 이번 턴 명시 존 선택(body.selected_db_ids)이 pending 보존분보다
        우선한다(요청 스코프 계약 — 사용자가 새로 고르면 그 값이 이긴다)."""
        from src.api.routes.query import _build_turn_input_state
        from src.api.schemas import QueryRequest

        body = QueryRequest(
            query="[양식 미해결 항목 답변]",
            form_fill_answers={"도입일자": {"action": "blank", "value": None}},
            selected_db_ids=["polestar_b0"],
        )
        checkpoint = {
            "pending_form_fill": {
                "uploaded_file": b"xlsx-bytes",
                "file_type": "xlsx",
                "original_query": "양식 채워줘",
                "unresolved": ["도입일자"],
                "db_ids": ["polestar_cm_gp"],
            }
        }
        delta = _build_turn_input_state(body, "t1", checkpoint, {"sub": "u1"})
        assert delta["selected_db_ids"] == ["polestar_b0"]

    def test_route_legacy_pending_without_db_ids(self):
        """FIX-26 하위호환: 구버전 pending(db_ids 부재)은 기존 동작(존 미복원)을 유지한다."""
        from src.api.routes.query import _build_turn_input_state
        from src.api.schemas import QueryRequest

        body = QueryRequest(
            query="[양식 미해결 항목 답변]",
            form_fill_answers={"도입일자": {"action": "blank", "value": None}},
        )
        checkpoint = {
            "pending_form_fill": {
                "uploaded_file": b"xlsx-bytes",
                "file_type": "xlsx",
                "original_query": "양식 채워줘",
                "unresolved": ["도입일자"],
            }
        }
        delta = _build_turn_input_state(body, "t1", checkpoint, {"sub": "u1"})
        assert delta["selected_db_ids"] is None

    def test_route_ignores_answers_without_pending(self):
        from src.api.routes.query import _build_turn_input_state
        from src.api.schemas import QueryRequest

        body = QueryRequest(
            query="서버 목록 조회",
            form_fill_answers={"도입일자": {"action": "blank", "value": None}},
        )
        delta = _build_turn_input_state(body, "t1", {"pending_form_fill": None}, {"sub": "u1"})
        assert delta.get("uploaded_file") is None
        assert delta.get("form_fill_answers") is None

    def test_isolated_input_carries_form_fill_answers(self):
        """FIX-19: 오케스트레이션 격리 경계(_make_isolated_input)가 답변을 통과시켜야
        오버라이드가 파이프라인에 도달한다(누락 시 역질문 무한 반복 — 라이브 실측)."""
        from src.orchestration.subagents import _make_isolated_input

        answers = {"도입일자": {"action": "column", "value": "ctime"}}
        state = {
            "user_query": "여의도 센터의 서버들에 대해 양식을 채우시오",
            "parsed_requirements": {"original_query": "여의도 센터의 서버들에 대해 양식을 채우시오"},
            "template_structure": {"sheets": []},
            "form_fill_answers": answers,
        }
        task = {"task_id": "t1", "agent": "data_query", "sub_query": "양식 채우기"}
        isolated = _make_isolated_input(task, state, {})
        assert isolated["form_fill_answers"] == answers


class _FakeRedis:
    """form_memory 계약 테스트용 인메모리 Redis 대역(TTL은 기록만)."""

    def __init__(self):
        self.store: dict = {}
        self.ttls: dict = {}

    async def load_form_memory(self, signature):
        return self.store.get(signature)

    async def save_form_memory(self, signature, data, ttl_seconds):
        self.store[signature] = data
        self.ttls[signature] = ttl_seconds
        return True

    async def delete_form_memory(self, signature):
        self.ttls.pop(signature, None)
        return self.store.pop(signature, None) is not None

    async def touch_form_memory(self, signature, ttl_seconds):
        self.ttls[signature] = ttl_seconds


def _tpl(headers, title=""):
    return {"sheets": [{"name": "Sheet1", "headers": headers, "title_text": title}]}


_CFG7 = SimpleNamespace(query=SimpleNamespace(form_memory_ttl_days=7))
_CFG0 = SimpleNamespace(query=SimpleNamespace(form_memory_ttl_days=0))


class TestFormSignature:
    """양식 시그니처(D-151 Phase 3) — 헤더 집합만, 정규화 불변."""

    def test_whitespace_and_case_invariant(self):
        from src.utils.schema_utils import form_signature

        a = form_signature(_tpl(["IP 주소", "서버명"]))
        b = form_signature(_tpl(["IP주소", "서버명 "]))
        assert a is not None and a == b

    def test_title_and_sheet_name_ignored(self):
        from src.utils.schema_utils import form_signature

        a = form_signature(_tpl(["서버명"], title="금감원 제출용"))
        b = form_signature({"sheets": [{"name": "다른시트", "headers": ["서버명"], "title_text": ""}]})
        assert a == b

    def test_different_fields_differ_and_empty_none(self):
        from src.utils.schema_utils import form_signature

        assert form_signature(_tpl(["서버명"])) != form_signature(_tpl(["서버명", "IP"]))
        assert form_signature(_tpl([])) is None
        assert form_signature(None) is None


class TestFormMemoryStore:
    """확인 이력 저장소 계약 — TTL sliding·병합·삭제·기능 OFF."""

    def _patch(self, monkeypatch, fake):
        import src.schema_cache.form_memory as fm

        async def _fake_get_redis(app_config):
            return fake

        monkeypatch.setattr(fm, "_get_redis", _fake_get_redis)
        return fm

    async def test_save_load_touch_cycle(self, monkeypatch):
        fake = _FakeRedis()
        fm = self._patch(monkeypatch, fake)
        tpl = _tpl(["서버명", "도입일자"], title="서버 목록")

        display = await fm.save_form_memory_entries(
            tpl, {"도입일자": {"action": "column", "value": "ctime"}},
            "여의도 서버 양식 채워줘", _CFG7,
        )
        assert display == "서버 목록"
        sig = list(fake.store)[0]
        assert fake.ttls[sig] == 7 * 86400

        sig2, answers, meta = await fm.load_form_memory_answers(tpl, _CFG7)
        assert sig2 == sig
        assert answers == {"도입일자": {"action": "column", "value": "ctime", "origin": "memory"}}
        # sliding: 적용 시 use_count 증가 + TTL 재설정
        assert fake.store[sig]["use_count"] == 1

    async def test_merge_keeps_existing_fields(self, monkeypatch):
        fake = _FakeRedis()
        fm = self._patch(monkeypatch, fake)
        tpl = _tpl(["a", "b"])
        await fm.save_form_memory_entries(tpl, {"a": {"action": "blank", "value": None}}, "q", _CFG7)
        await fm.save_form_memory_entries(tpl, {"b": {"action": "literal", "value": "x"}}, "q", _CFG7)
        sig = list(fake.store)[0]
        assert set(fake.store[sig]["fields"]) == {"a", "b"}

    async def test_delete_field_and_all(self, monkeypatch):
        fake = _FakeRedis()
        fm = self._patch(monkeypatch, fake)
        tpl = _tpl(["a", "b"])
        await fm.save_form_memory_entries(
            tpl,
            {"a": {"action": "blank", "value": None}, "b": {"action": "literal", "value": "x"}},
            "q", _CFG7,
        )
        removed, _ = await fm.delete_form_memory_entries(tpl, _CFG7, ["a"])
        assert removed == 1
        sig = list(fake.store)[0]
        assert set(fake.store[sig]["fields"]) == {"b"}
        removed, _ = await fm.delete_form_memory_entries(tpl, _CFG7, None)
        assert removed == 1 and not fake.store

    async def test_ttl_zero_disables_all(self, monkeypatch):
        fake = _FakeRedis()
        fm = self._patch(monkeypatch, fake)
        tpl = _tpl(["a"])
        assert await fm.save_form_memory_entries(tpl, {"a": {"action": "blank", "value": None}}, "q", _CFG0) is None
        _, answers, _ = await fm.load_form_memory_answers(tpl, _CFG0)
        assert answers == {} and not fake.store


class TestFormMemoryNotesAndSaveGate:
    """origin 분리 표시 + 저장 게이트(옵트인·적용·answer-origin만, C3)."""

    def test_notes_split_by_origin(self):
        from src.nodes.output_generator import _append_form_fill_notes

        state = {
            "form_fill_overrides": {
                "도입일자": {"action": "column", "value": "ctime", "applied": True, "origin": "answer"},
                "용도": {"action": "literal", "value": "웹서버", "applied": True, "origin": "memory"},
            },
        }
        out = _append_form_fill_notes("응답", state, fill_stats={"도입일자": 3, "용도": 3})
        assert "[사용자 답변 적용 내역]" in out and "도입일자" in out
        assert "[확인 이력 적용]" in out and "용도" in out
        assert "기억 삭제" in out  # 변경 방법 안내

    async def test_save_gate_only_applied_answer_origin(self, monkeypatch):
        import importlib

        og = importlib.import_module("src.nodes.output_generator")

        captured: dict = {}

        async def _fake_save(template, entries, original_query, app_config):
            captured["entries"] = entries
            return "서버 목록"

        async def _fake_text(*a, **k):
            return "본문"

        def _fake_doc(state, fmt):
            return {"file_bytes": b"x", "file_name": "r.xlsx", "total_filled": 3,
                    "fill_stats": {"서버명": 3}}

        monkeypatch.setattr(og, "save_form_memory_entries", _fake_save)
        monkeypatch.setattr(og, "_generate_text_response", _fake_text)
        monkeypatch.setattr(og, "_generate_document_file", _fake_doc)
        state = {
            "organized_data": {"rows": [{"서버명": "w1"}], "column_mapping": None,
                               "resolved_mapping": None, "is_sufficient": True,
                               "summary": "", "sheet_mappings": None},
            "parsed_requirements": {"output_format": "xlsx", "original_query": "양식 채워줘"},
            "template_structure": _tpl(["서버명"]),
            "uploaded_file": b"xlsx",
            "form_fill_remember": True,
            "form_fill_overrides": {
                "도입일자": {"action": "column", "value": "ctime", "applied": True, "origin": "answer"},
                "용도": {"action": "literal", "value": "웹서버", "applied": True, "origin": "memory"},
                "구분": {"action": "column", "value": "없는칼럼", "applied": False, "origin": "answer"},
            },
        }
        cfg = SimpleNamespace(query=SimpleNamespace(form_memory_ttl_days=7))
        result = await og.output_generator(state, llm=object(), app_config=cfg)
        # 적용된 answer-origin만 저장 — memory 재적용분·검증 탈락분 제외(C3)
        assert set(captured["entries"]) == {"도입일자"}
        assert "[기억 저장]" in result["final_response"]

    async def test_no_save_without_opt_in(self, monkeypatch):
        import importlib

        og = importlib.import_module("src.nodes.output_generator")

        called: list = []

        async def _fake_save(*a, **k):
            called.append(1)
            return "x"

        async def _fake_text(*a, **k):
            return "본문"

        def _fake_doc(state, fmt):
            return {"file_bytes": b"x", "file_name": "r.xlsx", "total_filled": 1,
                    "fill_stats": {"서버명": 1}}

        monkeypatch.setattr(og, "save_form_memory_entries", _fake_save)
        monkeypatch.setattr(og, "_generate_text_response", _fake_text)
        monkeypatch.setattr(og, "_generate_document_file", _fake_doc)
        state = {
            "organized_data": {"rows": [{"서버명": "w1"}], "column_mapping": None,
                               "resolved_mapping": None, "is_sufficient": True,
                               "summary": "", "sheet_mappings": None},
            "parsed_requirements": {"output_format": "xlsx"},
            "template_structure": _tpl(["서버명"]),
            "uploaded_file": b"xlsx",
            "form_fill_overrides": {
                "도입일자": {"action": "column", "value": "ctime", "applied": True, "origin": "answer"},
            },
        }
        cfg = SimpleNamespace(query=SimpleNamespace(form_memory_ttl_days=7))
        await og.output_generator(state, llm=object(), app_config=cfg)
        assert not called


class TestFormMemoryCommands:
    """③.45 조회·삭제 pre-check — 결정적 단락(LLM 미호출), direct_response."""

    def _state(self):
        return {
            "user_query": "",
            "template_structure": _tpl(["서버명", "도입일자"], title="서버 목록"),
            # FIX-21 계약: 업로드 턴은 field_mapper가 mapped_db_ids를 항상 세팅한다 —
            # 이력 명령(②.7)이 ③(mapped_db_ids)·②.5(selected_db_ids)보다 선행해야
            # 채우기로 오탈취되지 않는다(라이브 실측 2026-08-03: B0 채움 시도).
            "mapped_db_ids": ["polestar_b0"],
            "selected_db_ids": ["polestar_cm_yd"],
        }

    async def _run(self, monkeypatch, query, answers, deleted_capture=None):
        import src.schema_cache.form_memory as fm
        from src.orchestration.intent_planner import intent_planner

        async def _fake_load(template, app_config, touch=True, signature=None):
            meta = {"display_name": "서버 목록", "created_at": "2026-07-31T10:00:00",
                    "use_count": 2} if answers else None
            return "sig", dict(answers), meta

        async def _fake_delete(template, app_config, field_names=None, signature=None):
            if deleted_capture is not None:
                deleted_capture.append(field_names)
            n = len(field_names) if field_names else len(answers)
            return n, "서버 목록"

        monkeypatch.setattr(fm, "load_form_memory_answers", _fake_load)
        monkeypatch.setattr(fm, "delete_form_memory_entries", _fake_delete)
        state = self._state()
        state["user_query"] = query
        from unittest.mock import AsyncMock

        llm = AsyncMock()
        result = await intent_planner(
            state, llm=llm,
            app_config=SimpleNamespace(query=SimpleNamespace(form_memory_ttl_days=7)),
        )
        llm.ainvoke.assert_not_called()
        return result["task_plan"][0]

    async def test_view_shows_entries(self, monkeypatch):
        t = await self._run(
            monkeypatch, "이 양식에 기억된 답 보여줘",
            {"도입일자": {"action": "column", "value": "ctime", "origin": "memory"}},
        )
        assert t["agent"] == "general_inference"
        assert "도입일자" in t["direct_response"]
        assert "기억된 답이 1건 있습니다" in t["direct_response"]

    async def test_delete_specific_field(self, monkeypatch):
        captured: list = []
        t = await self._run(
            monkeypatch, "도입일자 기억 삭제해줘",
            {"도입일자": {"action": "column", "value": "ctime", "origin": "memory"},
             "용도": {"action": "literal", "value": "웹서버", "origin": "memory"}},
            deleted_capture=captured,
        )
        assert captured == [["도입일자"]]
        assert "삭제했습니다" in t["direct_response"]

    async def test_delete_all_requires_all_keyword(self, monkeypatch):
        captured: list = []
        t = await self._run(
            monkeypatch, "이 양식 기억 전부 삭제해줘",
            {"도입일자": {"action": "column", "value": "ctime", "origin": "memory"}},
            deleted_capture=captured,
        )
        assert captured == [None]
        assert "모두 삭제" in t["direct_response"]

    async def test_delete_unmatched_does_not_delete(self, monkeypatch):
        captured: list = []
        t = await self._run(
            monkeypatch, "그거 기억 지워줘",
            {"도입일자": {"action": "column", "value": "ctime", "origin": "memory"}},
            deleted_capture=captured,
        )
        assert captured == []  # 특정 실패 — 침묵 오삭제 금지
        assert "특정하지 못했습니다" in t["direct_response"]

    async def test_view_empty_memory_guides(self, monkeypatch):
        t = await self._run(monkeypatch, "이 양식 기억 보여줘", {})
        assert "기억된 답이 없습니다" in t["direct_response"]

    def test_memory_command_skips_file_zone_clarification(self):
        """FIX-20: 이력 조회·삭제는 DB 조회가 없어 파일 경로 존 역질문을 스킵해야
        ③.45에 도달한다(라이브 실측 2026-08-03: 존 역질문이 조회를 가로챔)."""
        from src.api.routes.query import _file_zone_clarification_or_none
        from src.orchestration.intent_planner import is_form_memory_command

        assert is_form_memory_command("이 양식에 기억된 답 보여줘")
        assert is_form_memory_command("도입일자 기억 삭제해줘")
        assert not is_form_memory_command("여의도 서버들에 대해 양식을 채우시오")

        cfg = SimpleNamespace(multi_db=SimpleNamespace(get_active_db_ids=lambda: []))
        assert _file_zone_clarification_or_none("이 양식에 기억된 답 보여줘", None, cfg) is None
        assert _file_zone_clarification_or_none("기억 전부 삭제", None, cfg) is None
        # 일반 폼필(존 미지정)은 기존대로 역질문 발동
        assert _file_zone_clarification_or_none("양식 채워줘", None, cfg) is not None

    def test_memory_device_term_not_hijacked(self):
        """'(주)기억장치'(메모리 관용 명사)는 이력 명령이 아니다 — 정상 폼필이
        존 역질문 스킵·이력 조회로 오탈취되지 않아야 한다(FIX-20 사이드이펙트 교정)."""
        from src.api.routes.query import _file_zone_clarification_or_none
        from src.orchestration.intent_planner import is_form_memory_command

        assert not is_form_memory_command("이 양식으로 주기억장치 사용현황 보여줘")
        assert not is_form_memory_command("기억장치 사용률 조회")
        # 진짜 이력 명령은 여전히 판정
        assert is_form_memory_command("주기억장치 양식에 기억된 답 보여줘")

        cfg = SimpleNamespace(multi_db=SimpleNamespace(get_active_db_ids=lambda: []))
        # 주기억장치 폼필(존 미지정)은 존 역질문이 정상 발동해야 한다
        assert _file_zone_clarification_or_none(
            "이 양식으로 주기억장치 사용현황 보여줘", None, cfg
        ) is not None

    async def test_replanner_skips_direct_response_tasks(self):
        """FIX-22: 결정적 direct_response(이력 조회 등)는 최종 응답 — replanner LLM
        재평가가 '내용 미제공' 오판으로 DB 조회 후속을 만들면 채우기로 회귀한다
        (라이브 실측 2026-08-03)."""
        from unittest.mock import AsyncMock

        from src.orchestration.replanner import replanner

        llm = AsyncMock()
        state = {
            "user_query": "이 양식에 기억된 답 보여줘",
            "replan_count": 0,
            "task_plan": [{
                "task_id": "t1", "agent": "general_inference",
                "sub_query": "이 양식에 기억된 답 보여줘",
                "direct_response": "'서버 목록'에 기억된 답 7건 ...",
                "status": "completed",
            }],
            "task_results": {"t1": {"final_response": "'서버 목록'에 기억된 답 7건 ..."}},
        }
        result = await replanner(
            state, llm=llm,
            app_config=SimpleNamespace(max_replan=2),
        )
        assert result["needs_replan"] is False
        llm.ainvoke.assert_not_called()

    async def test_no_file_delete_uses_last_form_signature(self, monkeypatch):
        """FIX-23: 파일 재첨부 없는 삭제가 직전 양식 시그니처로 실제 실행되어야 한다
        (라이브 실측 2026-08-03: 커버리지 밖 → LLM 환각 '삭제했다' + Redis 잔존)."""
        import src.schema_cache.form_memory as fm
        from unittest.mock import AsyncMock

        from src.orchestration.intent_planner import intent_planner

        captured: list = []

        async def _fake_load(template, app_config, touch=True, signature=None):
            return signature, {"구분": {"action": "literal", "value": "K리전(공동존)", "origin": "memory"}}, \
                {"display_name": "서버 목록", "created_at": "2026-07-31", "use_count": 1}

        async def _fake_delete(template, app_config, field_names=None, signature=None):
            captured.append((signature, field_names))
            return 1, "서버 목록"

        monkeypatch.setattr(fm, "load_form_memory_answers", _fake_load)
        monkeypatch.setattr(fm, "delete_form_memory_entries", _fake_delete)
        llm = AsyncMock()
        state = {
            "user_query": "구분 기억 삭제",
            "template_structure": None,           # 파일 재첨부 없음
            "last_form_signature": "sig-prev",    # 직전 양식 턴이 보존
        }
        result = await intent_planner(
            state, llm=llm,
            app_config=SimpleNamespace(query=SimpleNamespace(form_memory_ttl_days=7)),
        )
        llm.ainvoke.assert_not_called()
        assert captured == [("sig-prev", ["구분"])]
        assert "삭제했습니다" in result["task_plan"][0]["direct_response"]
        assert result["last_form_signature"] == "sig-prev"  # 컨텍스트 유지

    async def test_no_file_no_signature_guides_instead_of_hallucinating(self):
        """FIX-23: 시그니처도 없으면 LLM으로 새지 않고 결정적 안내(환각 성공 차단)."""
        from unittest.mock import AsyncMock

        from src.orchestration.intent_planner import intent_planner

        llm = AsyncMock()
        state = {"user_query": "구분 기억 삭제", "template_structure": None}
        result = await intent_planner(
            state, llm=llm,
            app_config=SimpleNamespace(query=SimpleNamespace(form_memory_ttl_days=7)),
        )
        llm.ainvoke.assert_not_called()
        assert "양식 파일을" in result["task_plan"][0]["direct_response"]

    async def test_fill_turn_preserves_signature(self):
        """FIX-23: 채우기 턴(③ mapped_db_ids 경유 포함)이 last_form_signature를 보존한다."""
        from unittest.mock import AsyncMock

        from src.orchestration.intent_planner import intent_planner
        from src.utils.schema_utils import form_signature

        tpl = _tpl(["서버명", "도입일자"], title="서버 목록")
        llm = AsyncMock()
        state = {
            "user_query": "여의도 서버들에 대해 양식을 채우시오",
            "template_structure": tpl,
            "mapped_db_ids": ["polestar_cm_yd"],
        }
        result = await intent_planner(
            state, llm=llm,
            app_config=SimpleNamespace(query=SimpleNamespace(form_memory_ttl_days=7)),
        )
        llm.ainvoke.assert_not_called()
        assert result["task_plan"][0]["agent"] == "data_query"
        assert result["last_form_signature"] == form_signature(tpl)

    async def test_delete_save_back_failure_reports_zero(self, monkeypatch):
        """FIX-23: 필드 삭제 저장-백 실패는 '삭제 성공'으로 보고되면 안 된다."""
        import src.schema_cache.form_memory as fm

        class _FailSaveRedis(_FakeRedis):
            async def save_form_memory(self, signature, data, ttl_seconds):
                return False  # 저장 실패

        fake = _FailSaveRedis()
        fake.store["s1"] = {"display_name": "d", "fields": {
            "a": {"action": "blank", "value": None},
            "b": {"action": "blank", "value": None},
        }}

        async def _fake_get_redis(app_config):
            return fake

        monkeypatch.setattr(fm, "_get_redis", _fake_get_redis)
        removed, _ = await fm.delete_form_memory_entries(
            None, _CFG7, ["a"], signature="s1",
        )
        assert removed == 0

    async def test_field_mapper_skips_memory_command_turn(self):
        """FIX-24: 이력 명령 턴은 field_mapper 매핑 자체를 스킵 — 위치어 없는 질의의
        전 DB 유사어 LLM 발견(413 재시도 수십 초)이 낭비 실행되던 라이브 실측 교정."""
        from src.nodes.field_mapper import field_mapper

        state = {
            "user_query": "이 양식에 저장된 값들을 보여줘",
            "template_structure": _tpl(["서버명", "도입일자"]),
            "parsed_requirements": {},
        }
        # llm=None: 스킵 경로는 LLM 생성 전에 반환해야 한다(도달 시 create_llm 시도로 실패)
        result = await field_mapper(state, llm=None, app_config=None)
        assert result["current_node"] == "field_mapper"
        assert result.get("column_mapping") is None
        assert result.get("mapped_db_ids") is None


class TestWriterShortKeySubstringGuard:
    """FIX-25 — 비고=IP값 근본 원인 재현·차단 (라이브 실측 2026-08-03 아티팩트 기반).

    SQL에 AS "비고"가 없고 CSV도 4칼럼뿐인데 Excel 비고가 IP로 채워졌다 — writer의
    부분 매칭 폴백이 행 키 "IP"(2글자)를 'descr**ip**tion'·'**ip**address' 등 거의
    모든 매핑 문자열에 매칭한 것."""

    _ROW = {"서버명": "web-01", "IP": "10.0.0.1",
            "OS(버전정보)": "RHEL8", "제조사(모델명)": "Dell(R740)"}

    def test_short_key_never_substring_matched(self):
        from src.document.excel_writer import _get_value_from_row

        # 낡은 비고 매핑 후보 전부 — 어느 것도 'IP' 키에 오매칭되면 안 된다
        for col in ("polestar.lvw_logical_link_topology.description",
                    "cmm_resource.ipaddress", "description", "IPAM_INFO.DESCRIPTION"):
            assert _get_value_from_row(dict(self._ROW), col, None) is None, col

    def test_legit_ip_fill_survives_via_reverse_mapping(self):
        from src.document.excel_writer import _get_value_from_row

        reverse = {"cmm_resource.ipaddress": "IP"}
        assert _get_value_from_row(dict(self._ROW), "cmm_resource.ipaddress", reverse) == "10.0.0.1"

    def test_dropped_fields_force_mapping_none_single(self):
        """FIX-15 제외 필드는 state 매핑도 None(불변식) — writer 오채움 원천 차단."""
        from src.nodes.query_generator import _try_build_form_fill_pivot_sql

        state = {
            "column_mapping": {
                "서버명": "cmm_resource.name",
                "비고": "polestar.lvw_logical_link_topology.description",  # 타 테이블 → 제외
            },
            "schema_info": {
                "tables": {"cmm_resource": {"columns": [{"name": "name", "type": "text"}]}},
                **_EAV_META,
            },
            "template_structure": {"sheets": [{"title_text": "서버 목록", "name": "Sheet1"}]},
            "active_db_engine": "postgresql",
            "active_db_id": "",
        }
        result = _try_build_form_fill_pivot_sql(state, 1000, "양식 채워줘")
        assert result is not None
        assert "비고" not in result["sql"]
        assert result["mapping_updates"]["비고"] is None
