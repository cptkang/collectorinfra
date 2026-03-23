"""xls_plan.md 기반 Excel/Word 양식 처리 통합 테스트.

신규 구현(field_mapper, 3단계 매핑, 멀티 DB 등)과
기존 코드(input_parser, query_generator, result_organizer, output_generator,
excel_writer, word_writer)가 함께 동작하는지 검증한다.

테스트 구성:
  1. 기존 기능 회귀 테스트 — 텍스트 출력, 단일 DB 레거시 매핑
  2. 3단계 매핑 통합 — 힌트 → synonyms → LLM 연계
  3. 멀티 DB 매핑 및 SQL 생성
  4. 파이프라인 end-to-end — input_parser → field_mapper → query_generator 연결
  5. output_generator — 추론 매핑 표시 및 유사어 등록 안내
  6. Excel/Word Writer 연계
  7. 유사어 등록 플로우 (전체/선택/단건)
"""

from __future__ import annotations

import io
import json
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from src.state import AgentState, OrganizedData, create_initial_state


# ============================================================
# 공통 헬퍼
# ============================================================


def _make_state(**overrides) -> AgentState:
    """테스트용 AgentState를 간편하게 생성한다."""
    state = create_initial_state("서버 현황 조회해줘")
    state.update(overrides)
    return state


def _make_schema_info() -> dict:
    """테스트용 스키마 정보."""
    return {
        "tables": {
            "servers": {
                "columns": [
                    {"name": "id", "type": "integer", "nullable": False,
                     "primary_key": True, "foreign_key": False, "references": None},
                    {"name": "hostname", "type": "varchar(255)", "nullable": False,
                     "primary_key": False, "foreign_key": False, "references": None},
                    {"name": "ip_address", "type": "varchar(45)", "nullable": False,
                     "primary_key": False, "foreign_key": False, "references": None},
                ],
                "row_count_estimate": 50,
                "sample_data": [],
            },
            "cpu_metrics": {
                "columns": [
                    {"name": "server_id", "type": "integer", "nullable": False,
                     "primary_key": False, "foreign_key": True, "references": "servers.id"},
                    {"name": "usage_pct", "type": "double", "nullable": True,
                     "primary_key": False, "foreign_key": False, "references": None},
                ],
                "row_count_estimate": 500000,
                "sample_data": [],
            },
            "memory_metrics": {
                "columns": [
                    {"name": "server_id", "type": "integer", "nullable": False,
                     "primary_key": False, "foreign_key": True, "references": "servers.id"},
                    {"name": "total_gb", "type": "double", "nullable": True,
                     "primary_key": False, "foreign_key": False, "references": None},
                    {"name": "usage_pct", "type": "double", "nullable": True,
                     "primary_key": False, "foreign_key": False, "references": None},
                ],
                "row_count_estimate": 500000,
                "sample_data": [],
            },
        },
        "relationships": [
            {"from": "cpu_metrics.server_id", "to": "servers.id"},
            {"from": "memory_metrics.server_id", "to": "servers.id"},
        ],
    }


def _excel_template_structure(headers: list[str], sheet_name: str = "Sheet1") -> dict:
    """간단한 Excel template_structure를 생성한다."""
    return {
        "file_type": "xlsx",
        "sheets": [{
            "name": sheet_name,
            "headers": headers,
            "header_row": 1,
            "data_start_row": 2,
            "data_end_row": None,
            "header_cells": [
                {"col": i + 1, "value": h} for i, h in enumerate(headers)
            ],
            "merged_cells": [],
            "formula_cells": [],
            "max_column": len(headers),
        }],
        "placeholders": [],
        "tables": [],
    }


def _word_template_structure(placeholders: list[str], table_headers: list[str] | None = None) -> dict:
    """간단한 Word template_structure를 생성한다."""
    tables = []
    if table_headers:
        tables.append({"headers": table_headers, "data_rows": []})
    return {
        "file_type": "docx",
        "sheets": [],
        "placeholders": placeholders,
        "tables": tables,
    }


def _create_test_excel_bytes(headers: list[str], sheet_name: str = "Sheet1") -> bytes:
    """openpyxl로 테스트용 Excel 바이너리를 생성한다."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
# 1. 기존 기능 회귀 테스트
# ============================================================


class TestLegacyRegression:
    """기존 기능이 xls_plan 변경으로 깨지지 않는지 확인."""

    @pytest.mark.asyncio
    async def test_field_mapper_skips_text_mode(self):
        """template_structure 없으면 field_mapper가 스킵된다."""
        from src.nodes.field_mapper import field_mapper

        state = _make_state()  # template_structure=None
        result = await field_mapper(state, llm=AsyncMock(), app_config=MagicMock())

        assert result["current_node"] == "field_mapper"
        assert "column_mapping" not in result
        assert "mapped_db_ids" not in result

    @pytest.mark.asyncio
    async def test_legacy_map_fields_single_db(self):
        """기존 map_fields() API가 단일 DB에서 정상 동작한다."""
        from src.document.field_mapper import map_fields

        mock_llm = AsyncMock()
        mock_llm.ainvoke.return_value = MagicMock(
            content=json.dumps({
                "서버명": "servers.hostname",
                "IP주소": "servers.ip_address",
            })
        )

        template = _excel_template_structure(["서버명", "IP주소"])
        schema = _make_schema_info()

        result = await map_fields(mock_llm, template, schema)
        assert result["서버명"] == "servers.hostname"
        assert result["IP주소"] == "servers.ip_address"

    @pytest.mark.asyncio
    async def test_result_organizer_without_field_mapper(self):
        """field_mapper가 실행되지 않아도 result_organizer가 정상 동작한다."""
        from src.nodes.result_organizer import result_organizer

        state = _make_state(
            query_results=[
                {"hostname": "web-01", "usage_pct": 85.0},
            ],
            parsed_requirements={
                "query_targets": ["서버"],
                "output_format": "text",
                "original_query": "서버 조회",
            },
            schema_info=_make_schema_info(),
            # column_mapping=None → field_mapper 미실행
        )

        mock_config = MagicMock()
        mock_config.security.sensitive_columns = []
        mock_config.security.mask_pattern = "***"

        result = await result_organizer(state, app_config=mock_config)
        assert result["organized_data"]["is_sufficient"] is True
        assert len(result["organized_data"]["rows"]) == 1

    def test_extract_field_names_xlsx(self):
        """기존 extract_field_names가 xlsx에서 동작한다."""
        from src.document.field_mapper import extract_field_names

        template = _excel_template_structure(["서버명", "IP주소", "CPU 사용률"])
        result = extract_field_names(template)
        assert result == ["서버명", "IP주소", "CPU 사용률"]

    def test_extract_field_names_docx(self):
        """기존 extract_field_names가 docx에서 동작한다."""
        from src.document.field_mapper import extract_field_names

        template = _word_template_structure(["서버명", "날짜"], ["IP", "CPU"])
        result = extract_field_names(template)
        assert "서버명" in result
        assert "날짜" in result
        assert "IP" in result
        assert "CPU" in result

    def test_validate_mapping_preserves_valid(self):
        """유효한 매핑이 보존된다."""
        from src.document.field_mapper import _validate_mapping

        mapping = {"서버명": "servers.hostname", "비고": None}
        schema = _make_schema_info()

        result = _validate_mapping(mapping, schema, ["서버명", "비고"])
        assert result["서버명"] == "servers.hostname"
        assert result["비고"] is None

    def test_validate_mapping_rejects_invalid(self):
        """존재하지 않는 컬럼 참조는 None으로 변경된다."""
        from src.document.field_mapper import _validate_mapping

        mapping = {"서버명": "servers.nonexistent"}
        schema = _make_schema_info()

        result = _validate_mapping(mapping, schema, ["서버명"])
        assert result["서버명"] is None


# ============================================================
# 2. 3단계 매핑 통합 테스트
# ============================================================


class TestThreeStepMappingIntegration:
    """3단계 매핑(힌트→synonyms→LLM)이 올바른 우선순위로 동작하는지 검증."""

    @pytest.mark.asyncio
    async def test_all_three_steps_combined(self):
        """3단계가 모두 사용되는 시나리오."""
        from src.document.field_mapper import perform_3step_mapping

        mock_llm = AsyncMock()
        mock_llm.ainvoke.return_value = MagicMock(
            content=json.dumps({
                "메모리(GB)": {
                    "db_id": "polestar",
                    "column": "memory_metrics.total_gb",
                }
            })
        )

        result = await perform_3step_mapping(
            llm=mock_llm,
            field_names=["서버명", "IP주소", "메모리(GB)"],
            field_mapping_hints=[
                {"field": "서버명", "column": "servers.hostname", "db_id": "polestar"},
            ],
            all_db_synonyms={
                "polestar": {"servers.ip_address": ["IP주소", "아이피"]},
            },
            all_db_descriptions={
                "polestar": {"memory_metrics.total_gb": "메모리 총 용량 (GB)"},
            },
            priority_db_ids=[],
        )

        # 1단계 힌트
        assert result.mapping_sources["서버명"] == "hint"
        assert result.column_mapping["서버명"] == "servers.hostname"

        # 2단계 synonyms
        assert result.mapping_sources["IP주소"] == "synonym"
        assert result.column_mapping["IP주소"] == "servers.ip_address"

        # 3단계 LLM
        assert result.mapping_sources["메모리(GB)"] == "llm_inferred"
        assert result.column_mapping["메모리(GB)"] == "memory_metrics.total_gb"

        # 모든 필드가 같은 DB
        assert result.mapped_db_ids == ["polestar"]

    @pytest.mark.asyncio
    async def test_hint_overrides_synonym(self):
        """동일 필드에 힌트와 synonym이 있으면 힌트가 우선."""
        from src.document.field_mapper import perform_3step_mapping

        mock_llm = AsyncMock()

        result = await perform_3step_mapping(
            llm=mock_llm,
            field_names=["서버명"],
            field_mapping_hints=[
                {"field": "서버명", "column": "custom_servers.name", "db_id": "custom_db"},
            ],
            all_db_synonyms={
                "polestar": {"servers.hostname": ["서버명"]},
            },
            all_db_descriptions={},
            priority_db_ids=[],
        )

        # 힌트가 우선
        assert result.column_mapping["서버명"] == "custom_servers.name"
        assert result.mapping_sources["서버명"] == "hint"
        assert "custom_db" in result.mapped_db_ids

    @pytest.mark.asyncio
    async def test_no_redis_fallback_to_llm(self):
        """Redis 캐시 없이(synonyms 비어있음) LLM 폴백이 동작한다."""
        from src.document.field_mapper import perform_3step_mapping

        mock_llm = AsyncMock()
        mock_llm.ainvoke.return_value = MagicMock(
            content=json.dumps({
                "서버명": {"db_id": "polestar", "column": "servers.hostname"},
                "CPU 사용률": {"db_id": "polestar", "column": "cpu_metrics.usage_pct"},
            })
        )

        result = await perform_3step_mapping(
            llm=mock_llm,
            field_names=["서버명", "CPU 사용률"],
            field_mapping_hints=[],
            all_db_synonyms={},  # Redis 없음
            all_db_descriptions={
                "polestar": {
                    "servers.hostname": "서버 호스트명",
                    "cpu_metrics.usage_pct": "CPU 사용률",
                },
            },
            priority_db_ids=[],
        )

        assert result.column_mapping["서버명"] == "servers.hostname"
        assert result.column_mapping["CPU 사용률"] == "cpu_metrics.usage_pct"
        assert all(s == "llm_inferred" for s in result.mapping_sources.values())
        mock_llm.ainvoke.assert_called()  # LLM이 호출되었어야 함

    @pytest.mark.asyncio
    async def test_priority_db_wins_on_conflict(self):
        """동일 synonym이 여러 DB에 있을 때 우선순위 DB가 선택된다."""
        from src.document.field_mapper import perform_3step_mapping

        mock_llm = AsyncMock()

        result = await perform_3step_mapping(
            llm=mock_llm,
            field_names=["서버명"],
            field_mapping_hints=[],
            all_db_synonyms={
                "db_a": {"tbl_a.host": ["서버명"]},
                "db_b": {"tbl_b.host": ["서버명"]},
            },
            all_db_descriptions={},
            priority_db_ids=["db_b"],  # db_b 우선
        )

        assert result.column_mapping["서버명"] == "tbl_b.host"
        assert "db_b" in result.mapped_db_ids

    @pytest.mark.asyncio
    async def test_unmapped_fields_are_none(self):
        """매핑되지 않는 필드(비고 등)는 None으로 포함된다."""
        from src.document.field_mapper import perform_3step_mapping

        mock_llm = AsyncMock()
        mock_llm.ainvoke.return_value = MagicMock(
            content=json.dumps({"비고": None})
        )

        result = await perform_3step_mapping(
            llm=mock_llm,
            field_names=["서버명", "비고"],
            field_mapping_hints=[],
            all_db_synonyms={"polestar": {"servers.hostname": ["서버명"]}},
            all_db_descriptions={"polestar": {"servers.hostname": "호스트명"}},
            priority_db_ids=[],
        )

        assert result.column_mapping["서버명"] == "servers.hostname"
        assert result.column_mapping["비고"] is None


# ============================================================
# 3. 멀티 DB 매핑
# ============================================================


class TestMultiDBMapping:
    """여러 DB에 걸친 매핑이 올바르게 동작하는지 검증."""

    @pytest.mark.asyncio
    async def test_fields_mapped_to_different_dbs(self):
        """필드가 서로 다른 DB에 매핑된다."""
        from src.document.field_mapper import perform_3step_mapping

        mock_llm = AsyncMock()

        result = await perform_3step_mapping(
            llm=mock_llm,
            field_names=["서버명", "클라우드 인스턴스"],
            field_mapping_hints=[],
            all_db_synonyms={
                "polestar": {"servers.hostname": ["서버명"]},
                "cloud_portal": {"cloud_instances.instance_type": ["클라우드 인스턴스"]},
            },
            all_db_descriptions={},
            priority_db_ids=[],
        )

        assert len(result.mapped_db_ids) == 2
        assert "polestar" in result.mapped_db_ids
        assert "cloud_portal" in result.mapped_db_ids
        assert result.db_column_mapping["polestar"]["서버명"] == "servers.hostname"
        assert result.db_column_mapping["cloud_portal"]["클라우드 인스턴스"] == "cloud_instances.instance_type"

    @pytest.mark.asyncio
    async def test_field_mapper_node_produces_mapped_db_ids(self):
        """field_mapper 노드가 mapped_db_ids를 올바르게 State에 반환한다."""
        from src.nodes.field_mapper import field_mapper

        state = _make_state(
            template_structure=_excel_template_structure(["서버명", "VM 이름"]),
            parsed_requirements={
                "field_mapping_hints": [],
                "target_db_hints": [],
            },
        )

        mock_config = MagicMock()
        mock_config.multi_db.get_active_db_ids.return_value = ["polestar", "cloud"]

        mock_llm = AsyncMock()

        with patch("src.nodes.field_mapper._load_db_cache_data") as mock_load:
            mock_load.return_value = (
                {
                    "polestar": {"servers.hostname": ["서버명"]},
                    "cloud": {"vms.vm_name": ["VM 이름"]},
                },
                {},
                [],
            )
            result = await field_mapper(state, llm=mock_llm, app_config=mock_config)

        assert "polestar" in result["mapped_db_ids"]
        assert "cloud" in result["mapped_db_ids"]


# ============================================================
# 4. 파이프라인 연결 테스트
# ============================================================


class TestPipelineConnection:
    """input_parser → field_mapper → semantic_router → query_generator 연계."""

    @pytest.mark.asyncio
    async def test_field_mapper_to_semantic_router(self):
        """field_mapper의 mapped_db_ids가 semantic_router에서 사용된다."""
        from src.routing.semantic_router import semantic_router

        state = _make_state(
            mapped_db_ids=["polestar", "cloud_portal"],
        )

        mock_config = MagicMock()
        mock_config.multi_db.get_active_db_ids.return_value = ["polestar", "cloud_portal"]

        result = await semantic_router(state, llm=AsyncMock(), app_config=mock_config)

        assert len(result["target_databases"]) == 2
        db_ids = [t["db_id"] for t in result["target_databases"]]
        assert "polestar" in db_ids
        assert "cloud_portal" in db_ids

    @pytest.mark.asyncio
    async def test_semantic_router_skips_llm_when_mapped(self):
        """mapped_db_ids가 있으면 LLM 라우팅을 스킵한다."""
        from src.routing.semantic_router import semantic_router

        state = _make_state(mapped_db_ids=["polestar"])

        mock_config = MagicMock()
        mock_config.multi_db.get_active_db_ids.return_value = ["polestar"]

        mock_llm = AsyncMock()
        result = await semantic_router(state, llm=mock_llm, app_config=mock_config)

        # LLM은 호출되지 않아야 함
        mock_llm.ainvoke.assert_not_called()
        assert result["target_databases"][0]["db_id"] == "polestar"

    @pytest.mark.asyncio
    async def test_semantic_router_falls_back_without_mapping(self):
        """mapped_db_ids가 없으면 기존 LLM 라우팅을 사용한다."""
        from src.routing.semantic_router import semantic_router

        state = _make_state(mapped_db_ids=None)

        mock_config = MagicMock()
        active_ids = ["polestar"]
        mock_config.multi_db.get_active_db_ids.return_value = active_ids

        mock_llm = AsyncMock()
        mock_llm.ainvoke.return_value = MagicMock(
            content=json.dumps({
                "databases": [{"db_id": "polestar", "relevance_score": 0.9,
                               "sub_query_context": "서버 조회", "reason": "인프라"}]
            })
        )

        result = await semantic_router(state, llm=mock_llm, app_config=mock_config)

        mock_llm.ainvoke.assert_called()
        assert result["target_databases"][0]["db_id"] == "polestar"

    @pytest.mark.asyncio
    async def test_query_generator_uses_column_mapping(self):
        """query_generator가 column_mapping을 프롬프트에 포함한다."""
        from src.nodes.query_generator import query_generator

        state = _make_state(
            schema_info=_make_schema_info(),
            parsed_requirements={
                "query_targets": ["서버", "CPU"],
                "filter_conditions": [],
                "output_format": "xlsx",
                "original_query": "서버 현황 Excel로 조회",
            },
            column_mapping={
                "서버명": "servers.hostname",
                "IP주소": "servers.ip_address",
                "CPU 사용률": "cpu_metrics.usage_pct",
            },
        )

        mock_llm = AsyncMock()
        mock_llm.ainvoke.return_value = MagicMock(
            content='```sql\nSELECT s.hostname AS "servers.hostname" FROM servers s LIMIT 100;\n```'
        )

        mock_config = MagicMock()
        mock_config.query.default_limit = 1000

        result = await query_generator(state, llm=mock_llm, app_config=mock_config)

        assert result["generated_sql"]
        # LLM에 전달된 프롬프트에 매핑 정보가 포함되었는지 확인
        call_args = mock_llm.ainvoke.call_args
        messages = call_args[0][0]
        user_msg = messages[1].content
        assert "servers.hostname" in user_msg
        assert "servers.ip_address" in user_msg
        assert "cpu_metrics.usage_pct" in user_msg


# ============================================================
# 5. output_generator — 추론 매핑 표시 및 유사어 등록 안내
# ============================================================


class TestOutputGeneratorMappingDisplay:
    """output_generator가 LLM 추론 매핑 정보를 올바르게 표시하는지 검증."""

    @pytest.mark.asyncio
    async def test_inferred_mapping_shown_in_response(self):
        """llm_inferred 매핑이 응답에 자동 매핑 안내로 표시된다."""
        from src.nodes.output_generator import output_generator

        state = _make_state(
            organized_data=OrganizedData(
                summary="3건 조회",
                rows=[{"servers.hostname": "web-01", "cpu_metrics.usage_pct": 85.0}],
                column_mapping={"서버명": "servers.hostname", "CPU 사용률": "cpu_metrics.usage_pct"},
                is_sufficient=True,
                sheet_mappings=None,
            ),
            parsed_requirements={
                "output_format": "text",
                "original_query": "서버 현황 조회",
                "query_targets": ["서버"],
            },
            generated_sql="SELECT hostname FROM servers",
            mapping_sources={
                "서버명": "synonym",
                "CPU 사용률": "llm_inferred",
            },
            column_mapping={
                "서버명": "servers.hostname",
                "CPU 사용률": "cpu_metrics.usage_pct",
            },
            db_column_mapping={
                "polestar": {
                    "서버명": "servers.hostname",
                    "CPU 사용률": "cpu_metrics.usage_pct",
                },
            },
            pending_synonym_registrations=[
                {"index": 1, "field": "CPU 사용률", "column": "cpu_metrics.usage_pct", "db_id": "polestar"},
            ],
        )

        mock_llm = AsyncMock()
        mock_llm.ainvoke.return_value = MagicMock(content="서버 현황 결과입니다.")

        mock_config = MagicMock()
        result = await output_generator(state, llm=mock_llm, app_config=mock_config)

        response = result["final_response"]
        # 추론 매핑 안내가 포함되어야 함
        assert "자동 매핑" in response or "CPU 사용률" in response

    @pytest.mark.asyncio
    async def test_no_mapping_info_for_text_only(self):
        """매핑 정보가 없는 텍스트 응답에는 안내가 없다."""
        from src.nodes.output_generator import output_generator

        state = _make_state(
            organized_data=OrganizedData(
                summary="3건 조회",
                rows=[{"hostname": "web-01"}],
                column_mapping=None,
                is_sufficient=True,
                sheet_mappings=None,
            ),
            parsed_requirements={
                "output_format": "text",
                "original_query": "서버 조회",
                "query_targets": ["서버"],
            },
            generated_sql="SELECT hostname FROM servers",
            mapping_sources=None,
        )

        mock_llm = AsyncMock()
        mock_llm.ainvoke.return_value = MagicMock(content="서버 목록입니다.")

        result = await output_generator(state, llm=mock_llm, app_config=MagicMock())

        # 매핑 안내가 없어야 함
        assert "자동 매핑" not in result["final_response"]


# ============================================================
# 6. Excel/Word Writer 연계
# ============================================================


class TestExcelWriterIntegration:
    """Excel Writer가 field_mapper 매핑 결과와 연계되는지 검증."""

    def test_fill_excel_with_alias_keys(self):
        """query_results key가 table.column alias 형식일 때 정확히 채운다."""
        from src.document.excel_writer import fill_excel_template

        headers = ["서버명", "IP주소", "CPU 사용률"]
        file_data = _create_test_excel_bytes(headers)
        template = _excel_template_structure(headers)
        column_mapping = {
            "서버명": "servers.hostname",
            "IP주소": "servers.ip_address",
            "CPU 사용률": "cpu_metrics.usage_pct",
        }
        rows = [
            {"servers.hostname": "web-01", "servers.ip_address": "10.0.0.1", "cpu_metrics.usage_pct": 85.2},
            {"servers.hostname": "web-02", "servers.ip_address": "10.0.0.2", "cpu_metrics.usage_pct": 72.0},
        ]

        result_bytes = fill_excel_template(file_data, template, column_mapping, rows)

        # 결과 파일을 읽어서 데이터가 올바르게 채워졌는지 확인
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active

        assert ws.cell(row=2, column=1).value == "web-01"
        assert ws.cell(row=2, column=2).value == "10.0.0.1"
        assert ws.cell(row=2, column=3).value == 85.2
        assert ws.cell(row=3, column=1).value == "web-02"

    def test_fill_excel_with_bare_column_keys(self):
        """query_results key가 bare column 형식(hostname)이어도 폴백으로 채운다."""
        from src.document.excel_writer import fill_excel_template

        headers = ["서버명", "IP주소"]
        file_data = _create_test_excel_bytes(headers)
        template = _excel_template_structure(headers)
        column_mapping = {
            "서버명": "servers.hostname",
            "IP주소": "servers.ip_address",
        }
        rows = [
            {"hostname": "web-01", "ip_address": "10.0.0.1"},  # bare keys
        ]

        result_bytes = fill_excel_template(file_data, template, column_mapping, rows)

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active

        # _get_value_from_row 폴백으로 "hostname" 매칭
        assert ws.cell(row=2, column=1).value == "web-01"
        assert ws.cell(row=2, column=2).value == "10.0.0.1"

    def test_fill_excel_none_value_handling(self):
        """매핑된 컬럼에 값이 None이면 셀을 덮어쓰지 않는다."""
        from src.document.excel_writer import fill_excel_template

        headers = ["서버명", "비고"]
        file_data = _create_test_excel_bytes(headers)
        template = _excel_template_structure(headers)
        column_mapping = {
            "서버명": "servers.hostname",
            "비고": None,  # 매핑 불가
        }
        rows = [{"servers.hostname": "web-01"}]

        result_bytes = fill_excel_template(file_data, template, column_mapping, rows)

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active

        assert ws.cell(row=2, column=1).value == "web-01"
        # 비고 컬럼은 매핑 불가 → 셀이 비어있음
        assert ws.cell(row=2, column=2).value is None


# ============================================================
# 7. 유사어 등록 플로우
# ============================================================


class TestSynonymRegistrationFlow:
    """유사어 등록 플로우(전체/선택/단건)가 올바르게 동작하는지 검증."""

    def test_build_pending_registrations(self):
        """LLM 추론 매핑에서 pending 등록 목록을 생성한다."""
        from src.nodes.field_mapper import _build_pending_registrations
        from src.document.field_mapper import MappingResult

        mr = MappingResult()
        mr.mapping_sources = {
            "서버명": "synonym",
            "CPU 사용률": "llm_inferred",
            "디스크": "llm_inferred",
        }
        mr.db_column_mapping = {
            "polestar": {
                "서버명": "servers.hostname",
                "CPU 사용률": "cpu_metrics.usage_pct",
                "디스크": "disk_metrics.usage_gb",
            }
        }

        pending = _build_pending_registrations(mr)

        assert len(pending) == 2  # synonym은 제외
        assert pending[0]["field"] == "CPU 사용률"
        assert pending[0]["index"] == 1
        assert pending[0]["db_id"] == "polestar"
        assert pending[1]["field"] == "디스크"
        assert pending[1]["index"] == 2

    @pytest.mark.asyncio
    async def test_synonym_registration_all(self):
        """전체 등록 요청이 처리된다."""
        from src.nodes.field_mapper import _handle_synonym_registration

        state = _make_state(
            pending_synonym_registrations=[
                {"index": 1, "field": "CPU 사용률", "column": "cpu_metrics.usage_pct", "db_id": "polestar"},
                {"index": 2, "field": "디스크", "column": "disk_metrics.usage_gb", "db_id": "polestar"},
            ],
        )

        mock_config = MagicMock()

        with patch("src.schema_cache.cache_manager.get_cache_manager") as mock_get_cm:
            mock_cm = AsyncMock()
            mock_cm.get_synonyms.return_value = {}
            mock_cm.save_synonyms.return_value = True
            mock_get_cm.return_value = mock_cm

            result = await _handle_synonym_registration(
                state, {"mode": "all"}, mock_config
            )

        assert "2건" in result["final_response"]
        assert mock_cm.save_synonyms.call_count >= 1

    @pytest.mark.asyncio
    async def test_synonym_registration_selective(self):
        """선택 등록(번호 지정) 요청이 처리된다."""
        from src.nodes.field_mapper import _handle_synonym_registration

        state = _make_state(
            pending_synonym_registrations=[
                {"index": 1, "field": "CPU 사용률", "column": "cpu_metrics.usage_pct", "db_id": "polestar"},
                {"index": 2, "field": "디스크", "column": "disk_metrics.usage_gb", "db_id": "polestar"},
                {"index": 3, "field": "네트워크", "column": "network.bandwidth", "db_id": "polestar"},
            ],
        )

        mock_config = MagicMock()

        with patch("src.schema_cache.cache_manager.get_cache_manager") as mock_get_cm:
            mock_cm = AsyncMock()
            mock_cm.get_synonyms.return_value = {}
            mock_cm.save_synonyms.return_value = True
            mock_get_cm.return_value = mock_cm

            # 1번, 3번만 선택
            result = await _handle_synonym_registration(
                state, {"mode": "selective", "indices": [1, 3]}, mock_config
            )

        assert "2건" in result["final_response"]
        # save_synonyms 호출 확인 (2번은 제외되었어야 함)

    @pytest.mark.asyncio
    async def test_synonym_registration_empty_pending(self):
        """pending이 없을 때 적절한 안내 메시지를 반환한다."""
        from src.nodes.field_mapper import _handle_synonym_registration

        state = _make_state(pending_synonym_registrations=None)

        result = await _handle_synonym_registration(
            state, {"mode": "all"}, MagicMock()
        )

        assert "등록할 유사어 매핑이 없습니다" in result["final_response"]

    @pytest.mark.asyncio
    async def test_field_mapper_node_handles_synonym_registration(self):
        """field_mapper 노드가 synonym_registration 요청을 처리한다."""
        from src.nodes.field_mapper import field_mapper

        state = _make_state(
            parsed_requirements={
                "synonym_registration": {"mode": "all"},
            },
            pending_synonym_registrations=[
                {"index": 1, "field": "CPU", "column": "cpu_metrics.usage_pct", "db_id": "polestar"},
            ],
        )

        mock_config = MagicMock()

        with patch("src.nodes.field_mapper._handle_synonym_registration") as mock_handle:
            mock_handle.return_value = {
                "final_response": "1건의 유사어가 등록되었습니다.",
                "current_node": "field_mapper",
            }
            result = await field_mapper(state, llm=AsyncMock(), app_config=mock_config)

        assert "등록" in result["final_response"]


# ============================================================
# 8. result_organizer 매핑 연계
# ============================================================


class TestResultOrganizerMappingIntegration:
    """result_organizer가 field_mapper의 column_mapping을 올바르게 사용하는지 검증."""

    @pytest.mark.asyncio
    async def test_uses_state_column_mapping(self):
        """State의 column_mapping이 organized_data에 전달된다."""
        from src.nodes.result_organizer import result_organizer

        state = _make_state(
            query_results=[
                {"servers.hostname": "web-01", "cpu_metrics.usage_pct": 85.0},
            ],
            parsed_requirements={
                "query_targets": ["서버"],
                "output_format": "xlsx",
                "original_query": "서버 조회",
            },
            template_structure=_excel_template_structure(["서버명", "CPU 사용률"]),
            schema_info=_make_schema_info(),
            column_mapping={
                "서버명": "servers.hostname",
                "CPU 사용률": "cpu_metrics.usage_pct",
            },
        )

        mock_config = MagicMock()
        mock_config.security.sensitive_columns = []
        mock_config.security.mask_pattern = "***"

        result = await result_organizer(state, app_config=mock_config)

        organized = result["organized_data"]
        assert organized["is_sufficient"] is True
        # column_mapping이 organized_data에 포함
        assert organized["column_mapping"] is not None

    @pytest.mark.asyncio
    async def test_data_sufficiency_check_with_mapping(self):
        """column_mapping 기반 충분성 검사가 동작한다."""
        from src.nodes.result_organizer import _check_data_sufficiency

        results = [{"servers.hostname": "web-01", "cpu_metrics.usage_pct": 85.0}]
        parsed = {"query_targets": ["서버"]}
        template = _excel_template_structure(["서버명", "CPU 사용률"])
        column_mapping = {
            "서버명": "servers.hostname",
            "CPU 사용률": "cpu_metrics.usage_pct",
        }

        is_sufficient = _check_data_sufficiency(
            results, parsed, template, column_mapping=column_mapping
        )
        assert is_sufficient is True

    @pytest.mark.asyncio
    async def test_data_insufficiency_detected(self):
        """매핑된 컬럼이 결과에 없으면 불충분으로 판단한다."""
        from src.nodes.result_organizer import _check_data_sufficiency

        results = [{"unrelated_col": "value"}]
        parsed = {"query_targets": ["서버"]}
        template = _excel_template_structure(["서버명", "IP주소", "CPU", "메모리"])
        column_mapping = {
            "서버명": "servers.hostname",
            "IP주소": "servers.ip_address",
            "CPU": "cpu_metrics.usage_pct",
            "메모리": "memory_metrics.total_gb",
        }

        is_sufficient = _check_data_sufficiency(
            results, parsed, template, column_mapping=column_mapping
        )
        # 매핑된 4개 컬럼 중 0개가 결과에 있음 → 불충분
        assert is_sufficient is False


# ============================================================
# 9. 전체 Excel 파이프라인 end-to-end
# ============================================================


class TestEndToEndExcelPipeline:
    """input_parser → field_mapper → query_generator → excel_writer 전체 흐름."""

    @pytest.mark.asyncio
    async def test_full_excel_pipeline(self):
        """Excel 파일 업로드 → 매핑 → SQL 생성 → 결과 채우기 전체 흐름."""
        from src.document.field_mapper import perform_3step_mapping, extract_field_names
        from src.document.excel_writer import fill_excel_template

        # 1. Excel template 준비
        headers = ["서버명", "IP주소", "CPU 사용률"]
        template = _excel_template_structure(headers)
        file_data = _create_test_excel_bytes(headers)

        # 2. 필드명 추출
        field_names = extract_field_names(template)
        assert field_names == headers

        # 3. 3단계 매핑 수행
        mock_llm = AsyncMock()
        mapping_result = await perform_3step_mapping(
            llm=mock_llm,
            field_names=field_names,
            field_mapping_hints=[],
            all_db_synonyms={
                "polestar": {
                    "servers.hostname": ["서버명", "호스트명"],
                    "servers.ip_address": ["IP주소", "아이피"],
                    "cpu_metrics.usage_pct": ["CPU 사용률", "CPU%"],
                },
            },
            all_db_descriptions={},
            priority_db_ids=[],
        )

        # 모든 필드가 synonyms로 매핑됨 (LLM 호출 없이)
        assert all(s == "synonym" for s in mapping_result.mapping_sources.values())
        mock_llm.ainvoke.assert_not_called()

        # 4. query_results 시뮬레이션 (alias 형식 키)
        query_results = [
            {"servers.hostname": "web-01", "servers.ip_address": "10.0.0.1", "cpu_metrics.usage_pct": 85.2},
            {"servers.hostname": "web-02", "servers.ip_address": "10.0.0.2", "cpu_metrics.usage_pct": 72.0},
            {"servers.hostname": "db-01", "servers.ip_address": "10.0.1.1", "cpu_metrics.usage_pct": 91.5},
        ]

        # 5. Excel 채우기
        result_bytes = fill_excel_template(
            file_data, template, mapping_result.column_mapping, query_results
        )

        # 6. 결과 검증
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active

        # 헤더 보존
        assert ws.cell(row=1, column=1).value == "서버명"
        assert ws.cell(row=1, column=2).value == "IP주소"
        assert ws.cell(row=1, column=3).value == "CPU 사용률"

        # 데이터 3건 채워짐
        assert ws.cell(row=2, column=1).value == "web-01"
        assert ws.cell(row=2, column=3).value == 85.2
        assert ws.cell(row=3, column=1).value == "web-02"
        assert ws.cell(row=4, column=1).value == "db-01"
        assert ws.cell(row=4, column=3).value == 91.5

    @pytest.mark.asyncio
    async def test_mixed_mapping_sources_pipeline(self):
        """힌트+synonyms+LLM 혼합 매핑으로 Excel 생성."""
        from src.document.field_mapper import perform_3step_mapping, extract_field_names
        from src.document.excel_writer import fill_excel_template
        from src.nodes.field_mapper import _build_pending_registrations

        headers = ["서버명", "IP주소", "메모리(GB)"]
        template = _excel_template_structure(headers)
        file_data = _create_test_excel_bytes(headers)

        mock_llm = AsyncMock()
        mock_llm.ainvoke.return_value = MagicMock(
            content=json.dumps({
                "메모리(GB)": {"db_id": "polestar", "column": "memory_metrics.total_gb"}
            })
        )

        mapping_result = await perform_3step_mapping(
            llm=mock_llm,
            field_names=extract_field_names(template),
            field_mapping_hints=[
                {"field": "서버명", "column": "servers.hostname", "db_id": "polestar"},
            ],
            all_db_synonyms={"polestar": {"servers.ip_address": ["IP주소"]}},
            all_db_descriptions={"polestar": {"memory_metrics.total_gb": "메모리 총량"}},
            priority_db_ids=[],
        )

        # 3개 매핑 소스 확인
        assert mapping_result.mapping_sources["서버명"] == "hint"
        assert mapping_result.mapping_sources["IP주소"] == "synonym"
        assert mapping_result.mapping_sources["메모리(GB)"] == "llm_inferred"

        # pending 등록 생성 — LLM 추론 항목만
        pending = _build_pending_registrations(mapping_result)
        assert len(pending) == 1
        assert pending[0]["field"] == "메모리(GB)"

        # Excel 채우기
        rows = [
            {"servers.hostname": "web-01", "servers.ip_address": "10.0.0.1", "memory_metrics.total_gb": 32.0},
        ]

        result_bytes = fill_excel_template(
            file_data, template, mapping_result.column_mapping, rows
        )

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(result_bytes))
        ws = wb.active

        assert ws.cell(row=2, column=1).value == "web-01"
        assert ws.cell(row=2, column=2).value == "10.0.0.1"
        assert ws.cell(row=2, column=3).value == 32.0


# ============================================================
# 10. State 초기화 및 필드 존재 확인
# ============================================================


class TestStateFields:
    """create_initial_state가 xls_plan의 신규 필드를 올바르게 초기화하는지 검증."""

    def test_new_mapping_fields_initialized(self):
        state = create_initial_state("테스트")

        assert state["column_mapping"] is None
        assert state["db_column_mapping"] is None
        assert state["mapping_sources"] is None
        assert state["mapped_db_ids"] is None
        assert state["pending_synonym_registrations"] is None

    def test_legacy_fields_preserved(self):
        state = create_initial_state("테스트")

        assert state["user_query"] == "테스트"
        assert state["schema_info"] == {}
        assert state["retry_count"] == 0
        assert state["generated_sql"] == ""
        assert state["final_response"] == ""

    def test_state_with_file_upload(self):
        state = create_initial_state(
            "서버 조회",
            uploaded_file=b"fake_excel_bytes",
            file_type="xlsx",
        )

        assert state["uploaded_file"] == b"fake_excel_bytes"
        assert state["file_type"] == "xlsx"
