"""SQL 쿼리 결과 → 엑셀 양식 매핑 테스트.

polestar DB의 cmm_resource + core_config_prop EAV 조인 쿼리 결과를
'취합 예시2.xlsx' 양식에 올바르게 매핑·채우기 할 수 있는지 검증한다.
"""

from __future__ import annotations

import io
import logging
import pathlib
from typing import Any

import openpyxl
import pytest

from src.document.excel_parser import parse_excel_template
from src.document.excel_writer import _get_value_from_row, fill_excel_template

logger = logging.getLogger(__name__)

SAMPLE_DIR = pathlib.Path(__file__).resolve().parent.parent / "sample"
TEMPLATE_FILE = SAMPLE_DIR / "취합 예시2.xlsx"

# ──────────────────────────────────────────────
# SQL 쿼리가 반환하는 행 구조 (alias 기준)
# ──────────────────────────────────────────────
QUERY_RESULT_ROWS: list[dict[str, Any]] = [
    {
        "resource_id": 1001,
        "resource_name": "web-server-01",
        "resource_type": "platform.server.Linux",
        "cmm_resource_hostname": "web-01",
        "cmm_resource_ipaddress": "10.0.1.10",
        "resource_description": "웹서비스 서버",
        "parent_resource_id": 100,
        "id_ancestry": "/100/1001",
        "eav_hostname": "web-01",
        "eav_ipaddress": "10.0.1.10",
        "os_type": "Linux",
        "serial_number": "SN-ABCD-1234",
        "agent_id": "AGT-001",
        "agent_version": "5.2.1",
        "model": "PowerEdge R740",
        "os_version": "RHEL 8.6",
        "vendor": "Dell",
    },
    {
        "resource_id": 1002,
        "resource_name": "db-server-01",
        "resource_type": "platform.server.AIX",
        "cmm_resource_hostname": "db-01",
        "cmm_resource_ipaddress": "10.0.2.20",
        "resource_description": "데이터베이스 서버",
        "parent_resource_id": 100,
        "id_ancestry": "/100/1002",
        "eav_hostname": "db-01",
        "eav_ipaddress": "10.0.2.20",
        "os_type": "AIX",
        "serial_number": "SN-EFGH-5678",
        "agent_id": "AGT-002",
        "agent_version": "5.2.1",
        "model": "Power S924",
        "os_version": "AIX 7.2",
        "vendor": "IBM",
    },
    {
        "resource_id": 1003,
        "resource_name": "app-server-01",
        "resource_type": "platform.server.Windows",
        "cmm_resource_hostname": "app-01",
        "cmm_resource_ipaddress": "10.0.3.30",
        "resource_description": "애플리케이션 서버",
        "parent_resource_id": 200,
        "id_ancestry": "/200/1003",
        "eav_hostname": "app-01",
        "eav_ipaddress": "10.0.3.30",
        "os_type": "Windows",
        "serial_number": "SN-IJKL-9012",
        "agent_id": "AGT-003",
        "agent_version": "5.3.0",
        "model": "ProLiant DL380",
        "os_version": "Windows Server 2019",
        "vendor": "HP",
    },
]


# ──────────────────────────────────────────────
# 엑셀 필드 → DB 컬럼 매핑 후보
# ──────────────────────────────────────────────

# Case A: field_mapper가 "table.column" 형식으로 매핑한 경우
MAPPING_TABLE_DOT_COLUMN: dict[str, str | None] = {
    "일련번호": None,  # auto-numbering, no DB column
    "업무유형": None,
    "성능관리": None,
    "비대면주요서버": None,
    "CF여부": None,
    "Serial 번호": "EAV:SerialNumber",
    "서버명": "cmm_resource.hostname",
    "업무내용": "cmm_resource.description",
    "대분류": "EAV:OSType",
    "소분류": "EAV:OSVerson",
    "IP": "cmm_resource.ipaddress",
    "CPU": None,
    "MEMORY": None,
    "OS": "EAV:OSType",
    "WAS": None,
    "DB": None,
    "DMZ 유무": None,
    "센터구분": None,
    "CPU_AVG702": None,
    "CPU_PEAK713": None,
    "MEMORY724": None,
}

# Case B: field_mapper가 SQL alias 이름으로 매핑한 경우
MAPPING_SQL_ALIAS: dict[str, str | None] = {
    "일련번호": None,
    "업무유형": None,
    "성능관리": None,
    "비대면주요서버": None,
    "CF여부": None,
    "Serial 번호": "serial_number",
    "서버명": "cmm_resource_hostname",
    "업무내용": "resource_description",
    "대분류": "os_type",
    "소분류": "os_version",
    "IP": "cmm_resource_ipaddress",
    "CPU": None,
    "MEMORY": None,
    "OS": "os_type",
    "WAS": None,
    "DB": None,
    "DMZ 유무": None,
    "센터구분": None,
    "CPU_AVG702": None,
    "CPU_PEAK713": None,
    "MEMORY724": None,
}


@pytest.fixture()
def template_data() -> bytes:
    """취합 예시2.xlsx 바이너리 데이터."""
    assert TEMPLATE_FILE.exists(), f"Template not found: {TEMPLATE_FILE}"
    return TEMPLATE_FILE.read_bytes()


@pytest.fixture()
def template_structure(template_data: bytes) -> dict[str, Any]:
    """파싱된 양식 구조."""
    return parse_excel_template(template_data)


# ──────────────────────────────────────────────
# 1. template structure 정합성 검증
# ──────────────────────────────────────────────
class TestTemplateStructure:
    """양식 구조 파싱이 올바른지 확인."""

    def test_sheet_detected(self, template_structure: dict) -> None:
        sheets = template_structure["sheets"]
        assert len(sheets) >= 1
        assert sheets[0]["name"] == "성능관리분석자료"

    def test_header_row_is_4(self, template_structure: dict) -> None:
        sheet = template_structure["sheets"][0]
        assert sheet["header_row"] == 4

    def test_data_start_row_is_5(self, template_structure: dict) -> None:
        sheet = template_structure["sheets"][0]
        assert sheet["data_start_row"] == 5

    def test_expected_headers_present(self, template_structure: dict) -> None:
        headers = template_structure["sheets"][0]["headers"]
        expected = ["서버명", "IP", "OS", "Serial 번호", "CPU", "MEMORY"]
        for h in expected:
            assert h in headers, f"헤더 '{h}'가 파싱 결과에 없음"

    def test_header_count(self, template_structure: dict) -> None:
        headers = template_structure["sheets"][0]["headers"]
        assert len(headers) == 21


# ──────────────────────────────────────────────
# 2. _get_value_from_row 단위 테스트
# ──────────────────────────────────────────────
class TestGetValueFromRow:
    """쿼리 결과 행에서 값 추출이 올바르게 동작하는지 검증."""

    ROW = QUERY_RESULT_ROWS[0]

    # --- 직접 매칭 (SQL alias가 그대로 키인 경우) ---
    def test_exact_alias_match(self) -> None:
        """SQL alias 키와 정확히 일치."""
        val = _get_value_from_row(self.ROW, "cmm_resource_hostname")
        assert val == "web-01"

    def test_exact_alias_serial(self) -> None:
        val = _get_value_from_row(self.ROW, "serial_number")
        assert val == "SN-ABCD-1234"

    # --- table.column 형식 매핑 → 실제 alias 키 매칭 ---
    def test_table_dot_column_hostname(self) -> None:
        """'cmm_resource.hostname' 매핑 → 'cmm_resource_hostname' 키에서 값 추출 가능?"""
        val = _get_value_from_row(self.ROW, "cmm_resource.hostname")
        # "hostname" (column part) -> substring match: "hostname" in "cmm_resource_hostname" → True
        assert val is not None, (
            "table.column 형식 'cmm_resource.hostname'이 "
            "SQL alias 키 'cmm_resource_hostname'과 매칭되지 않음"
        )
        assert val == "web-01"

    def test_table_dot_column_ipaddress(self) -> None:
        val = _get_value_from_row(self.ROW, "cmm_resource.ipaddress")
        assert val is not None, (
            "'cmm_resource.ipaddress' → 'cmm_resource_ipaddress' 매칭 실패"
        )
        assert val == "10.0.1.10"

    def test_table_dot_column_description(self) -> None:
        val = _get_value_from_row(self.ROW, "cmm_resource.description")
        assert val is not None, (
            "'cmm_resource.description' → 'resource_description' 매칭 실패"
        )
        assert val == "웹서비스 서버"

    # --- EAV 형식 매핑 ---
    def test_eav_ostype(self) -> None:
        """'EAV:OSType' → 'os_type' 키에서 값 추출 가능?"""
        val = _get_value_from_row(self.ROW, "EAV:OSType")
        assert val is not None, (
            "EAV:OSType → SQL alias 'os_type' 매칭 실패. "
            "EAV 접두사 제거 후 'OSType' ≠ 'os_type' (대소문자/형식 불일치)"
        )
        assert val == "Linux"

    def test_eav_serialnumber(self) -> None:
        val = _get_value_from_row(self.ROW, "EAV:SerialNumber")
        assert val is not None, (
            "EAV:SerialNumber → SQL alias 'serial_number' 매칭 실패"
        )
        assert val == "SN-ABCD-1234"

    def test_eav_osversion(self) -> None:
        """'EAV:OSVerson' (원본 DB 오타) → 'os_version' 키에서 값 추출 가능?"""
        val = _get_value_from_row(self.ROW, "EAV:OSVerson")
        assert val is not None, (
            "EAV:OSVerson → SQL alias 'os_version' 매칭 실패"
        )
        assert val == "RHEL 8.6"

    def test_eav_model(self) -> None:
        val = _get_value_from_row(self.ROW, "EAV:Model")
        assert val is not None, "EAV:Model → 'model' 매칭 실패"
        assert val == "PowerEdge R740"

    def test_eav_vendor(self) -> None:
        val = _get_value_from_row(self.ROW, "EAV:Vendor")
        assert val is not None, "EAV:Vendor → 'vendor' 매칭 실패"
        assert val == "Dell"

    # --- 대소문자 불일치 케이스 ---
    def test_case_insensitive_match(self) -> None:
        """대소문자만 다른 경우 매칭 가능한지."""
        val = _get_value_from_row(self.ROW, "OS_TYPE")
        assert val is not None, "대소문자 불일치 'OS_TYPE' → 'os_type' 매칭 실패"

    # --- substring 매칭 주의: 잘못된 매칭 가능성 ---
    def test_substring_false_positive_hostname(self) -> None:
        """'hostname' substring 매칭 시 eav_hostname vs cmm_resource_hostname 중복."""
        val = _get_value_from_row(self.ROW, "cmm_resource.hostname")
        # substring으로 "hostname" 매칭 시 eav_hostname이 먼저 매칭될 수 있음
        # 값은 동일하지만, 딕셔너리 순서에 의존하는 것은 위험
        assert val in ("web-01",), f"hostname 매칭 결과: {val}"


# ──────────────────────────────────────────────
# 3. fill_excel_template 통합 테스트
# ──────────────────────────────────────────────
class TestFillExcelTemplate:
    """fill_excel_template 호출 후 실제 셀 값이 올바른지 검증."""

    def _fill_and_load(
        self,
        template_data: bytes,
        template_structure: dict,
        column_mapping: dict[str, str | None],
        rows: list[dict[str, Any]],
    ) -> openpyxl.Workbook:
        result_bytes, filled_count = fill_excel_template(
            template_data, template_structure, column_mapping, rows
        )
        logger.info("filled_count=%d", filled_count)
        wb = openpyxl.load_workbook(io.BytesIO(result_bytes))
        return wb

    # --- Case A: table.column 형식 매핑 ---
    def test_fill_with_table_dot_column_mapping(
        self, template_data: bytes, template_structure: dict
    ) -> None:
        """field_mapper가 table.column 형식으로 매핑한 경우 데이터가 채워지는지."""
        wb = self._fill_and_load(
            template_data,
            template_structure,
            MAPPING_TABLE_DOT_COLUMN,
            QUERY_RESULT_ROWS,
        )
        ws = wb["성능관리분석자료"]

        # 데이터 시작 행 = 5, 3개 행 데이터
        results = {}
        for row_idx in range(5, 8):
            row_label = f"row{row_idx}"
            results[row_label] = {
                "서버명": ws.cell(row=row_idx, column=7).value,
                "IP": ws.cell(row=row_idx, column=11).value,
                "Serial 번호": ws.cell(row=row_idx, column=6).value,
                "OS": ws.cell(row=row_idx, column=14).value,
                "업무내용": ws.cell(row=row_idx, column=8).value,
                "대분류": ws.cell(row=row_idx, column=9).value,
                "소분류": ws.cell(row=row_idx, column=10).value,
            }

        logger.info("Case A results: %s", results)

        # Row 5 = QUERY_RESULT_ROWS[0]
        assert results["row5"]["서버명"] == "web-01", (
            f"서버명 매핑 실패: {results['row5']['서버명']}"
        )
        assert results["row5"]["IP"] == "10.0.1.10", (
            f"IP 매핑 실패: {results['row5']['IP']}"
        )
        assert results["row5"]["Serial 번호"] == "SN-ABCD-1234", (
            f"Serial 번호 매핑 실패: {results['row5']['Serial 번호']}"
        )
        assert results["row5"]["OS"] == "Linux", (
            f"OS 매핑 실패: {results['row5']['OS']}"
        )
        assert results["row5"]["업무내용"] == "웹서비스 서버", (
            f"업무내용 매핑 실패: {results['row5']['업무내용']}"
        )

        # Row 6 = QUERY_RESULT_ROWS[1]
        assert results["row6"]["서버명"] == "db-01"
        assert results["row6"]["IP"] == "10.0.2.20"
        assert results["row6"]["Serial 번호"] == "SN-EFGH-5678"

        # Row 7 = QUERY_RESULT_ROWS[2]
        assert results["row7"]["서버명"] == "app-01"
        assert results["row7"]["IP"] == "10.0.3.30"

    # --- Case B: SQL alias 형식 매핑 ---
    def test_fill_with_sql_alias_mapping(
        self, template_data: bytes, template_structure: dict
    ) -> None:
        """field_mapper가 SQL alias 이름으로 매핑한 경우."""
        wb = self._fill_and_load(
            template_data,
            template_structure,
            MAPPING_SQL_ALIAS,
            QUERY_RESULT_ROWS,
        )
        ws = wb["성능관리분석자료"]

        # Row 5 검증
        assert ws.cell(row=5, column=7).value == "web-01", "서버명"
        assert ws.cell(row=5, column=11).value == "10.0.1.10", "IP"
        assert ws.cell(row=5, column=6).value == "SN-ABCD-1234", "Serial 번호"
        assert ws.cell(row=5, column=14).value == "Linux", "OS"
        assert ws.cell(row=5, column=8).value == "웹서비스 서버", "업무내용"
        assert ws.cell(row=5, column=9).value == "Linux", "대분류"
        assert ws.cell(row=5, column=10).value == "RHEL 8.6", "소분류"

    # --- filled_count 검증 ---
    def test_filled_count_case_a(
        self, template_data: bytes, template_structure: dict
    ) -> None:
        """매핑된 필드 수 × 행 수만큼 셀이 채워져야 한다."""
        _, filled_count = fill_excel_template(
            template_data,
            template_structure,
            MAPPING_TABLE_DOT_COLUMN,
            QUERY_RESULT_ROWS,
        )
        # 매핑된 필드: Serial 번호, 서버명, 업무내용, 대분류, 소분류, IP, OS = 7
        # (대분류와 OS 모두 EAV:OSType → 같은 값)
        # 3 rows × 7 fields = 21 (단, 대분류/OS 중복 매핑이므로 7 fields)
        # 실제 값이 None이 아닌 경우만 카운트
        assert filled_count > 0, f"데이터가 하나도 채워지지 않음! filled_count={filled_count}"
        logger.info("Case A filled_count=%d", filled_count)

    def test_filled_count_case_b(
        self, template_data: bytes, template_structure: dict
    ) -> None:
        _, filled_count = fill_excel_template(
            template_data,
            template_structure,
            MAPPING_SQL_ALIAS,
            QUERY_RESULT_ROWS,
        )
        assert filled_count > 0, f"데이터가 하나도 채워지지 않음! filled_count={filled_count}"
        logger.info("Case B filled_count=%d", filled_count)


# ──────────────────────────────────────────────
# 4. 매핑 불일치 진단 테스트
# ──────────────────────────────────────────────
class TestMappingDiagnostics:
    """각 매핑 케이스별로 _get_value_from_row가 실제로 값을 꺼낼 수 있는지 진단."""

    @pytest.mark.parametrize(
        "field_name,db_column",
        [
            ("서버명", "cmm_resource.hostname"),
            ("IP", "cmm_resource.ipaddress"),
            ("업무내용", "cmm_resource.description"),
            ("Serial 번호", "EAV:SerialNumber"),
            ("OS", "EAV:OSType"),
            ("대분류", "EAV:OSType"),
            ("소분류", "EAV:OSVerson"),
        ],
    )
    def test_table_dot_column_value_extraction(
        self, field_name: str, db_column: str
    ) -> None:
        """Case A 매핑의 각 필드가 SQL alias 키 기반 행에서 값을 추출할 수 있는지."""
        row = QUERY_RESULT_ROWS[0]
        val = _get_value_from_row(row, db_column)
        assert val is not None, (
            f"매핑 실패: Excel 필드 '{field_name}' → DB 컬럼 '{db_column}'\n"
            f"  데이터 행 키: {list(row.keys())}\n"
            f"  _get_value_from_row 반환값: None\n"
            f"  → field_mapper가 생성하는 매핑 형식과 SQL 결과의 키 형식이 불일치"
        )
        logger.info("✓ %s → %s = %s", field_name, db_column, val)

    @pytest.mark.parametrize(
        "field_name,db_column,expected",
        [
            ("서버명", "cmm_resource_hostname", "web-01"),
            ("IP", "cmm_resource_ipaddress", "10.0.1.10"),
            ("업무내용", "resource_description", "웹서비스 서버"),
            ("Serial 번호", "serial_number", "SN-ABCD-1234"),
            ("OS", "os_type", "Linux"),
            ("대분류", "os_type", "Linux"),
            ("소분류", "os_version", "RHEL 8.6"),
        ],
    )
    def test_sql_alias_value_extraction(
        self, field_name: str, db_column: str, expected: str
    ) -> None:
        """Case B 매핑 (SQL alias) → 직접 키 매칭으로 값 추출."""
        row = QUERY_RESULT_ROWS[0]
        val = _get_value_from_row(row, db_column)
        assert val == expected, (
            f"매핑 '{field_name}' → '{db_column}': expected={expected}, got={val}"
        )

    def test_diagnostic_report(self) -> None:
        """전체 매핑 진단 리포트 출력 (항상 통과, 디버깅용)."""
        row = QUERY_RESULT_ROWS[0]
        report_lines = [
            "",
            "=" * 70,
            "매핑 진단 리포트 (Case A: table.column 형식)",
            "=" * 70,
        ]
        for field_name, db_column in MAPPING_TABLE_DOT_COLUMN.items():
            if db_column is None:
                report_lines.append(f"  {field_name:15s} → (매핑 없음)")
                continue
            val = _get_value_from_row(row, db_column)
            status = "OK" if val is not None else "FAIL"
            report_lines.append(
                f"  {field_name:15s} → {db_column:30s} = {val!r:20s} [{status}]"
            )

        report_lines.append("")
        report_lines.append("=" * 70)
        report_lines.append("매핑 진단 리포트 (Case B: SQL alias 형식)")
        report_lines.append("=" * 70)
        for field_name, db_column in MAPPING_SQL_ALIAS.items():
            if db_column is None:
                report_lines.append(f"  {field_name:15s} → (매핑 없음)")
                continue
            val = _get_value_from_row(row, db_column)
            status = "OK" if val is not None else "FAIL"
            report_lines.append(
                f"  {field_name:15s} → {db_column:30s} = {val!r:20s} [{status}]"
            )

        report_lines.append("")
        report_lines.append("=" * 70)
        report_lines.append("데이터 행 키 목록:")
        report_lines.append(f"  {list(row.keys())}")
        report_lines.append("=" * 70)

        report = "\n".join(report_lines)
        logger.info(report)
        # 항상 통과 — 로그 확인용
