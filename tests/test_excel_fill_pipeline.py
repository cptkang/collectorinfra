"""Excel 데이터 채우기 파이프라인 통합 테스트.

Plan 35: Excel 데이터 미채움 버그 수정 검증.
- Redis CSV 캐시 (히트/미스/fallback)
- CSV 헤더 vs column_mapping 정합성 검증
- fill_excel_template 반환 타입 (bytes, int)
- _get_value_from_row 다양한 키 형식 매칭
- Silent Failure 제거 확인
"""

from __future__ import annotations

import io
import logging
from dataclasses import asdict
from typing import Any
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import Workbook

from src.document.excel_csv_converter import (
    CsvSheetData,
    _compute_file_hash,
    excel_to_csv,
    excel_to_csv_cached,
)
from src.document.excel_writer import (
    _get_value_from_row,
    fill_excel_template,
)


# ---------------------------------------------------------------------------
# Helper: 간단한 Excel 파일 바이너리 생성
# ---------------------------------------------------------------------------

def _make_test_excel(
    headers: list[str],
    data_rows: list[list[Any]] | None = None,
    sheet_name: str = "Sheet1",
) -> bytes:
    """테스트용 Excel 파일 바이너리를 생성한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 헤더 쓰기
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # 데이터 쓰기
    if data_rows:
        for row_offset, row_data in enumerate(data_rows):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=2 + row_offset, column=col_idx, value=value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _make_template_structure(
    headers: list[str],
    sheet_name: str = "Sheet1",
) -> dict:
    """fill_excel_template에 전달할 template_structure를 생성한다."""
    header_cells = [
        {"col": i + 1, "value": h} for i, h in enumerate(headers)
    ]
    return {
        "sheets": [
            {
                "name": sheet_name,
                "header_cells": header_cells,
                "data_start_row": 2,
                "formula_cells": [],
            }
        ]
    }


# ===========================================================================
# Phase 1: CSV 캐시 테스트
# ===========================================================================


class TestCsvRedisCache:
    """Redis CSV 캐시 관련 테스트."""

    @pytest.mark.asyncio
    async def test_csv_cache_hit(self):
        """동일 파일 업로드 시 Redis CSV 캐시 히트 확인."""
        headers = ["서버명", "IP", "CPU"]
        file_bytes = _make_test_excel(headers, [["svr1", "10.0.0.1", "80"]])

        # 첫 호출: 캐시 미스 -> 변환 수행 -> 저장
        mock_redis_cache = AsyncMock()
        mock_redis_cache.load_csv_cache = AsyncMock(return_value=None)
        mock_redis_cache.save_csv_cache = AsyncMock()

        mock_cache_mgr = MagicMock()
        mock_cache_mgr.redis_available = True
        mock_cache_mgr._redis_cache = mock_redis_cache

        result1 = await excel_to_csv_cached(
            file_bytes, cache_manager=mock_cache_mgr
        )
        assert len(result1) > 0
        mock_redis_cache.save_csv_cache.assert_called_once()

        # 두 번째 호출: 캐시 히트
        file_hash = _compute_file_hash(file_bytes)
        cached_data = {k: asdict(v) for k, v in result1.items()}
        mock_redis_cache.load_csv_cache = AsyncMock(return_value=cached_data)

        result2 = await excel_to_csv_cached(
            file_bytes, cache_manager=mock_cache_mgr
        )
        assert len(result2) > 0
        # 캐시에서 로드했으므로 save 호출은 1회만 (첫 호출)
        assert mock_redis_cache.save_csv_cache.call_count == 1

    @pytest.mark.asyncio
    async def test_csv_cache_miss_different_file(self):
        """다른 파일 업로드 시 Redis 캐시 미스 확인."""
        file1 = _make_test_excel(["A", "B"])
        file2 = _make_test_excel(["X", "Y", "Z"])

        mock_redis_cache = AsyncMock()
        mock_redis_cache.load_csv_cache = AsyncMock(return_value=None)
        mock_redis_cache.save_csv_cache = AsyncMock()

        mock_cache_mgr = MagicMock()
        mock_cache_mgr.redis_available = True
        mock_cache_mgr._redis_cache = mock_redis_cache

        await excel_to_csv_cached(file1, cache_manager=mock_cache_mgr)
        await excel_to_csv_cached(file2, cache_manager=mock_cache_mgr)

        # 두 파일 모두 캐시 미스이므로 save가 2회 호출됨
        assert mock_redis_cache.save_csv_cache.call_count == 2

    @pytest.mark.asyncio
    async def test_csv_cache_fallback_redis_unavailable(self):
        """Redis 미사용 시 fallback으로 직접 변환 수행 확인."""
        headers = ["서버명", "IP"]
        file_bytes = _make_test_excel(headers, [["svr1", "10.0.0.1"]])

        # cache_manager=None -> Redis 없이 직접 변환
        result = await excel_to_csv_cached(file_bytes, cache_manager=None)
        assert len(result) > 0
        for sheet_data in result.values():
            assert isinstance(sheet_data, CsvSheetData)
            assert "서버명" in sheet_data.headers
            assert "IP" in sheet_data.headers

    @pytest.mark.asyncio
    async def test_csv_cache_fallback_redis_error(self):
        """Redis 장애 시 graceful fallback 확인."""
        headers = ["서버명"]
        file_bytes = _make_test_excel(headers)

        mock_redis_cache = AsyncMock()
        mock_redis_cache.load_csv_cache = AsyncMock(side_effect=Exception("Redis down"))
        mock_redis_cache.save_csv_cache = AsyncMock(side_effect=Exception("Redis down"))

        mock_cache_mgr = MagicMock()
        mock_cache_mgr.redis_available = True
        mock_cache_mgr._redis_cache = mock_redis_cache

        # Redis 장애에도 변환은 성공해야 함
        result = await excel_to_csv_cached(
            file_bytes, cache_manager=mock_cache_mgr
        )
        assert len(result) > 0


# ===========================================================================
# Phase 1: 매핑 검증 테스트
# ===========================================================================


class TestValidateMappingAgainstCsv:
    """CSV 헤더 vs column_mapping 정합성 검증 테스트."""

    def test_validate_mapping_all_mapped(self, caplog):
        """모든 헤더가 매핑된 경우."""
        from src.nodes.output_generator import _validate_mapping_against_csv

        csv_data = {
            "Sheet1": {"headers": ["서버명", "IP"], "example_rows": []}
        }
        mapping = {"서버명": "cmm_resource.hostname", "IP": "cmm_resource.ip_address"}

        with caplog.at_level(logging.INFO):
            _validate_mapping_against_csv(csv_data, mapping)

        assert "100%" in caplog.text

    def test_validate_mapping_partial(self, caplog):
        """일부만 매핑된 경우."""
        from src.nodes.output_generator import _validate_mapping_against_csv

        csv_data = {
            "Sheet1": {"headers": ["서버명", "IP", "CPU"], "example_rows": []}
        }
        mapping = {"서버명": "cmm_resource.hostname", "IP": None, "CPU": None}

        with caplog.at_level(logging.INFO):
            _validate_mapping_against_csv(csv_data, mapping)

        assert "33%" in caplog.text

    def test_validate_mapping_zero_ratio(self, caplog):
        """매핑률 0% 경고."""
        from src.nodes.output_generator import _validate_mapping_against_csv

        csv_data = {
            "Sheet1": {"headers": ["서버명", "IP"], "example_rows": []}
        }
        mapping = {"서버명": None, "IP": None}

        with caplog.at_level(logging.WARNING):
            _validate_mapping_against_csv(csv_data, mapping)

        assert "0%" in caplog.text

    def test_validate_mapping_key_mismatch(self, caplog):
        """column_mapping 키가 CSV 헤더와 불일치."""
        from src.nodes.output_generator import _validate_mapping_against_csv

        csv_data = {
            "Sheet1": {"headers": ["서버명", "IP"], "example_rows": []}
        }
        mapping = {"server_name": "cmm_resource.hostname", "ip_addr": "cmm_resource.ip"}

        with caplog.at_level(logging.WARNING):
            _validate_mapping_against_csv(csv_data, mapping)

        assert "불일치" in caplog.text


# ===========================================================================
# Phase 2: Silent Failure 제거 테스트
# ===========================================================================


class TestFillExcelTemplate:
    """fill_excel_template 반환 타입 및 Silent Failure 제거 테스트."""

    def test_fill_returns_tuple(self):
        """fill_excel_template가 (bytes, int) 튜플을 반환하는지 확인."""
        headers = ["서버명", "IP"]
        file_bytes = _make_test_excel(headers)
        template = _make_template_structure(headers)
        mapping = {"서버명": "hostname", "IP": "ip_address"}
        rows = [{"hostname": "svr1", "ip_address": "10.0.0.1"}]

        result = fill_excel_template(
            file_data=file_bytes,
            template_structure=template,
            column_mapping=mapping,
            rows=rows,
        )
        assert isinstance(result, tuple)
        assert len(result) == 2
        file_out, total_filled = result
        assert isinstance(file_out, bytes)
        assert total_filled > 0

    def test_fill_empty_mapping_returns_zero(self):
        """column_mapping이 모두 None일 때 total_filled=0 반환 확인."""
        headers = ["서버명", "IP"]
        file_bytes = _make_test_excel(headers)
        template = _make_template_structure(headers)
        mapping = {"서버명": None, "IP": None}
        rows = [{"hostname": "svr1", "ip_address": "10.0.0.1"}]

        file_out, total_filled = fill_excel_template(
            file_data=file_bytes,
            template_structure=template,
            column_mapping=mapping,
            rows=rows,
        )
        assert isinstance(file_out, bytes)
        assert total_filled == 0

    def test_fill_key_mismatch_case_insensitive(self):
        """DB 결과 키와 매핑 값 불일치 시 case-insensitive 매칭 확인."""
        headers = ["서버명"]
        file_bytes = _make_test_excel(headers)
        template = _make_template_structure(headers)
        mapping = {"서버명": "cmm_resource.hostname"}
        # DB 결과의 키가 대문자
        rows = [{"HOSTNAME": "svr1"}]

        file_out, total_filled = fill_excel_template(
            file_data=file_bytes,
            template_structure=template,
            column_mapping=mapping,
            rows=rows,
        )
        assert total_filled > 0

    def test_fill_multiple_rows(self):
        """여러 행 데이터 채우기."""
        headers = ["서버명", "IP"]
        file_bytes = _make_test_excel(headers)
        template = _make_template_structure(headers)
        mapping = {"서버명": "hostname", "IP": "ip"}
        rows = [
            {"hostname": "svr1", "ip": "10.0.0.1"},
            {"hostname": "svr2", "ip": "10.0.0.2"},
            {"hostname": "svr3", "ip": "10.0.0.3"},
        ]

        file_out, total_filled = fill_excel_template(
            file_data=file_bytes,
            template_structure=template,
            column_mapping=mapping,
            rows=rows,
        )
        # 3행 x 2컬럼 = 6셀
        assert total_filled == 6

    def test_fill_data_verified_in_output(self):
        """채워진 데이터가 출력 Excel에 실제로 존재하는지 확인."""
        from openpyxl import load_workbook

        headers = ["서버명", "IP"]
        file_bytes = _make_test_excel(headers)
        template = _make_template_structure(headers)
        mapping = {"서버명": "hostname", "IP": "ip"}
        rows = [{"hostname": "test-server", "ip": "192.168.1.1"}]

        file_out, total_filled = fill_excel_template(
            file_data=file_bytes,
            template_structure=template,
            column_mapping=mapping,
            rows=rows,
        )
        assert total_filled > 0

        # 출력 Excel 검증
        wb = load_workbook(io.BytesIO(file_out))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "test-server"
        assert ws.cell(row=2, column=2).value == "192.168.1.1"
        wb.close()


# ===========================================================================
# Phase 4: _get_value_from_row 매칭 강화 테스트
# ===========================================================================


class TestGetValueFromRow:
    """다양한 키 형식에서 값 추출 확인."""

    def test_exact_match(self):
        """정확한 키 매칭."""
        row = {"cmm_resource.hostname": "svr1"}
        assert _get_value_from_row(row, "cmm_resource.hostname") == "svr1"

    def test_column_only_match(self):
        """table.column에서 column 부분만 매칭."""
        row = {"hostname": "svr1"}
        assert _get_value_from_row(row, "cmm_resource.hostname") == "svr1"

    def test_eav_attr_match(self):
        """EAV:attr 형식 매칭."""
        row = {"OSType": "Linux"}
        assert _get_value_from_row(row, "EAV:OSType") == "Linux"

    def test_eav_attr_case_insensitive(self):
        """EAV:attr 대소문자 무시 매칭."""
        row = {"ostype": "Linux"}
        assert _get_value_from_row(row, "EAV:OSType") == "Linux"

    def test_case_insensitive_match(self):
        """대소문자 무시 매칭."""
        row = {"HOSTNAME": "svr1"}
        assert _get_value_from_row(row, "cmm_resource.hostname") == "svr1"

    def test_reverse_mapping_match(self):
        """역방향 매핑 (한글 필드명) 매칭."""
        row = {"서버명": "svr1"}
        reverse = {"cmm_resource.hostname": "서버명"}
        assert (
            _get_value_from_row(row, "cmm_resource.hostname", reverse) == "svr1"
        )

    def test_partial_match_substring(self):
        """부분 매칭 (substring) 폴백."""
        row = {"server_hostname": "svr1"}
        assert _get_value_from_row(row, "hostname") == "svr1"

    def test_partial_match_reverse(self):
        """부분 매칭: 키가 매핑값의 일부."""
        row = {"host": "svr1"}
        # "host" is contained in "hostname"
        assert _get_value_from_row(row, "cmm_resource.hostname") == "svr1"

    def test_no_match_returns_none(self):
        """매칭되지 않으면 None 반환."""
        row = {"unrelated_column": "value"}
        assert _get_value_from_row(row, "cmm_resource.hostname") is None

    def test_none_value_preserved(self):
        """키가 매칭되지만 값이 None인 경우."""
        row = {"hostname": None}
        # exact column match 성공하지만 값이 None
        assert _get_value_from_row(row, "cmm_resource.hostname") is None


# ===========================================================================
# 해시 함수 테스트
# ===========================================================================


class TestComputeFileHash:
    """파일 해시 계산 테스트."""

    def test_same_content_same_hash(self):
        """동일 내용은 동일 해시."""
        data = b"test content"
        assert _compute_file_hash(data) == _compute_file_hash(data)

    def test_different_content_different_hash(self):
        """다른 내용은 다른 해시."""
        assert _compute_file_hash(b"a") != _compute_file_hash(b"b")

    def test_hash_is_hex_string(self):
        """해시 결과가 16진수 문자열."""
        h = _compute_file_hash(b"test")
        assert len(h) == 64  # SHA-256 hex length
        assert all(c in "0123456789abcdef" for c in h)
