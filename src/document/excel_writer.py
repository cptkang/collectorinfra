"""Excel 양식 데이터 채우기 모듈.

openpyxl을 사용하여 조회 결과를 Excel 양식에 채워넣고,
원본의 병합 셀, 서식, 수식을 보존한다.
"""

from __future__ import annotations

import copy
import io
import logging
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def fill_excel_template(
    file_data: bytes,
    template_structure: dict[str, Any],
    column_mapping: dict[str, Optional[str]],
    rows: list[dict[str, Any]],
    *,
    sheet_mappings: list[dict[str, Any]] | None = None,
    target_sheets: list[str] | None = None,
) -> bytes:
    """Excel 양식에 조회 결과를 채워넣는다.

    멀티시트 지원:
    - sheet_mappings가 제공되면 시트별 독립 매핑/데이터를 사용한다.
    - target_sheets가 제공되면 해당 시트만 처리한다.
    - 둘 다 없으면 기존 방식대로 모든 시트에 동일한 매핑/데이터를 적용한다.

    Args:
        file_data: 원본 Excel 파일 바이너리
        template_structure: 양식 구조 정보 (excel_parser 출력)
        column_mapping: 필드-컬럼 매핑 (field_mapper 출력, 단일 시트용)
        rows: 조회 결과 행 목록 (단일 시트용)
        sheet_mappings: 시트별 매핑 결과 목록 (멀티시트용, 선택)
        target_sheets: 대상 시트명 목록 (None이면 전체 시트)

    Returns:
        데이터가 채워진 Excel 파일 바이너리

    Raises:
        ValueError: 파일을 처리할 수 없는 경우
    """
    try:
        wb = load_workbook(io.BytesIO(file_data), data_only=False)
    except Exception as e:
        raise ValueError(f"Excel 파일을 읽을 수 없습니다: {e}") from e

    sheets_info = template_structure.get("sheets", [])
    target_set = set(target_sheets) if target_sheets else None

    # 시트별 매핑 딕셔너리 구축
    per_sheet_mapping: dict[str, dict[str, Optional[str]]] = {}
    per_sheet_rows: dict[str, list[dict[str, Any]]] = {}
    if sheet_mappings:
        for sm in sheet_mappings:
            sname = sm.get("sheet_name", "")
            per_sheet_mapping[sname] = sm.get("column_mapping") or {}
            per_sheet_rows[sname] = sm.get("rows", rows)

    total_filled = 0
    for sheet_info in sheets_info:
        sheet_name = sheet_info.get("name", "")

        # target_sheets 필터링
        if target_set and sheet_name not in target_set:
            logger.debug("시트 '%s': target_sheets에 포함되지 않아 스킵", sheet_name)
            continue

        if sheet_name not in wb.sheetnames:
            logger.warning("시트 '%s'를 찾을 수 없어 스킵", sheet_name)
            continue

        ws = wb[sheet_name]

        # 시트별 매핑이 있으면 해당 매핑 사용, 없으면 공통 매핑 사용
        sheet_col_mapping = per_sheet_mapping.get(sheet_name, column_mapping)
        sheet_rows = per_sheet_rows.get(sheet_name, rows)

        if sheet_col_mapping:
            _fill_sheet(ws, sheet_info, sheet_col_mapping, sheet_rows)
            total_filled += len(sheet_rows)
        else:
            logger.debug("시트 '%s': 매핑 정보가 없어 스킵", sheet_name)

    # 바이너리로 저장
    output = io.BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)

    logger.info("Excel 파일 생성 완료: %d건 데이터 채움 (시트 수: %d)", total_filled, len(sheets_info))
    return output.getvalue()


def _fill_sheet(
    ws: Worksheet,
    sheet_info: dict[str, Any],
    column_mapping: dict[str, Optional[str]],
    rows: list[dict[str, Any]],
) -> None:
    """단일 시트에 데이터를 채운다.

    Args:
        ws: openpyxl Worksheet 객체
        sheet_info: 시트 구조 정보
        column_mapping: 필드-컬럼 매핑
        rows: 조회 결과 행 목록
    """
    header_cells = sheet_info.get("header_cells", [])
    data_start_row = sheet_info.get("data_start_row", 2)
    formula_cells_set = set(sheet_info.get("formula_cells", []))

    # 헤더별 열 인덱스와 매핑된 DB 컬럼 매칭
    col_assignments: list[tuple[int, str]] = []
    for hc in header_cells:
        col_idx = hc["col"]
        header_name = hc["value"]
        mapped = column_mapping.get(header_name)
        if mapped:
            col_assignments.append((col_idx, mapped))

    if not col_assignments:
        logger.warning("시트 '%s': 매핑된 컬럼이 없어 데이터 채우기 스킵", ws.title)
        return

    # 매칭 실패한 헤더 경고 로그
    mapped_headers = {hc["value"] for hc in header_cells if column_mapping.get(hc["value"])}
    all_headers = {hc["value"] for hc in header_cells}
    unmapped = all_headers - mapped_headers
    if unmapped:
        logger.info("시트 '%s': 매핑 안 된 헤더=%s", ws.title, unmapped)

    # 서식 참조용: 데이터 시작 행의 기존 셀 스타일 수집
    style_cache = _collect_row_styles(ws, data_start_row, [ca[0] for ca in col_assignments])

    # 데이터 채우기
    for row_offset, data_row in enumerate(rows):
        target_row = data_start_row + row_offset

        for col_idx, db_column in col_assignments:
            cell_coord = ws.cell(row=target_row, column=col_idx).coordinate
            if cell_coord in formula_cells_set:
                continue  # 수식 셀은 건너뜀

            # DB 컬럼에서 값 추출 (table.column -> column)
            value = _get_value_from_row(data_row, db_column)

            cell = ws.cell(row=target_row, column=col_idx)
            # None 값 처리: 매핑된 컬럼에 값이 없으면 원본 셀 값 유지
            if value is None:
                continue
            cell.value = value

            # 서식 적용
            if col_idx in style_cache:
                _apply_style(cell, style_cache[col_idx])


def _collect_row_styles(
    ws: Worksheet,
    row_idx: int,
    col_indices: list[int],
) -> dict[int, dict[str, Any]]:
    """특정 행의 셀 스타일을 수집한다.

    Args:
        ws: Worksheet 객체
        row_idx: 행 번호
        col_indices: 스타일을 수집할 열 인덱스 목록

    Returns:
        {열 인덱스: 스타일 정보} 딕셔너리
    """
    styles: dict[int, dict[str, Any]] = {}

    for col_idx in col_indices:
        cell: Cell = ws.cell(row=row_idx, column=col_idx)
        styles[col_idx] = {
            "font": copy.copy(cell.font) if cell.font else None,
            "fill": copy.copy(cell.fill) if cell.fill else None,
            "border": copy.copy(cell.border) if cell.border else None,
            "alignment": copy.copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format,
        }

    return styles


def _apply_style(cell: Cell, style: dict[str, Any]) -> None:
    """셀에 스타일을 적용한다.

    Args:
        cell: 대상 셀
        style: 스타일 정보 딕셔너리
    """
    if style.get("font"):
        cell.font = style["font"]
    if style.get("fill"):
        cell.fill = style["fill"]
    if style.get("border"):
        cell.border = style["border"]
    if style.get("alignment"):
        cell.alignment = style["alignment"]
    if style.get("number_format"):
        cell.number_format = style["number_format"]


def _get_value_from_row(
    data_row: dict[str, Any],
    db_column: str,
) -> Any:
    """데이터 행에서 DB 컬럼에 해당하는 값을 추출한다.

    "table.column" 형식의 키와 "column" 형식의 키를 모두 시도한다.

    Args:
        data_row: 조회 결과 행
        db_column: "table.column" 형식의 DB 컬럼명

    Returns:
        추출된 값 또는 None
    """
    # 1. "table.column" 형식으로 시도
    if db_column in data_row:
        return data_row[db_column]

    # 2. "column" 부분만으로 시도
    if "." in db_column:
        col_name = db_column.split(".", 1)[1]
        if col_name in data_row:
            return data_row[col_name]

    # 3. 대소문자 무시 검색
    lower_col = db_column.lower()
    for key, value in data_row.items():
        if key.lower() == lower_col or (
            "." in db_column and key.lower() == db_column.split(".", 1)[1].lower()
        ):
            return value

    return None
