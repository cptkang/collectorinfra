"""Excel 양식 구조 분석 모듈.

openpyxl을 사용하여 Excel 파일의 헤더, 데이터 영역, 병합 셀, 수식 셀 정보를 추출한다.
"""

from __future__ import annotations

import io
import logging
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# 헤더 행으로 판단하기 위한 최소 비어있지 않은 셀 수
_MIN_HEADER_CELLS = 2

# 헤더를 탐색할 최대 행 수
_MAX_HEADER_SEARCH_ROWS = 20

# 연속 빈 행이 이 수를 초과하면 데이터 영역 끝으로 판단
_MAX_CONSECUTIVE_EMPTY_ROWS = 3

# 2행 병합 헤더 결합 시 그룹 헤더와 서브 헤더를 잇는 구분자 (D-145)
# normalize_field_name이 보존하는 문자여야 한다
_HEADER_JOIN_SEPARATOR = "|"


def parse_excel_template(file_data: bytes) -> dict[str, Any]:
    """Excel 양식 파일의 구조를 분석한다.

    Args:
        file_data: Excel 파일 바이너리 데이터

    Returns:
        양식 구조 딕셔너리 (template_structure 형식)

    Raises:
        ValueError: 파일을 읽을 수 없는 경우
    """
    try:
        wb = load_workbook(io.BytesIO(file_data), data_only=False)
    except Exception as e:
        raise ValueError(f"Excel 파일을 읽을 수 없습니다: {e}") from e

    sheets: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        sheet_info = _analyze_sheet(ws)
        if sheet_info is not None:
            sheets.append(sheet_info)

    wb.close()

    logger.info("Excel 양식 분석 완료: %d개 시트", len(sheets))

    return {
        "file_type": "xlsx",
        "sheets": sheets,
        "placeholders": [],
        "tables": [],
    }


def _analyze_sheet(ws: Worksheet) -> dict[str, Any] | None:
    """단일 시트의 구조를 분석한다.

    Args:
        ws: openpyxl Worksheet 객체

    Returns:
        시트 구조 딕셔너리 또는 None (빈 시트)
    """
    block_top, header_row, header_cells = _detect_header_block(ws)
    if header_row is None:
        logger.debug("시트 '%s': 헤더를 찾을 수 없어 스킵", ws.title)
        return None

    headers = [cell["value"] for cell in header_cells]
    data_start_row = header_row + 1
    max_column = max(cell["col"] for cell in header_cells) if header_cells else 1

    # 데이터 영역 끝 탐지
    data_end_row = _detect_data_end_row(ws, data_start_row, max_column)

    # 병합 셀 정보
    merged_cells = [str(rng) for rng in ws.merged_cells.ranges]

    # 수식 셀 정보
    formula_cells = _detect_formula_cells(ws, data_start_row, data_end_row, max_column)

    # 헤더 블록 위 제목·유의사항 텍스트 — 양식 문맥 판정(월 시리즈 인식기의 리소스 판정 등)에
    # 쓰인다. 시트 이름(Sheet1)만으로는 양식 종류를 알 수 없는 실물 양식이 많다(D-146).
    title_text = _collect_title_text(ws, block_top)

    return {
        "name": ws.title,
        "headers": headers,
        "title_text": title_text,
        # header_row는 헤더 블록의 "하단 행" — 모든 소비처가 data_start_row =
        # header_row + 1 계약에 의존하므로 2행 블록에서도 이 계약을 유지한다(D-145).
        "header_row": header_row,
        "data_start_row": data_start_row,
        "data_end_row": data_end_row,
        "header_cells": header_cells,
        "header_block_rows": list(range(block_top, header_row + 1)),
        "merged_cells": merged_cells,
        "formula_cells": formula_cells,
        "max_column": max_column,
    }


def _collect_title_text(ws: Worksheet, block_top: int) -> str:
    """헤더 블록 위(1행~블록 상단-1행)의 비어있지 않은 셀 값을 공백으로 이어 반환한다."""
    parts: list[str] = []
    for row_idx in range(1, min(block_top, _MAX_HEADER_SEARCH_ROWS + 1)):
        for col_idx in range(1, (ws.max_column or 1) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None and str(value).strip():
                parts.append(_normalize_header_value(str(value)))
    return " ".join(parts)


def _normalize_header_value(value: str) -> str:
    """헤더 셀 값을 정규화한다.

    줄바꿈, 다중 공백, Unicode 정규화를 적용한다.
    """
    from src.utils.schema_utils import normalize_field_name

    return normalize_field_name(value)


def _detect_header_row(
    ws: Worksheet,
) -> tuple[int | None, list[dict[str, Any]]]:
    """헤더 행을 자동 탐지한다 (헤더 블록 탐지의 호환 래퍼).

    2행 병합 헤더 블록이 결합된 경우에도 (블록 하단 행, 결합된 셀 목록)을 반환하므로
    기존 소비처의 `data_start_row = header_row + 1` 계약이 그대로 유지된다(D-145).
    excel_csv_converter 등 폼필 외 경로도 이 래퍼를 통해 동일하게 해석한다(경로 대칭).

    Returns:
        (헤더 행 번호(1-based, 블록이면 하단 행), 헤더 셀 목록) 또는 (None, [])
    """
    _block_top, bottom_row, cells = _detect_header_block(ws)
    return bottom_row, cells


def _detect_header_block(
    ws: Worksheet,
) -> tuple[int | None, int | None, list[dict[str, Any]]]:
    """헤더 블록(1행 또는 2행 병합 헤더)을 자동 탐지한다.

    1) 탐색 범위 내에서 비어있지 않은 셀이 가장 많은 행을 기준 헤더로 판단한다.
    2) 기준 행의 인접 행이 병합 구조로 결합된 2행 헤더(그룹 헤더 + 서브 헤더)이면
       두 행을 열 단위로 결합해 복합 필드명("그룹|서브")을 생성한다(D-145).
       결합 조건 미충족 시 기존 단일 행 동작과 완전히 동일하다.

    Returns:
        (블록 상단 행, 블록 하단 행, 헤더 셀 목록) 또는 (None, None, [])
    """
    best_row: int | None = None
    best_cells: list[dict[str, Any]] = []

    for row_idx in range(1, min(ws.max_row or 1, _MAX_HEADER_SEARCH_ROWS) + 1):
        cells: list[dict[str, Any]] = []
        for col_idx in range(1, (ws.max_column or 1) + 1):
            cell: Cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is not None and str(value).strip():
                cells.append({
                    "col": col_idx,
                    "value": _normalize_header_value(str(value)),
                })

        if len(cells) >= _MIN_HEADER_CELLS and len(cells) > len(best_cells):
            best_row = row_idx
            best_cells = cells

    if best_row is None:
        return None, None, []

    # 2행 병합 헤더 블록 결합 시도: (위 행 + 기준 행) 우선, 다음 (기준 행 + 아래 행)
    for top, bottom in ((best_row - 1, best_row), (best_row, best_row + 1)):
        if top < 1 or bottom > (ws.max_row or 1):
            continue
        combined = _try_combine_header_rows(ws, top, bottom)
        if combined is not None:
            return top, bottom, combined

    return best_row, best_row, best_cells


def _row_raw_values(ws: Worksheet, row_idx: int) -> dict[int, Any]:
    """행의 비어있지 않은 원시 값을 {열: 값}으로 수집한다."""
    values: dict[int, Any] = {}
    for col_idx in range(1, (ws.max_column or 1) + 1):
        value = ws.cell(row=row_idx, column=col_idx).value
        if value is not None and str(value).strip():
            values[col_idx] = value
    return values


def _is_numeric_value(value: Any) -> bool:
    """셀 값이 수치(데이터 행 신호)인지 판단한다."""
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    try:
        float(str(value).strip().replace(",", ""))
        return True
    except ValueError:
        return False


def _numeric_ratio(values: list[Any]) -> float:
    """값 목록 중 수치 값의 비율을 계산한다."""
    if not values:
        return 0.0
    return sum(1 for v in values if _is_numeric_value(v)) / len(values)


def _effective_header_value(
    ws: Worksheet,
    merged_ranges: list[Any],
    row_idx: int,
    col_idx: int,
) -> tuple[Any | None, Any]:
    """병합 셀을 고려한 유효 값을 조회한다.

    (row, col)이 병합 범위 안이면 앵커(최상좌) 값을, 아니면 셀 자신의 값을 반환한다.

    Returns:
        (소속 병합 범위 또는 None, 유효 값)
    """
    for rng in merged_ranges:
        if rng.min_row <= row_idx <= rng.max_row and rng.min_col <= col_idx <= rng.max_col:
            return rng, ws.cell(row=rng.min_row, column=rng.min_col).value
    return None, ws.cell(row=row_idx, column=col_idx).value


def _try_combine_header_rows(
    ws: Worksheet,
    top: int,
    bottom: int,
) -> list[dict[str, Any]] | None:
    """인접 2행이 병합 헤더 블록이면 열 단위로 결합한 헤더 셀 목록을 반환한다.

    보수적 3중 게이트(D-145 — 기존 단일 헤더 템플릿 오결합 방지):
    1. 두 행 모두 비어있지 않은 셀이 _MIN_HEADER_CELLS 이상
    2. 병합 증거: 두 행에 정확히 걸친 **세로 병합**이 존재 (필수).
       가로 병합만으로는 결합하지 않는다 — 부분 그룹 행(예: sample/취합 예시2.xlsx
       3행 "자원현황")이 세로 병합 없이 가로 병합만 갖는 실측 오결합 사례가 있어,
       두 행을 구조적으로 묶는 세로 병합을 필수 증거로 요구한다.
    3. 두 행 모두 수치 값 비율 50% 이하 (데이터 행 오인 방지)

    Returns:
        결합된 헤더 셀 목록, 결합 조건 미충족 시 None
    """
    top_raw = _row_raw_values(ws, top)
    bot_raw = _row_raw_values(ws, bottom)

    # 게이트 1: 양쪽 모두 헤더로 볼 만한 셀 수
    if len(top_raw) < _MIN_HEADER_CELLS or len(bot_raw) < _MIN_HEADER_CELLS:
        return None

    # 게이트 3: 데이터 행 오인 방지
    if _numeric_ratio(list(top_raw.values())) > 0.5 or _numeric_ratio(list(bot_raw.values())) > 0.5:
        return None

    merged_ranges = list(ws.merged_cells.ranges)

    # 게이트 2: 세로 병합 증거 (필수)
    has_vertical_span = any(
        rng.min_row == top and rng.max_row == bottom for rng in merged_ranges
    )
    if not has_vertical_span:
        return None

    # 열 단위 결합: 세로 병합(동일 범위)은 단독 값, 그 외 상·하 값 존재 시 "상단|하단"
    cells: list[dict[str, Any]] = []
    for col_idx in range(1, (ws.max_column or 1) + 1):
        top_rng, top_val = _effective_header_value(ws, merged_ranges, top, col_idx)
        bot_rng, bot_val = _effective_header_value(ws, merged_ranges, bottom, col_idx)

        if top_rng is not None and top_rng is bot_rng:
            # 두 행에 걸친 동일 병합(세로 병합): 앵커 값 단독
            parts = [top_val]
        else:
            parts = [top_val, bot_val]

        norm_parts = [
            _normalize_header_value(str(p))
            for p in parts
            if p is not None and str(p).strip()
        ]
        if not norm_parts:
            continue
        cells.append({
            "col": col_idx,
            "value": _HEADER_JOIN_SEPARATOR.join(norm_parts),
        })

    return cells if len(cells) >= _MIN_HEADER_CELLS else None


def _detect_data_end_row(
    ws: Worksheet,
    data_start_row: int,
    max_column: int,
) -> int | None:
    """데이터 영역의 끝 행을 탐지한다.

    연속으로 빈 행이 _MAX_CONSECUTIVE_EMPTY_ROWS개 이상이면 데이터 영역 끝으로 판단한다.
    끝을 찾지 못하면 None (자동 확장 가능)을 반환한다.

    Args:
        ws: openpyxl Worksheet 객체
        data_start_row: 데이터 시작 행 (1-based)
        max_column: 데이터 영역 최대 열

    Returns:
        데이터 끝 행 (1-based) 또는 None
    """
    max_row = ws.max_row or data_start_row
    consecutive_empty = 0
    last_data_row = None

    for row_idx in range(data_start_row, max_row + 1):
        is_empty = True
        for col_idx in range(1, max_column + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None and str(cell_val).strip():
                is_empty = False
                break

        if is_empty:
            consecutive_empty += 1
            if consecutive_empty >= _MAX_CONSECUTIVE_EMPTY_ROWS:
                return last_data_row
        else:
            consecutive_empty = 0
            last_data_row = row_idx

    return last_data_row


def _detect_formula_cells(
    ws: Worksheet,
    data_start_row: int,
    data_end_row: int | None,
    max_column: int,
) -> list[str]:
    """수식이 포함된 셀 위치를 탐지한다.

    Args:
        ws: openpyxl Worksheet 객체
        data_start_row: 데이터 시작 행
        data_end_row: 데이터 끝 행
        max_column: 최대 열

    Returns:
        수식 셀 주소 목록 (예: ["D2", "E3"])
    """
    formula_cells: list[str] = []
    end_row = data_end_row or data_start_row + 100  # 최대 100행 탐색

    for row_idx in range(data_start_row, min(end_row + 1, (ws.max_row or 0) + 1)):
        for col_idx in range(1, max_column + 1):
            cell: Cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append(cell.coordinate)

    return formula_cells
