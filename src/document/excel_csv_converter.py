"""Excel -> CSV 변환 모듈.

Excel 파일의 각 시트에서 헤더와 예시 데이터를 추출하여
CsvSheetData 구조로 변환한다. LLM 컨텍스트 보강 목적으로 사용된다.
CSV 변환이 실패하면 기존 excel_parser의 template_structure 기반 폴백을 수행한다.
"""

from __future__ import annotations

import csv
import hashlib
import io
import logging
from dataclasses import asdict, dataclass
from datetime import date, datetime, time
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.document.excel_parser import (
    _detect_data_end_row,
    _detect_header_row,
    parse_excel_template,
)

logger = logging.getLogger(__name__)

# 예시 데이터 최대 행 수 (LLM 토큰 절약)
_MAX_EXAMPLE_ROWS = 50


class CsvConversionError(Exception):
    """CSV 변환 실패 시 폴백 판단을 위한 예외."""


@dataclass
class CsvSheetData:
    """시트별 헤더와 예시 데이터를 구조화한 데이터클래스."""

    sheet_name: str
    headers: list[str]
    example_rows: list[list[str]]
    csv_text: str
    header_row_index: int
    data_start_row: int


def excel_to_csv(
    file_data: bytes, sheet_name: str | None = None
) -> dict[str, CsvSheetData]:
    """Excel 파일을 시트별 CsvSheetData로 변환한다.

    Args:
        file_data: Excel 파일 바이너리
        sheet_name: 특정 시트만 변환 (None이면 전체)

    Returns:
        {"시트명": CsvSheetData, ...}

    Raises:
        ValueError: 파일을 읽을 수 없는 경우
    """
    try:
        wb = load_workbook(io.BytesIO(file_data), data_only=True)
    except Exception as e:
        raise ValueError(f"Excel 파일을 읽을 수 없습니다: {e}") from e

    result: dict[str, CsvSheetData] = {}

    for ws in wb.worksheets:
        if sheet_name is not None and ws.title != sheet_name:
            continue

        try:
            sheet_data = _extract_csv_sheet_data(ws)
        except CsvConversionError:
            logger.info(
                "시트 '%s': CSV 변환 실패, template_structure 폴백 수행", ws.title
            )
            sheet_data = _extract_from_template_structure(ws, file_data)

        result[ws.title] = sheet_data

    wb.close()

    logger.info("Excel -> CSV 변환 완료: %d개 시트", len(result))
    return result


def _extract_csv_sheet_data(ws: Worksheet) -> CsvSheetData:
    """단일 시트에서 헤더와 예시 데이터를 CSV 형태로 추출한다.

    Args:
        ws: openpyxl Worksheet 객체

    Returns:
        CsvSheetData 인스턴스

    Raises:
        CsvConversionError: 헤더 탐지 실패 시
    """
    header_row, header_cells = _detect_header_row(ws)
    if header_row is None:
        raise CsvConversionError(f"시트 '{ws.title}': 헤더를 탐지할 수 없습니다")

    headers = [cell["value"] for cell in header_cells]
    header_col_indices = [cell["col"] for cell in header_cells]
    max_column = max(header_col_indices) if header_col_indices else 1
    data_start_row = header_row + 1

    # 데이터 영역 끝 탐지
    data_end_row = _detect_data_end_row(ws, data_start_row, max_column)

    # 예시 데이터 행 추출
    example_rows: list[list[str]] = []
    if data_end_row is not None:
        end = min(data_end_row, data_start_row + _MAX_EXAMPLE_ROWS - 1)
        for row_idx in range(data_start_row, end + 1):
            row_values: list[str] = []
            for col_idx in header_col_indices:
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_values.append(_format_cell_value(cell_value))
            example_rows.append(row_values)

    # CSV 텍스트 생성
    csv_text = _build_csv_text(headers, example_rows)

    return CsvSheetData(
        sheet_name=ws.title,
        headers=headers,
        example_rows=example_rows,
        csv_text=csv_text,
        header_row_index=header_row,
        data_start_row=data_start_row,
    )


def _extract_from_template_structure(
    ws: Worksheet, file_data: bytes
) -> CsvSheetData:
    """template_structure 기반 폴백으로 헤더를 추출한다.

    CSV 변환이 실패한 시트에 대해 기존 excel_parser의
    parse_excel_template()을 활용하여 헤더 정보를 추출한다.

    Args:
        ws: openpyxl Worksheet 객체
        file_data: 원본 Excel 파일 바이너리

    Returns:
        CsvSheetData 인스턴스 (example_rows 비어있음)
    """
    template = parse_excel_template(file_data)

    headers: list[str] = []
    header_row_index = 1
    data_start_row = 2

    for sheet_info in template.get("sheets", []):
        if sheet_info.get("name") == ws.title:
            headers = sheet_info.get("headers", [])
            header_row_index = sheet_info.get("header_row", 1)
            data_start_row = sheet_info.get("data_start_row", header_row_index + 1)
            break

    csv_text = _build_csv_text(headers, [])

    return CsvSheetData(
        sheet_name=ws.title,
        headers=headers,
        example_rows=[],
        csv_text=csv_text,
        header_row_index=header_row_index,
        data_start_row=data_start_row,
    )


def _format_cell_value(value: object) -> str:
    """셀 값을 CSV에 적합한 문자열로 변환한다.

    Args:
        value: openpyxl 셀 값

    Returns:
        문자열 표현
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value)


def _build_csv_text(headers: list[str], rows: list[list[str]]) -> str:
    """헤더와 데이터 행으로 CSV 텍스트를 생성한다.

    Args:
        headers: 헤더 문자열 목록
        rows: 데이터 행 목록

    Returns:
        CSV 형식 텍스트
    """
    output = io.StringIO()
    writer = csv.writer(output)
    if headers:
        writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


def _compute_file_hash(file_data: bytes) -> str:
    """파일 바이너리의 SHA-256 해시를 계산한다."""
    return hashlib.sha256(file_data).hexdigest()


async def excel_to_csv_cached(
    file_data: bytes,
    cache_manager: Any | None = None,
    sheet_name: str | None = None,
) -> dict[str, CsvSheetData]:
    """Redis 캐시를 활용하는 Excel -> CSV 변환.

    조회 순서:
    1. Redis 캐시 조회 (file_hash 키)
    2. 캐시 미스 시 CSV 변환 수행
    3. 변환 결과를 Redis에 저장

    Redis 미사용/장애 시 변환 수행 후 캐시 저장 스킵 (graceful fallback).

    Args:
        file_data: Excel 파일 바이너리
        cache_manager: SchemaCacheManager 인스턴스 (None이면 캐시 없이 변환)
        sheet_name: 특정 시트만 변환 (None이면 전체)

    Returns:
        {"시트명": CsvSheetData, ...}
    """
    file_hash = _compute_file_hash(file_data)

    # 1. Redis 캐시 조회
    if cache_manager and cache_manager.redis_available:
        try:
            cached = await cache_manager._redis_cache.load_csv_cache(file_hash)
            if cached:
                logger.info("CSV 캐시 히트 (Redis, hash=%s...)", file_hash[:12])
                result = {
                    k: CsvSheetData(**v) for k, v in cached.items()
                }
                if sheet_name:
                    return {k: v for k, v in result.items() if k == sheet_name}
                return result
        except Exception as e:
            logger.debug("Redis CSV 캐시 조회 실패, fallback 변환: %s", e)

    # 2. Fallback: CSV 변환 수행
    result = excel_to_csv(file_data, sheet_name)

    # 3. Redis에 저장 (비동기, 실패 무시)
    if cache_manager and cache_manager.redis_available:
        try:
            serializable = {k: asdict(v) for k, v in result.items()}
            await cache_manager._redis_cache.save_csv_cache(file_hash, serializable)
        except Exception as e:
            logger.debug("Redis CSV 캐시 저장 실패: %s", e)

    return result
